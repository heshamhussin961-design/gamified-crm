"""
🎮 AlSaeb CRM - Flask API Backend
=============================================
API endpoints للتعامل مع النظام

Setup:
  pip install flask flask-cors supabase python-dotenv

Run:
  python app.py
"""

import contextlib
import csv
import hashlib
import hmac
import io
import json
import logging
import math
import os
import re
import smtplib
import threading
import time
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

from flask import Flask, Response, jsonify, render_template, request, send_file, send_from_directory
from flask_cors import CORS

# ---- Logging -----------------------------------------------------------------
# Structured logging — set LOG_LEVEL=DEBUG for verbose, LOG_FORMAT=json for ECS-friendly.
logging.basicConfig(
    level=os.getenv('LOG_LEVEL', 'INFO').upper(),
    format='%(asctime)s %(levelname)s [%(name)s] %(message)s',
)
logger = logging.getLogger('alsaeb')

try:
    from supabase import Client, create_client
except ImportError:
    logger.error('supabase package missing — run: pip install -r requirements.txt')
    raise

from dotenv import load_dotenv

load_dotenv()

# ---- Sentry (optional) -------------------------------------------------------
# Activates only if SENTRY_DSN env var is present. Safe to leave unset.
SENTRY_DSN = os.getenv('SENTRY_DSN', '').strip()
_sentry_enabled = False
if SENTRY_DSN:
    try:
        import sentry_sdk
        from sentry_sdk.integrations.flask import FlaskIntegration
        sentry_sdk.init(
            dsn=SENTRY_DSN,
            integrations=[FlaskIntegration()],
            traces_sample_rate=float(os.getenv('SENTRY_TRACES_SAMPLE_RATE', '0.1')),
            environment=os.getenv('FLASK_ENV', 'production'),
            release=os.getenv('SENTRY_RELEASE') or os.getenv('GIT_SHA'),
            send_default_pii=False,
        )
        _sentry_enabled = True
        logger.info('Sentry enabled')
    except ImportError:
        logger.warning('SENTRY_DSN set but sentry-sdk not installed — skipping')
    except Exception as e:
        logger.error('Sentry init failed: %s', e)

import ai_client  # noqa: E402
import whatsapp_client  # noqa: E402

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max upload size


# ==================== API Versioning ====================
# Strategy: routes are defined once at /api/<path>. A WSGI middleware rewrites
# /api/v1/<path> → /api/<path> BEFORE Flask routing runs (before_request is too late).
# Unversioned /api/<path> still works but gets a Deprecation header to nudge clients
# toward /api/v1/.
class _APIVersionMiddleware:
    def __init__(self, wsgi_app):
        self._app = wsgi_app

    def __call__(self, environ, start_response):
        path = environ.get('PATH_INFO', '')
        if path.startswith('/api/v1/'):
            environ['PATH_INFO'] = '/api/' + path[len('/api/v1/'):]
            environ['ALSAEB_API_VERSION'] = '1'
            environ['ALSAEB_API_VERSIONED'] = '1'
        return self._app(environ, start_response)


app.wsgi_app = _APIVersionMiddleware(app.wsgi_app)
CURRENT_API_VERSION = '1'

ALLOWED_ORIGINS = [
    'https://al-saeb-crm.online',
    'https://www.al-saeb-crm.online',
    'http://localhost:5000',
    'http://127.0.0.1:5000',
]
CORS(app, origins=ALLOWED_ORIGINS, supports_credentials=True)

# ==================== Supabase Client ====================

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY')  # Service key للباك إند
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ==================== GPS / Geofence Config ====================
OFFICE_LAT = float(os.getenv('OFFICE_LAT', '0'))
OFFICE_LNG = float(os.getenv('OFFICE_LNG', '0'))
OFFICE_RADIUS_M = int(os.getenv('OFFICE_RADIUS_METERS', '200'))
REQUIRE_GPS_CHECKIN = os.getenv('REQUIRE_GPS_CHECKIN', 'false').lower() == 'true'

# ==================== SMTP Email Config ====================
SMTP_HOST = os.getenv('SMTP_HOST', '')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SMTP_USER = os.getenv('SMTP_USER', '')
SMTP_PASS = os.getenv('SMTP_PASS', '')
SMTP_FROM = os.getenv('SMTP_FROM', SMTP_USER)


# ─── Haversine distance (metres) between two lat/lng points ───
def haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ─── Send email (fire-and-forget, silently fails if SMTP not configured) ───
def send_email(to: str, subject: str, body_html: str) -> None:
    if not SMTP_HOST or not SMTP_USER:
        return

    def _send():
        with contextlib.suppress(Exception):
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = SMTP_FROM
            msg['To'] = to
            msg.attach(MIMEText(body_html, 'html', 'utf-8'))
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as srv:
                srv.ehlo()
                srv.starttls()
                srv.login(SMTP_USER, SMTP_PASS)
                srv.sendmail(SMTP_FROM, [to], msg.as_string())

    threading.Thread(target=_send, daemon=True).start()


# ─── Fire outbound webhooks for an event (fire-and-forget) ───
def fire_webhooks(event: str, payload: dict) -> None:
    def _fire():
        with contextlib.suppress(Exception):
            import requests as req_lib
            rows = supabase.table('webhooks').select('url,secret').eq(
                'is_active', True
            ).execute().data or []
            for wh in rows:
                events_row = wh.get('events') or []
                if event not in events_row and '*' not in events_row:
                    continue
                body = json.dumps({'event': event, 'data': payload, 'ts': datetime.utcnow().isoformat()})
                headers = {'Content-Type': 'application/json'}
                if wh.get('secret'):
                    sig = hmac.new(wh['secret'].encode(), body.encode(), hashlib.sha256).hexdigest()
                    headers['X-AlSaeb-Signature'] = f'sha256={sig}'
                req_lib.post(wh['url'], data=body, headers=headers, timeout=5)

    threading.Thread(target=_fire, daemon=True).start()


# ==================== CSRF Protection ====================
# Require X-Requested-With header on state-changing requests (POST/PUT/PATCH/DELETE)
# Browsers block cross-origin custom headers without CORS preflight approval
CSRF_SAFE_METHODS = {'GET', 'HEAD', 'OPTIONS'}
CSRF_EXEMPT_PATHS = {'/api/webhooks/whatsapp', '/webhook/facebook', '/webhook/lead/',
                     '/api/billing/webhooks/stripe', '/api/health'}

@app.before_request
def csrf_check():
    if request.method in CSRF_SAFE_METHODS:
        return
    path = request.path
    if any(path.startswith(p) for p in CSRF_EXEMPT_PATHS):
        return
    # API requests must have Authorization header (JWT) OR X-Requested-With
    if not request.headers.get('Authorization') and not request.headers.get('X-Requested-With'):
        return jsonify({'status': 'error', 'message': 'CSRF check failed'}), 403


# ==================== Security Headers ====================
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # API version negotiation hint — flagged so clients can migrate to /api/v1/.
    path = request.environ.get('PATH_INFO', '')
    if path.startswith('/api/'):
        response.headers['X-API-Version'] = CURRENT_API_VERSION
        if not request.environ.get('ALSAEB_API_VERSIONED'):
            response.headers['Deprecation'] = 'true'
            response.headers['Link'] = '</api/v1/>; rel="alternate"'
    return response


# ==================== PUBLIC ROUTES ====================

@app.route('/api/health')
def api_health():
    return jsonify({
        'message': 'Welcome to AlSaeb CRM API',
        'status': 'online',
        'version': '1.0.1 (Enterprise Build)',
        'admin_panel': '/admin',
    })


@app.route('/')
def landing():
    return '''<!doctype html>
<html lang="ar" dir="rtl">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>AL SAEB CRM</title>
<link rel="icon" type="image/png" href="/static/logo.png" />
<meta name="theme-color" content="#D4A847" />
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&family=IBM+Plex+Sans+Arabic:wght@400;600;700&display=swap" rel="stylesheet"/>
<style>
*{margin:0;padding:0;box-sizing:border-box}
:root{--bg:#0a0a0a;--card:#1a1a1a;--text:#e8e8e8;--muted:#888;--footer:#444;--bg-glow1:rgba(212,168,71,0.05);--bg-glow2:rgba(244,201,98,0.03);--card-border:rgba(212,168,71,0.15);--card-border-hover:rgba(212,168,71,0.45)}
[data-theme="light"]{--bg:#f5f5f5;--card:#ffffff;--text:#1a1a1a;--muted:#666;--footer:#999;--bg-glow1:rgba(212,168,71,0.1);--bg-glow2:rgba(244,201,98,0.06);--card-border:rgba(212,168,71,0.18);--card-border-hover:rgba(212,168,71,0.5)}
body{font-family:'IBM Plex Sans Arabic','Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:1.5rem;background-image:radial-gradient(ellipse at 20% 80%, var(--bg-glow1) 0%, transparent 50%),radial-gradient(ellipse at 80% 20%, var(--bg-glow2) 0%, transparent 50%);transition:background-color .3s,color .3s}
.logo-wrap{width:120px;height:120px;border-radius:24px;background:rgba(255,255,255,0.04);border:1px solid rgba(212,168,71,0.2);display:flex;align-items:center;justify-content:center;margin:0 auto 1.5rem;padding:12px;box-shadow:0 8px 32px rgba(212,168,71,0.15)}
.logo-wrap img{width:100%;height:100%;object-fit:contain;filter:drop-shadow(0 0 12px rgba(212,168,71,0.3))}
h1{font-family:'Inter','IBM Plex Sans Arabic',sans-serif;font-size:2.4rem;font-weight:800;background:linear-gradient(135deg,#F4C962,#D4A847,#A88128);-webkit-background-clip:text;background-clip:text;color:transparent;letter-spacing:3px;margin-bottom:.25rem}
.sub{color:var(--muted);font-size:.9rem;margin-bottom:2.5rem}
.cards{display:grid;grid-template-columns:repeat(2,1fr);gap:1rem;max-width:480px;width:100%}
@media(max-width:480px){.cards{grid-template-columns:1fr}}
.card{background:var(--card);border:1px solid var(--card-border);border-radius:14px;padding:1.5rem;text-align:center;text-decoration:none;color:var(--text);transition:all .25s}
.card:hover{border-color:var(--card-border-hover);box-shadow:0 4px 24px rgba(212,168,71,0.15);transform:translateY(-3px)}
.card .icon{font-size:2rem;margin-bottom:.75rem}
.card .title{font-family:'Inter',sans-serif;font-weight:700;font-size:1rem;margin-bottom:.25rem;color:#D4A847}
.card .desc{font-size:.75rem;color:var(--muted)}
.footer{margin-top:3rem;color:var(--footer);font-size:.75rem}
.footer span{color:#D4A847}
.theme-btn{position:fixed;top:1rem;left:1rem;background:rgba(212,168,71,0.12);border:1px solid rgba(212,168,71,0.3);color:#D4A847;width:44px;height:44px;border-radius:12px;font-size:18px;cursor:pointer;backdrop-filter:blur(10px);transition:transform .15s}
.theme-btn:hover{transform:scale(1.08)}
</style>
<script>
(function(){var s=localStorage.getItem('alsaeb_theme')||'dark';document.documentElement.setAttribute('data-theme',s);})();
function toggleTheme(){var c=document.documentElement.getAttribute('data-theme')||'dark';var n=c==='dark'?'light':'dark';document.documentElement.setAttribute('data-theme',n);localStorage.setItem('alsaeb_theme',n);var b=document.getElementById('theme-btn');if(b)b.textContent=n==='dark'?'🌙':'☀️';var m=document.querySelector('meta[name="theme-color"]');if(m)m.setAttribute('content',n==='dark'?'#0a0a0a':'#ffffff');}
document.addEventListener('DOMContentLoaded',function(){var t=document.documentElement.getAttribute('data-theme')||'dark';var b=document.getElementById('theme-btn');if(b)b.textContent=t==='dark'?'🌙':'☀️';});
</script>
</head>
<body>
<button id="theme-btn" class="theme-btn" onclick="toggleTheme()" title="تبديل المظهر">🌙</button>
<div class="logo-wrap"><img src="/static/logo.png" alt="AL SAEB"/></div>
<h1>AL SAEB</h1>
<p class="sub">نظام إدارة العملاء الذكي</p>
<div class="cards">
  <a href="/admin" class="card"><div class="icon">📊</div><div class="title">لوحة التحكم</div><div class="desc">إدارة الفريق والليدز والإعلانات</div></a>
  <a href="/agent" class="card"><div class="icon">🎯</div><div class="title">صفحة الموظف</div><div class="desc">متابعة الليدز والأداء</div></a>
  <a href="/game3d" class="card"><div class="icon">🎮</div><div class="title">العالم ثلاثي الأبعاد</div><div class="desc">المكتب 3D التفاعلي</div></a>
  <a href="/map" class="card"><div class="icon">🗺️</div><div class="title">الخريطة</div><div class="desc">رحلة الإنجاز</div></a>
</div>
<p class="footer">AL SAEB CRM <span>v1.1.0</span> — Gold Edition</p>
</body>
</html>'''


@app.route('/admin')
def admin_panel():
    """Serve the single-page admin panel."""
    return render_template('admin.html')


@app.route('/agent')
def agent_app():
    """Serve the mobile PWA for sales agents."""
    return render_template('agent.html')


@app.route('/map')
def journey_map():
    """Gamified RPG journey map — fog-of-war progression path."""
    return render_template('map.html')


@app.route('/game3d')
@app.route('/game')
def office_game_3d():
    """AlSaeb 3D Game — Three.js immersive CRM experience."""
    return render_template('game3d.html')


@app.route('/world')
def open_world():
    """AL SAEB World — new open-world city engine (Phase 0 foundation)."""
    return render_template('world.html')


@app.route('/favicon.ico')
def favicon():
    # Use brand logo as favicon (browsers accept PNG via .ico route)
    import os as _os
    if _os.path.exists(_os.path.join(app.static_folder, 'favicon.ico')):
        return send_from_directory('static', 'favicon.ico', mimetype='image/x-icon')
    return send_from_directory('static', 'logo.png', mimetype='image/png')


@app.route('/api/config')
def public_config():
    """
    Public config لاستخدام الـ client-side Supabase SDK.
    الـ anon key آمن للـ browser (بيحترم RLS policies).
    """
    return jsonify({
        'supabase_url':      os.getenv('SUPABASE_URL'),
        'supabase_anon_key': os.getenv('SUPABASE_ANON_KEY'),
        'vapid_public_key':  os.getenv('VAPID_PUBLIC_KEY', ''),
    })

@app.route('/api/debug/stats')
def debug_stats():
    """Temporary public endpoint to verify the 602 leads ingestion."""
    try:
        counts = supabase.rpc('get_pipeline_stats').execute()
        return jsonify({
            'status': 'debug_success',
            'pipeline': counts.data,
            'note': 'This is a public debug endpoint. Remove in production.'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Auth & RBAC Middleware ====================

# ──────────────── Auth + Rate Limit in-memory caches (Perf) ────────────────
# JWT token verification cache: avoids a Supabase HTTP call on every request.
# TTL is short (5 min) so revoked sessions don't linger long.
_token_cache = {}             # token_str -> (user_id, user_obj, expires_at)
_token_cache_lock = threading.Lock()
_TOKEN_CACHE_TTL = 300        # 5 minutes
_TOKEN_CACHE_MAX = 1000

def _verify_token_cached(token):
    """Verify Supabase JWT — uses local cache to avoid network calls per request."""
    if not token:
        return None
    now_ts = time.time()
    with _token_cache_lock:
        entry = _token_cache.get(token)
        if entry and entry[2] > now_ts:
            return entry[0], entry[1]
    try:
        user = supabase.auth.get_user(token)
        result = (user.user.id, user.user)
    except Exception:
        return None
    with _token_cache_lock:
        _token_cache[token] = (result[0], result[1], now_ts + _TOKEN_CACHE_TTL)
        if len(_token_cache) > _TOKEN_CACHE_MAX:
            # Drop expired entries
            stale = [t for t, e in list(_token_cache.items()) if e[2] < now_ts]
            for t in stale:
                _token_cache.pop(t, None)
    return result


# Rate-limit cache: in-memory sliding window (60s)
_rate_cache = {}              # (user_id, endpoint_key) -> [timestamps]
_rate_cache_lock = threading.Lock()

# Per-user response cache (for frequently-polled GET endpoints like /api/me, leaderboard)
_resp_cache = {}              # key -> (payload, expires_at)
_resp_cache_lock = threading.Lock()


def cached_response(ttl=15):
    """Caches the JSON response of a GET endpoint per user for `ttl` seconds.
    Use only for safe (idempotent) endpoints whose data tolerates a small staleness."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            uid = getattr(request, 'user_id', None) or 'anon'
            cache_key = (uid, request.path, request.query_string.decode('utf-8', 'ignore'))
            now_ts = time.time()
            with _resp_cache_lock:
                entry = _resp_cache.get(cache_key)
                if entry and entry[1] > now_ts:
                    cached_payload, _ = entry
                    # Return a fresh Response so headers can still be modified
                    resp = jsonify(cached_payload[0])
                    resp.status_code = cached_payload[1]
                    return resp
            result = f(*args, **kwargs)
            # Extract data + status from Flask response
            try:
                if isinstance(result, tuple):
                    body_resp, status = result
                else:
                    body_resp, status = result, 200
                payload_json = body_resp.get_json() if hasattr(body_resp, 'get_json') else None
                if payload_json is not None and status < 400:
                    with _resp_cache_lock:
                        _resp_cache[cache_key] = ((payload_json, status), now_ts + ttl)
                        if len(_resp_cache) > 2000:
                            stale = [k for k, e in list(_resp_cache.items()) if e[1] < now_ts]
                            for k in stale:
                                _resp_cache.pop(k, None)
            except Exception:
                pass
            return result
        return decorated
    return decorator


def invalidate_user_cache(user_id):
    """Drop cached responses for a user (call after mutations)."""
    if not user_id:
        return
    with _resp_cache_lock:
        for k in list(_resp_cache.keys()):
            if k[0] == user_id:
                _resp_cache.pop(k, None)


# Lock-check cache so we don't query the DB on every request
_locked_cache = {}  # user_id -> (is_locked, expires_at)
_locked_cache_lock = threading.Lock()
_LOCKED_CACHE_TTL = 30  # 30s — short so unlock takes effect quickly


def _is_user_locked(user_id):
    if not user_id:
        return False
    now_ts = time.time()
    with _locked_cache_lock:
        entry = _locked_cache.get(user_id)
        if entry and entry[1] > now_ts:
            return entry[0]
    try:
        r = supabase.table('employees').select('is_locked').eq('id', user_id).single().execute()
        locked = bool((r.data or {}).get('is_locked'))
    except Exception:
        locked = False
    with _locked_cache_lock:
        _locked_cache[user_id] = (locked, now_ts + _LOCKED_CACHE_TTL)
    return locked


def _invalidate_lock_cache(user_id):
    with _locked_cache_lock:
        _locked_cache.pop(user_id, None)


def require_auth(f):
    """Verifies the JWT token and attaches user to request context (cached)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return error_response('No token provided', 401)
        result = _verify_token_cached(token)
        if not result:
            return error_response('انتهت صلاحية الجلسة — سجل دخول مرة تانية', 401)
        # Block locked employees from accessing anything except their /api/me
        if _is_user_locked(result[0]):
            return error_response('حسابك مقفول — تواصل مع الأدمن لفتحه', 403)
        request.user_id = result[0]
        request.user = result[1]
        # Update last_seen_at (cheap — only every ~60s to avoid DB hammering)
        global _last_seen_throttle
        try:
            now_ts = time.time()
            last = _last_seen_throttle.get(result[0], 0)
            if now_ts - last > 60:
                _last_seen_throttle[result[0]] = now_ts
                # Fire and forget — non-blocking
                threading.Thread(target=_update_last_seen, args=(result[0],), daemon=True).start()
        except Exception:
            pass
        return f(*args, **kwargs)
    return decorated


_last_seen_throttle = {}

def _update_last_seen(user_id):
    with contextlib.suppress(Exception):
        supabase.table('employees').update({
            'last_seen_at': datetime.utcnow().isoformat()
        }).eq('id', user_id).execute()


def check_role(roles):
    """Enforces specific roles for access control."""
    def decorator(f):
        @wraps(f)
        @require_auth
        def decorated(*args, **kwargs):
            emp = supabase.table('employees').select('role').eq('id', request.user_id).single().execute()
            if not emp.data or emp.data['role'] not in roles:
                return error_response(f'Access denied. Required roles: {roles}', 403)
            return f(*args, **kwargs)
        return decorated
    return decorator


def audit_log(action_name, target_type=None):
    """Decorator يسجل أكشنز الأدمن في admin_audit_log. يتطبق بعد @check_role."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            response = f(*args, **kwargs)
            try:
                target_id = kwargs.get('project_id') or kwargs.get('team_id') \
                    or kwargs.get('comp_id') or kwargs.get('lead_id') \
                    or kwargs.get('user_id')
                supabase.table('admin_audit_log').insert({
                    'admin_id':    getattr(request, 'user_id', None),
                    'action':      action_name,
                    'target_type': target_type,
                    'target_id':   target_id,
                    'ip_address':  request.headers.get('X-Forwarded-For', request.remote_addr),
                    'user_agent':  request.headers.get('User-Agent', '')[:500],
                    'details': {
                        'method': request.method,
                        'path':   request.path,
                        'body':   (request.get_json(silent=True) or {}),
                    },
                }).execute()
            except Exception:
                pass  # audit failure shouldn't break the request
            return response
        return decorated
    return decorator


# Rate-limit backend selection — 'memory' (per-worker, fast, dev) or 'db' (shared, prod-safe).
# In multi-worker deployments (gunicorn -w N) memory backend allows up to N× the configured
# rate before throttling kicks in. Set RATE_LIMIT_BACKEND=db in production.
RATE_LIMIT_BACKEND = os.getenv('RATE_LIMIT_BACKEND', 'memory').lower()

# Negative cache: if a user/endpoint was throttled in the last 30s, refuse without DB lookup.
_rate_throttle_cache = {}        # (uid, endpoint_key) -> retry_at_ts
_rate_throttle_lock  = threading.Lock()
_RATE_THROTTLE_TTL   = 30


def _rate_limit_db(uid, endpoint_key, max_per_min):
    """Authoritative DB-backed rate-limit via check_rate_limit RPC.
    Returns True if the request is allowed, False if it should be 429'd.
    Fails OPEN (returns True) on DB error to avoid taking the app down with the DB.
    """
    try:
        res = supabase.rpc('check_rate_limit', {
            'p_employee_id': uid,
            'p_endpoint':    endpoint_key,
            'p_max_per_min': max_per_min,
        }).execute()
        return bool(res.data)
    except Exception as e:
        logger.warning('rate_limit DB check failed (fail-open): %s', e)
        return True


def rate_limit(endpoint_key, max_per_min=60):
    """Sliding-window rate limiter.
    Backend chosen by RATE_LIMIT_BACKEND env var:
      - 'memory' (default): per-worker in-memory; fast but lets up to N× through
                            on a multi-worker deployment.
      - 'db':               authoritative across workers via check_rate_limit RPC.
    """
    def decorator(f):
        @wraps(f)
        @require_auth
        def decorated(*args, **kwargs):
            uid = getattr(request, 'user_id', None)
            if not uid:
                return f(*args, **kwargs)

            if RATE_LIMIT_BACKEND == 'db':
                # Negative-cache: skip DB if recently throttled
                cache_key = (uid, endpoint_key)
                now_ts = time.time()
                with _rate_throttle_lock:
                    retry_at = _rate_throttle_cache.get(cache_key)
                if retry_at and retry_at > now_ts:
                    return error_response('Rate limit exceeded', 429)
                allowed = _rate_limit_db(uid, endpoint_key, max_per_min)
                if not allowed:
                    with _rate_throttle_lock:
                        _rate_throttle_cache[cache_key] = now_ts + _RATE_THROTTLE_TTL
                        # Light GC
                        if len(_rate_throttle_cache) > 5000:
                            for k, v in list(_rate_throttle_cache.items()):
                                if v < now_ts:
                                    _rate_throttle_cache.pop(k, None)
                    return error_response('Rate limit exceeded', 429)
                return f(*args, **kwargs)

            # Default: per-worker in-memory sliding window (no DB call per request).
            now_ts = time.time()
            key = (uid, endpoint_key)
            with _rate_cache_lock:
                hits = _rate_cache.get(key, [])
                hits = [t for t in hits if t > now_ts - 60]
                if len(hits) >= max_per_min:
                    return error_response('Rate limit exceeded', 429)
                hits.append(now_ts)
                _rate_cache[key] = hits
                # Light cleanup
                if len(_rate_cache) > 5000:
                    cutoff = now_ts - 60
                    for k in list(_rate_cache.keys()):
                        if not _rate_cache[k] or _rate_cache[k][-1] < cutoff:
                            _rate_cache.pop(k, None)
            return f(*args, **kwargs)
        return decorated
    return decorator


# ==================== Common Response Helpers ====================

def success_response(data, message="Success", status=200):
    return jsonify({
        'status': 'success',
        'message': message,
        'data': data,
        'timestamp': datetime.utcnow().isoformat()
    }), status

def error_response(message, status=400):
    return jsonify({
        'status': 'error',
        'message': message,
        'timestamp': datetime.utcnow().isoformat()
    }), status


# ==================== XP Configuration ====================

XP_CONFIG = {
    'whatsapp_sent':     {'xp': 10,  'coins': 0},
    'whatsapp_received': {'xp': 50,  'coins': 0},
    'call_made':         {'xp': 15,  'coins': 0},
    'call_received':     {'xp': 30,  'coins': 0},
    'meeting_booked':    {'xp': 200, 'coins': 20},
    'deal_closed':       {'xp': 500, 'coins': 50},
    'note_added':        {'xp': 5,   'coins': 0},
    'lead_upgraded':     {'xp': 25,  'coins': 5},
    'ai_hint_used':      {'xp': 2,   'coins': 0},
}


def get_power_hour_multiplier():
    """Check all active Power Hours / CEO Visits and return the highest XP multiplier (or 1.0)."""
    try:
        now = datetime.utcnow().isoformat()
        res = supabase.table('power_hours').select('multiplier, type') \
            .eq('is_active', True).gte('ends_at', now).execute()
        if res.data:
            return max(float(r.get('multiplier', 1)) for r in res.data)
    except Exception:
        pass
    return 1.0


# ==================== LEADS ENDPOINTS ====================

@app.route('/api/leads', methods=['GET'])
@require_auth
def get_leads():
    """جيب الليدز — الموظف يشوف بتوعه، الأدمن والمانيجر يشوفوا الكل"""
    try:
        status = request.args.get('status')
        quality = request.args.get('quality')
        project_id = request.args.get('project_id')
        search = request.args.get('search')
        assigned_to = request.args.get('assigned_to')  # للأدمن: فلتر بموظف معين
        ad_campaign_id = request.args.get('ad_campaign_id')  # فلتر بإعلان معين
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        # تحقق من الدور — الأدمن والمانيجر يشوفوا كل الليدز
        emp = supabase.table('employees').select('role').eq('id', request.user_id).single().execute()
        is_admin = emp.data and emp.data.get('role') in ('admin', 'manager')

        query = supabase.table('leads').select(
            '*, projects(name, slug)'
        ).eq('is_active', True)

        if not is_admin:
            query = query.eq('assigned_to', request.user_id)
        elif assigned_to:
            query = query.eq('assigned_to', assigned_to)

        if status:
            query = query.eq('status', status)
        if quality:
            query = query.eq('quality', quality)
        if project_id:
            query = query.eq('project_id', project_id)
        if ad_campaign_id:
            query = query.eq('ad_campaign_id', ad_campaign_id)
        if search:
            # بحث في الاسم + التليفون + الإيميل + الملخص الـ AI
            query = query.or_(
                f"name.ilike.%{search}%,"
                f"phone.ilike.%{search}%,"
                f"email.ilike.%{search}%,"
                f"ai_summary.ilike.%{search}%"
            )

        query = query.order('created_at', desc=True)
        query = query.range((page - 1) * per_page, page * per_page - 1)

        result = query.execute()
        return success_response({
            'leads': result.data,
            'page': page,
            'per_page': per_page,
            'is_admin_view': is_admin,
        })
    except Exception as e:
        return error_response(f"Lead Retrieval Error: {str(e)}", 500)


@app.route('/api/leads/<lead_id>', methods=['GET'])
@require_auth
def get_lead(lead_id):
    """تفاصيل ليد واحد مع تاريخ التواصل"""
    try:
        lead = supabase.table('leads').select(
            '*, projects(name, slug)'
        ).eq('id', lead_id).single().execute()
    except Exception:
        try:
            lead = supabase.table('leads').select('*').eq('id', lead_id).single().execute()
        except Exception:
            return error_response('الليد غير موجود', 404)

    actions_data, messages_data = [], []
    try:
        actions = supabase.table('actions_log').select('*').eq(
            'lead_id', lead_id
        ).order('created_at', desc=True).limit(20).execute()
        actions_data = actions.data or []
    except Exception:
        pass
    try:
        messages = supabase.table('whatsapp_messages').select('*').eq(
            'lead_id', lead_id
        ).order('created_at', desc=True).limit(50).execute()
        messages_data = messages.data or []
    except Exception:
        pass

    return jsonify({
        'lead': lead.data,
        'actions': actions_data,
        'messages': messages_data
    })


@app.route('/api/leads/<lead_id>/status', methods=['PATCH'])
@require_auth
def update_lead_status(lead_id):
    """
    تحديث حالة الليد عن طريق RPC اللي بيعمل:
    - التحقق من صحة التحويل (state machine)
    - تسجيل التاريخ في lead_status_history
    """
    try:
        data = request.json or {}
        new_status = data.get('status')
        reason = data.get('reason')

        valid_statuses = {'new', 'contacted', 'interested', 'meeting_set',
                          'negotiation', 'closed_won', 'closed_lost'}
        if new_status not in valid_statuses:
            return error_response('Invalid lead status provided')

        # تحقق إن الليد مخصص للموظف
        owner = supabase.table('leads').select('assigned_to').eq(
            'id', lead_id
        ).single().execute()
        if not owner.data:
            return error_response('Lead not found', 404)
        if owner.data['assigned_to'] != request.user_id:
            return error_response('Access denied', 403)

        result = supabase.rpc('change_lead_status', {
            'p_lead_id':     lead_id,
            'p_employee_id': request.user_id,
            'p_new_status':  new_status,
            'p_reason':      reason,
        }).execute()

        payload = result.data or {}
        if payload.get('error'):
            return error_response(payload, 400)

        # سجل الأكشن في actions_log عشان الـ quests والإحصائيات
        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id':     lead_id,
            'action':      'status_changed',
            'xp_earned':   0,
            'coins_earned': 0,
            'details':     payload,
            'ref_type':    'lead_status',
        }).execute()

        return success_response(payload, 'Lead status updated successfully')
    except Exception as e:
        return error_response(f"Lead Update Error: {str(e)}", 500)


@app.route('/api/leads', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('lead_create', target_type='lead')
def create_lead():
    """إنشاء ليد جديد يدوي (مع ربط بإعلان اختياري)"""
    try:
        body = request.get_json()
        phone = (body.get('phone') or '').strip()
        if not phone:
            return error_response('رقم التليفون مطلوب', 400)

        import re
        phone_clean = re.sub(r'\D', '', phone)

        row = {
            'phone': phone,
            'phone_clean': phone_clean,
            'name': body.get('name') or None,
            'email': body.get('email') or None,
            'project_id': body.get('project_id') or None,
            'assigned_to': body.get('assigned_to') or None,
            'source': body.get('source') or 'manual',
            'quality': body.get('quality', 'unknown'),
            'ad_campaign_id': body.get('ad_campaign_id') or None,
            'utm_source': body.get('utm_source') or None,
            'utm_medium': body.get('utm_medium') or None,
            'utm_campaign': body.get('utm_campaign') or None,
        }
        result = supabase.table('leads').insert(row).execute()
        return success_response(result.data[0] if result.data else {}, 'تم إنشاء الليد ✅', 201)
    except Exception as e:
        return error_response(f"Lead Create Error: {str(e)}", 500)


@app.route('/api/leads/<lead_id>/link-ad', methods=['PATCH'])
@check_role(['admin', 'manager'])
@audit_log('lead_link_ad', target_type='lead')
def link_lead_to_ad(lead_id):
    """ربط ليد بإعلان أو فك الارتباط"""
    try:
        body = request.get_json()
        updates = {
            'ad_campaign_id': body.get('ad_campaign_id') or None,
        }
        if 'utm_source' in body:
            updates['utm_source'] = body['utm_source'] or None
        if 'utm_medium' in body:
            updates['utm_medium'] = body['utm_medium'] or None
        if 'utm_campaign' in body:
            updates['utm_campaign'] = body['utm_campaign'] or None

        result = supabase.table('leads').update(updates).eq('id', lead_id).execute()
        return success_response(result.data[0] if result.data else {}, 'تم ربط الليد بالإعلان ✅')
    except Exception as e:
        return error_response(f"Link Ad Error: {str(e)}", 500)


@app.route('/api/leads/import', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('leads_import')
def import_leads_excel():
    """استيراد ليدز من ملف Excel"""
    import openpyxl

    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        return error_response('يرجى رفع ملف Excel (.xlsx)', 400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active
        rows_list = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows_list) < 2:
            return error_response('الملف فاضي — لازم يكون فيه صف عناوين + بيانات', 400)

        header_map = {
            'الاسم': 'name', 'name': 'name', 'اسم العميل': 'name', 'client name': 'name', 'lead name': 'name',
            'الموبايل': 'phone', 'phone': 'phone', 'رقم الموبايل': 'phone', 'الهاتف': 'phone', 'mobile': 'phone', 'رقم الهاتف': 'phone',
            'الإيميل': 'email', 'email': 'email', 'البريد': 'email', 'الايميل': 'email',
            'المصدر': 'source', 'source': 'source',
            'الحالة': 'status', 'status': 'status',
            'الجودة': 'quality', 'quality': 'quality',
            'الميزانية': 'budget', 'budget': 'budget',
            'المدينة': 'city', 'city': 'city',
            'البلد': 'country', 'country': 'country',
            'ملاحظات': 'notes', 'notes': 'notes',
            'الحملة': 'campaign', 'campaign': 'campaign', 'المشروع': 'campaign',
        }

        raw_headers = [str(h).strip().lower() if h else '' for h in rows_list[0]]
        col_map = {}
        for i, h in enumerate(raw_headers):
            if h in header_map:
                col_map[i] = header_map[h]

        if 'phone' not in col_map.values() and 'name' not in col_map.values():
            return error_response('الملف لازم يحتوي على عمود "الاسم" أو "الموبايل" على الأقل', 400)

        status_map = {
            'جديد': 'new', 'new': 'new', 'اتواصل': 'contacted', 'contacted': 'contacted',
            'مهتم': 'interested', 'interested': 'interested',
            'ميتنج': 'meeting_set', 'meeting_set': 'meeting_set', 'meeting': 'meeting_set',
            'تفاوض': 'negotiation', 'negotiation': 'negotiation',
            'فاز': 'closed_won', 'closed_won': 'closed_won', 'won': 'closed_won',
            'خسارة': 'closed_lost', 'closed_lost': 'closed_lost', 'lost': 'closed_lost',
        }

        created = 0
        skipped = 0
        duplicates = 0
        errors = []

        for row_idx, row in enumerate(rows_list[1:], start=2):
            parsed = {}
            for col_idx, field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    parsed[field] = str(val).strip()

            phone = parsed.get('phone', '').strip()
            name = parsed.get('name', '').strip()
            if not phone and not name:
                skipped += 1
                continue

            # تنظيف رقم الموبايل
            phone_clean = re.sub(r'\D', '', phone) if phone else ''

            # التحقق من التكرار بالموبايل
            if phone_clean:
                existing = supabase.table('leads').select('id').eq('phone_clean', phone_clean).limit(1).execute()
                if existing.data:
                    duplicates += 1
                    continue

            raw_status = parsed.get('status', '').strip().lower()
            lead_status = status_map.get(raw_status, 'new')

            lead_row = {
                'name': name or None,
                'phone': phone or None,
                'phone_clean': phone_clean or None,
                'email': parsed.get('email') or None,
                'source': parsed.get('source') or 'excel_import',
                'status': lead_status,
                'quality': parsed.get('quality') or 'unknown',
                'budget_range': parsed.get('budget') or None,
                'city': parsed.get('city') or None,
                'country': parsed.get('country') or None,
                'notes': parsed.get('notes') or None,
                'imported_from': f'excel_{f.filename}',
                'created_by': request.user_id,
            }

            try:
                supabase.table('leads').insert(lead_row).execute()
                created += 1
            except Exception as e:
                errors.append(f'صف {row_idx}: {str(e)[:80]}')

        wb.close()
        msg = f'تم استيراد {created} ليد'
        if duplicates:
            msg += f'، {duplicates} مكرر'
        if skipped:
            msg += f'، تم تخطي {skipped} صف فاضي'
        if errors:
            msg += f'، {len(errors)} خطأ'

        return success_response({
            'created': created,
            'duplicates': duplicates,
            'skipped': skipped,
            'errors': errors[:10],
        }, msg)

    except Exception as e:
        return error_response(f'خطأ في قراءة الملف: {str(e)}', 400)


@app.route('/api/leads/template')
@require_auth
def leads_template():
    """تحميل نموذج Excel لاستيراد الليدز"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الليدز'
    ws.sheet_view.rightToLeft = True

    headers = ['الاسم', 'الموبايل', 'الإيميل', 'المصدر', 'الحالة', 'الجودة', 'الميزانية', 'المدينة', 'البلد', 'ملاحظات']
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    sample = [
        ['أحمد محمد', '0501234567', 'ahmed@email.com', 'Facebook', 'جديد', 'warm', '500K-1M', 'دبي', 'الإمارات', 'مهتم بفيلا'],
        ['سارة علي', '0559876543', 'sara@email.com', 'Instagram', 'مهتم', 'hot', '1M-2M', 'أبوظبي', 'الإمارات', ''],
    ]
    for r, data in enumerate(sample, 2):
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    # تعليمات
    ws2 = wb.create_sheet('تعليمات')
    ws2.sheet_view.rightToLeft = True
    instructions = [
        'تعليمات استيراد الليدز:',
        '',
        '1. عمود "الاسم" أو "الموبايل" مطلوب على الأقل',
        '2. الحالة: جديد، اتواصل، مهتم، ميتنج، تفاوض، فاز، خسارة',
        '3. الجودة: hot, warm, cold, unknown',
        '4. الأرقام المكررة يتم تخطيها تلقائياً',
        '5. المصدر: Facebook, Instagram, Google, WhatsApp, Website, referral, أو أي نص',
    ]
    for r, txt in enumerate(instructions, 1):
        ws2.cell(row=r, column=1, value=txt).font = Font(size=12)
    ws2.column_dimensions['A'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='leads_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/leads/bulk-assign', methods=['POST'])
@check_role(['admin'])
@audit_log('bulk_assign', target_type='lead')
def bulk_assign_leads():
    """
    توزيع الليدز بسرعة الصاروخ باستخدام Database-side logic (RPC).
    هنا بنستخدم الـ Stored Procedure اللي عملناه في Postgres.
    """
    try:
        data = request.json or {}
        project_id = data.get('project_id')
        limit = data.get('limit', 1000) # الافتراضي توزيع 1000 ليد

        # تنفيذ التوزيع بطلب واحد فقط لقاعدة البيانات
        result = supabase.rpc('distribute_leads_to_agents', {
            'p_project_id': project_id,
            'p_limit': limit
        }).execute()

        if result.data and result.data.get('error'):
            return error_response(result.data['error'], 400)

        return success_response(result.data, "Bulk distribution completed successfully")
    except Exception as e:
        return error_response(f"Bulk Assignment Error: {str(e)}", 500)


# ==================== ACTIONS ENDPOINTS ====================

@app.route('/api/actions/whatsapp', methods=['POST'])
@rate_limit('whatsapp_send', max_per_min=30)
def send_whatsapp():
    """إرسال رسالة واتساب (+ XP)"""
    data = request.json or {}
    lead_id = data.get('lead_id')
    message = data.get('message')
    template_id = data.get('template_id')

    if not lead_id:
        return jsonify({'error': 'lead_id required'}), 400

    # جيب بيانات الليد
    lead = supabase.table('leads').select('*').eq('id', lead_id).single().execute()
    if not lead.data:
        return jsonify({'error': 'Lead not found'}), 404

    # لو في قالب، جهز الرسالة
    if template_id:
        template = supabase.table('message_templates').select('*').eq(
            'id', template_id
        ).single().execute()
        if template.data:
            message = template.data['template_text']
            message = message.replace('{{name}}', lead.data.get('name', 'عميلنا الكريم'))
            # TODO: replace {{project}} with actual project name

            # حدّث عداد استخدام القالب
            supabase.table('message_templates').update({
                'times_used': template.data['times_used'] + 1
            }).eq('id', template_id).execute()

    # 📱 إرسال فعلي عبر WhatsApp Business API
    if template_id and template.data:
        wa_resp = whatsapp_client.send_template(
            lead.data['phone_clean'],
            template.data['name'],
            language=template.data.get('language', 'ar'),
        )
    else:
        wa_resp = whatsapp_client.send_text(lead.data['phone_clean'], message or '')

    wa_status = 'sent' if wa_resp.get('ok') else ('skipped' if wa_resp.get('skipped') else 'failed')

    # سجل الرسالة
    supabase.table('whatsapp_messages').insert({
        'lead_id': lead_id,
        'employee_id': request.user_id,
        'direction': 'outbound',
        'message_text': message,
        'message_type': 'template' if template_id else 'text',
        'template_name': template.data['name'] if template_id and template.data else None,
        'wa_message_id': wa_resp.get('message_id'),
        'wa_status': wa_status,
    }).execute()

    # حدّث الليد
    supabase.table('leads').update({
        'status': 'contacted' if lead.data['status'] == 'new' else lead.data['status'],
        'last_contact_at': datetime.utcnow().isoformat(),
        'contact_count': lead.data.get('contact_count', 0) + 1,
        'updated_at': datetime.utcnow().isoformat()
    }).eq('id', lead_id).execute()

    # 🎮 أدي XP (with Power Hour multiplier)
    xp_config = XP_CONFIG['whatsapp_sent']
    ph_mult = get_power_hour_multiplier()
    final_xp = int(xp_config['xp'] * ph_mult)
    final_coins = int(xp_config['coins'] * ph_mult)
    xp_result = supabase.rpc('award_xp_and_coins', {
        'p_employee_id': request.user_id,
        'p_xp': final_xp,
        'p_coins': final_coins,
        'p_reason': 'إرسال رسالة واتساب',
        'p_ref_type': 'action',
    }).execute()

    # سجل الأكشن
    supabase.table('actions_log').insert({
        'employee_id': request.user_id,
        'lead_id': lead_id,
        'action': 'whatsapp_sent',
        'xp_earned': final_xp,
        'coins_earned': final_coins,
        'details': {'message_preview': message[:100] if message else None}
    }).execute()

    # حدّث الإحصائيات (atomic — يمنع race condition بين request متوازيين)
    supabase.rpc('increment_employee_counter', {
        'p_employee_id': request.user_id,
        'p_counter':     'total_messages',
        'p_delta':       1,
    }).execute()

    return jsonify({
        'success': True,
        'xp': xp_result.data if xp_result.data else xp_config
    })


@app.route('/api/actions/call', methods=['POST'])
@rate_limit('call_log', max_per_min=60)
def log_call():
    """تسجيل مكالمة (+ XP)"""
    data = request.json or {}
    lead_id = data.get('lead_id')
    duration = data.get('duration_seconds', 0)
    notes = data.get('notes')
    outcome = data.get('outcome')  # answered, no_answer, voicemail, callback

    # 🎮 XP حسب نتيجة المكالمة (with Power Hour multiplier)
    xp_config = XP_CONFIG['call_made'].copy()
    if outcome == 'answered' and duration > 60:
        xp_config['xp'] *= 2  # مكالمة طويلة = ضعف الـ XP
    ph_mult = get_power_hour_multiplier()
    xp_config['xp'] = int(xp_config['xp'] * ph_mult)
    xp_config['coins'] = int(xp_config['coins'] * ph_mult)

    try:
        # سجل الأكشن
        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id': lead_id,
            'action': 'call_made',
            'xp_earned': xp_config['xp'],
            'details': {
                'duration_seconds': duration,
                'outcome': outcome,
                'notes': notes
            }
        }).execute()

        # أدي XP
        supabase.rpc('award_xp_and_coins', {
            'p_employee_id': request.user_id,
            'p_xp': xp_config['xp'],
            'p_coins': xp_config['coins'],
            'p_reason': f'مكالمة - {outcome}',
            'p_ref_type': 'action',
        }).execute()

        # atomic counter bump
        supabase.rpc('increment_employee_counter', {
            'p_employee_id': request.user_id,
            'p_counter':     'total_calls',
            'p_delta':       1,
        }).execute()
    except Exception:
        pass

    return jsonify({'success': True, 'xp_earned': xp_config['xp']})


@app.route('/api/actions/close-deal', methods=['POST'])
@rate_limit('close_deal', max_per_min=20)
def close_deal():
    """إغلاق صفقة 🎉 (+ XP كبير + عملات)"""
    data = request.json or {}
    lead_id = data.get('lead_id')
    deal_value = data.get('deal_value', 0)
    notes = data.get('notes')

    # حدّث حالة الليد (لو في lead_id)
    if lead_id:
        supabase.table('leads').update({
            'status': 'closed_won',
            'updated_at': datetime.utcnow().isoformat()
        }).eq('id', lead_id).execute()

    # 🎮 XP + Coins كبيرة! (with Power Hour multiplier)
    xp_config = XP_CONFIG['deal_closed']
    ph_mult = get_power_hour_multiplier()

    # بونص حسب قيمة الصفقة
    bonus_multiplier = 1.0
    if deal_value > 1000000:   # أكتر من مليون
        bonus_multiplier = 3.0
    elif deal_value > 500000:
        bonus_multiplier = 2.0
    elif deal_value > 100000:
        bonus_multiplier = 1.5

    final_xp = int(xp_config['xp'] * bonus_multiplier * ph_mult)
    final_coins = int(xp_config['coins'] * bonus_multiplier * ph_mult)

    try:
        # أدي المكافأة
        xp_result = supabase.rpc('award_xp_and_coins', {
            'p_employee_id': request.user_id,
            'p_xp': final_xp,
            'p_coins': final_coins,
            'p_reason': f'إغلاق صفقة بقيمة {deal_value}',
            'p_ref_type': 'action',
        }).execute()

        # سجل الأكشن
        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id': lead_id,
            'action': 'deal_closed',
            'xp_earned': final_xp,
            'coins_earned': final_coins,
            'details': {
                'deal_value': deal_value,
                'bonus_multiplier': bonus_multiplier,
                'notes': notes
            }
        }).execute()

        # حدّث إحصائيات الموظف (atomic)
        supabase.rpc('increment_employee_counter', {
            'p_employee_id': request.user_id,
            'p_counter':     'total_deals',
            'p_delta':       1,
        }).execute()

        leveled_up = xp_result.data.get('leveled_up', False) if xp_result.data else False
    except Exception:
        leveled_up = False

    # 🔔 Webhook + Email trigger
    with contextlib.suppress(Exception):
        emp = supabase.table('employees').select('full_name,email').eq('id', request.user_id).single().execute().data or {}
        fire_webhooks('deal_closed', {
            'employee_id': request.user_id,
            'employee_name': emp.get('full_name'),
            'lead_id': lead_id,
            'deal_value': deal_value,
            'xp_earned': final_xp,
        })
        if emp.get('email'):
            send_email(
                emp['email'],
                '🏆 صفقة جديدة مغلقة!',
                f'<h2>مبروك يا {emp.get("full_name", "")}! 🎉</h2>'
                f'<p>تم إغلاق صفقة بقيمة <strong>{deal_value:,}</strong> بنجاح.</p>'
                f'<p>كسبت <strong>{final_xp} XP</strong> و <strong>{final_coins} عملة</strong>.</p>',
            )

    return jsonify({
        'success': True,
        'xp_earned': final_xp,
        'coins_earned': final_coins,
        'bonus_multiplier': bonus_multiplier,
        'leveled_up': leveled_up
    })


# ==================== QUESTS ENDPOINTS ====================

@app.route('/api/quests', methods=['GET'])
@require_auth
def get_quests():
    """جيب المهمات الحالية"""
    status = request.args.get('status', 'pending')
    try:
        result = supabase.table('quests').select(
            '*, leads(name, phone, projects(name))'
        ).eq('employee_id', request.user_id).eq('status', status).order(
            'priority', desc=True
        ).order('due_at').execute()
        return jsonify({'quests': result.data})
    except Exception:
        try:
            result = supabase.table('quests').select('*').eq('employee_id', request.user_id).execute()
            return jsonify({'quests': result.data or []})
        except Exception:
            return jsonify({'quests': []})


@app.route('/api/quests/<quest_id>/complete', methods=['POST'])
@require_auth
def complete_quest(quest_id):
    """إكمال مهمة"""
    try:
        quest = supabase.table('quests').select('*').eq(
            'id', quest_id
        ).eq('employee_id', request.user_id).single().execute()
    except Exception:
        return error_response('المهمة غير موجودة', 404)

    if not quest.data:
        return jsonify({'error': 'Quest not found'}), 404

    if quest.data['status'] == 'completed':
        return jsonify({'error': 'Quest already completed'}), 400

    # حدّث المهمة
    supabase.table('quests').update({
        'status': 'completed',
        'completed_at': datetime.utcnow().isoformat()
    }).eq('id', quest_id).execute()

    # 🎮 أدي المكافأة
    xp = quest.data['xp_reward']
    coins = quest.data['coin_reward']
    multiplier = float(quest.data.get('bonus_multiplier', 1.0))

    final_xp = int(xp * multiplier)
    final_coins = int(coins * multiplier)

    supabase.rpc('award_xp_and_coins', {
        'p_employee_id': request.user_id,
        'p_xp': final_xp,
        'p_coins': final_coins,
        'p_reason': f'إكمال مهمة: {quest.data["title"]}',
        'p_ref_type': 'quest',
        'p_ref_id': quest_id,
    }).execute()

    return jsonify({
        'success': True,
        'xp_earned': final_xp,
        'coins_earned': final_coins
    })


# ==================== LEADERBOARD ====================

@app.route('/api/leaderboard', methods=['GET'])
@require_auth
@cached_response(ttl=30)
def get_leaderboard():
    """لوحة الشرف (مخزن مؤقتاً 30 ثانية)"""
    period = request.args.get('period', 'weekly')

    try:
        result = supabase.table('leaderboard').select(
            '*, employees(full_name, avatar_url, level, team)'
        ).order('rank').limit(20).execute()
        board = result.data or []
    except Exception:
        # Fallback: build leaderboard from employees table
        emps = supabase.table('employees').select(
            'id, full_name, avatar_url, level, team, total_xp'
        ).eq('is_active', True).order('total_xp', desc=True).limit(20).execute()
        board = []
        for i, e in enumerate(emps.data or [], 1):
            board.append({
                'rank': i,
                'employee_id': e['id'],
                'score': e.get('total_xp', 0),
                'total_xp': e.get('total_xp', 0),
                'deals_closed': e.get('total_deals', 0),
                'employees': {
                    'full_name': e['full_name'],
                    'avatar_url': e.get('avatar_url'),
                    'level': e.get('level', 1),
                    'team': e.get('team')
                }
            })

    return jsonify({'leaderboard': board, 'period': period})


# ==================== STAMINA ====================

@app.route('/api/stamina', methods=['GET'])
@require_auth
def get_stamina():
    """Get current stamina for the logged-in user."""
    try:
        res = supabase.table('employees').select('stamina, max_stamina').eq('id', request.user_id).single().execute()
        return success_response({'stamina': res.data.get('stamina', 100), 'max_stamina': res.data.get('max_stamina', 100)})
    except Exception:
        return success_response({'stamina': 100, 'max_stamina': 100})


@app.route('/api/stamina/drain', methods=['POST'])
@require_auth
def drain_stamina():
    """Drain stamina when performing CRM actions."""
    data = request.get_json(silent=True) or {}
    amount = int(data.get('amount', 5))
    try:
        res = supabase.rpc('drain_stamina', {'p_employee_id': request.user_id, 'p_amount': amount}).execute()
        result = res.data
        if isinstance(result, list):
            result = result[0] if result else {}
        if not result.get('ok'):
            return error_response(result.get('error', 'stamina drain failed'), 400)
        return success_response(result)
    except Exception:
        # RPC not available — track client-side only
        return success_response({'ok': True, 'stamina': max(0, 100 - amount), 'drained': amount})


@app.route('/api/stamina/regen', methods=['POST'])
@require_auth
def regen_stamina():
    """Regenerate stamina (e.g. in break room)."""
    data = request.get_json(silent=True) or {}
    amount = int(data.get('amount', 10))
    try:
        res = supabase.rpc('regen_stamina', {'p_employee_id': request.user_id, 'p_amount': amount}).execute()
        result = res.data
        if isinstance(result, list):
            result = result[0] if result else {}
        return success_response(result)
    except Exception:
        return success_response({'ok': True, 'stamina': 100, 'regen': amount})


# ==================== POSITIONS (Game) ====================

@app.route('/api/positions', methods=['GET'])
@require_auth
def get_positions():
    """Get all online player positions for multiplayer."""
    try:
        from datetime import datetime, timedelta
        cutoff = (datetime.utcnow() - timedelta(minutes=5)).isoformat()
        res = supabase.table('positions').select('user_id, x, y, direction, scene, updated_at').gte('updated_at', cutoff).execute()
        players = []
        for p in (res.data or []):
            try:
                emp = supabase.table('employees').select('full_name, level, title, avatar_color, status').eq('id', p['user_id']).single().execute()
                p['employee'] = emp.data
            except Exception:
                p['employee'] = {}
            players.append(p)
        return success_response({'players': players})
    except Exception:
        # Table doesn't exist yet — return empty
        return success_response({'players': []})


@app.route('/api/positions/update', methods=['POST'])
@require_auth
def update_position():
    """Update player position in the game world."""
    data = request.get_json(silent=True) or {}
    x = int(data.get('x', 640))
    y = int(data.get('y', 860))
    direction = data.get('direction', 'down')
    try:
        supabase.rpc('update_position', {'p_user_id': request.user_id, 'p_x': x, 'p_y': y, 'p_direction': direction}).execute()
        return success_response({'ok': True})
    except Exception:
        # RPC/table not available — silently succeed
        return success_response({'ok': True})


# ==================== PRESENCE ====================

@app.route('/api/presence/status', methods=['POST'])
@require_auth
def update_presence():
    """Update player status (working, on_break, idle, in_meeting)."""
    data = request.get_json(silent=True) or {}
    status = data.get('status', 'working')
    if status not in ('working', 'on_break', 'idle', 'in_meeting', 'offline'):
        return error_response('invalid status', 400)
    try:
        supabase.table('employees').update({'status': status, 'updated_at': 'now()'}).eq('id', request.user_id).execute()
        return success_response({'status': status})
    except Exception:
        # status column may not exist yet
        return success_response({'status': status})


@app.route('/api/presence/online', methods=['GET'])
@require_auth
def get_online_users():
    """Get list of online users with their status."""
    try:
        res = supabase.table('employees').select('id, full_name, level, title, avatar_color, status, stamina').neq('status', 'offline').eq('is_active', True).execute()
        return success_response({'users': res.data or []})
    except Exception:
        try:
            res = supabase.table('employees').select('id, full_name, level, title').eq('is_active', True).execute()
            return success_response({'users': res.data or []})
        except Exception:
            return success_response({'users': []})


# ==================== BADGES (Server-side) ====================

@app.route('/api/badges', methods=['GET'])
@require_auth
def get_badges():
    """Get all badges with user's earned status."""
    try:
        all_badges = supabase.table('badges').select('*').execute()
        earned = supabase.table('user_badges').select('badge_id, earned_at').eq('user_id', request.user_id).execute()
        earned_map = {b['badge_id']: b['earned_at'] for b in (earned.data or [])}
        badges = []
        for b in (all_badges.data or []):
            b['earned'] = b['id'] in earned_map
            b['earned_at'] = earned_map.get(b['id'])
            badges.append(b)
        return success_response({'badges': badges})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/badges/check', methods=['POST'])
@require_auth
def check_badges():
    """Check and award any newly earned badges for the user."""
    try:
        # Get user stats
        emp = supabase.table('employees').select('*').eq('id', request.user_id).single().execute()
        emp_data = emp.data or {}

        # Get today's activities
        from datetime import datetime
        today = datetime.utcnow().strftime('%Y-%m-%d')
        activities = supabase.table('actions_log').select('*').eq('employee_id', request.user_id).gte('created_at', today).execute()
        acts = activities.data or []

        # Get already earned
        earned = supabase.table('user_badges').select('badge_id').eq('user_id', request.user_id).execute()
        earned_ids = {b['badge_id'] for b in (earned.data or [])}

        # Get all badges
        all_badges = supabase.table('badges').select('*').execute()
        newly_earned = []

        for badge in (all_badges.data or []):
            if badge['id'] in earned_ids:
                continue
            criteria = badge.get('criteria', {})
            btype = criteria.get('type', '')

            awarded = False
            if btype == 'deals_closed' and (emp_data.get('total_deals', 0) >= criteria.get('count', 1)):
                awarded = True
            elif btype == 'daily_calls':
                call_count = sum(1 for a in acts if a.get('action') in ('call_made',))
                if call_count >= criteria.get('count', 100):
                    awarded = True
            elif btype == 'streak':
                awarded = False  # Client-tracked
            elif btype == 'early_login':
                hour = datetime.utcnow().hour
                if hour < criteria.get('hour', 7):
                    awarded = True
            elif btype == 'late_work':
                hour = datetime.utcnow().hour
                if hour >= criteria.get('hour', 22):
                    awarded = True

            if awarded:
                with contextlib.suppress(Exception):
                    supabase.rpc('award_badge', {'p_user_id': request.user_id, 'p_badge_name': badge['name']}).execute()
                newly_earned.append(badge)

        return success_response({'newly_earned': newly_earned, 'count': len(newly_earned)})
    except Exception:
        # badges/user_badges tables may not exist yet
        return success_response({'newly_earned': [], 'count': 0})


# ==================== STORE ====================

@app.route('/api/store', methods=['GET'])
@require_auth
def get_store():
    """عرض منتجات المتجر"""
    try:
        emp = supabase.table('employees').select('level, syb_coins').eq(
            'id', request.user_id
        ).single().execute()
        coins = emp.data.get('syb_coins', 0)
        level = emp.data.get('level', 1)
    except Exception:
        coins, level = 0, 1

    try:
        items = supabase.table('store_items').select('*').eq(
            'is_available', True
        ).execute()
        for item in (items.data or []):
            item['can_afford'] = coins >= item.get('coin_price', 0)
            item['level_met'] = level >= item.get('level_required', 0)
            item['can_buy'] = item['can_afford'] and item['level_met']
        return jsonify({'items': items.data or [], 'my_coins': coins, 'my_level': level})
    except Exception:
        return jsonify({'items': [], 'my_coins': coins, 'my_level': level})


@app.route('/api/store/buy', methods=['POST'])
@require_auth
def buy_item():
    """شراء منتج من المتجر"""
    data = request.json or {}
    item_id = data.get('item_id')
    try:
        result = supabase.rpc('purchase_store_item', {
            'p_employee_id': request.user_id,
            'p_item_id': item_id
        }).execute()

        if result.data and result.data.get('error'):
            return jsonify({'error': result.data['error']}), 400

        return jsonify(result.data)
    except Exception as e:
        return error_response(f'خطأ في الشراء: {str(e)}', 500)


# ==================== EMPLOYEE PROFILE ====================

@app.route('/api/me', methods=['GET'])
@require_auth
@cached_response(ttl=10)
def get_profile():
    """بروفايل الموظف الحالي (يتخزن مؤقتاً 10 ثوانٍ لتخفيف الضغط)"""
    try:
        emp = supabase.table('employees').select('*').eq(
            'id', request.user_id
        ).single().execute()
    except Exception:
        return error_response('الموظف غير موجود', 404)

    # جيب المشتريات النشطة (الجدول ممكن مش موجود)
    purchases_data = []
    try:
        purchases = supabase.table('employee_purchases').select(
            '*, store_items(name, icon, item_type)'
        ).eq('employee_id', request.user_id).eq('is_active', True).execute()
        purchases_data = purchases.data
    except Exception:
        pass

    # جيب عدد المهمات المعلقة
    pending_count = 0
    try:
        pending_quests = supabase.table('quests').select(
            'id', count='exact'
        ).eq('employee_id', request.user_id).eq('status', 'pending').execute()
        pending_count = pending_quests.count
    except Exception:
        pass

    return jsonify({
        'employee': emp.data,
        'active_purchases': purchases_data,
        'pending_quests': pending_count
    })


@app.route('/api/me/mode', methods=['PATCH'])
@require_auth
def toggle_professional_mode():
    """الموظف يبدّل بين الوضع الاحترافي ووضع اللعب"""
    try:
        body = request.get_json() or {}
        mode = bool(body.get('professional_mode', False))
        supabase.table('employees').update(
            {'professional_mode': mode}
        ).eq('id', request.user_id).execute()
        return success_response({'professional_mode': mode},
                                'تم التبديل للوضع الاحترافي ✅' if mode else 'تم التبديل لوضع اللعب 🎮')
    except Exception as e:
        if 'professional_mode' in str(e) and ('not find' in str(e) or 'does not exist' in str(e)):
            return success_response({'professional_mode': mode},
                                    'شغّل schema_complete.sql في Supabase لتفعيل هذه الميزة')
        return error_response(f"Mode Toggle Error: {str(e)}", 500)


@app.route('/api/me/stats', methods=['GET'])
@require_auth
@cached_response(ttl=20)
def get_my_stats():
    """إحصائيات مفصلة للموظف (مخزن مؤقتاً 20 ثانية)"""
    days = int(request.args.get('days', 7))
    since = (datetime.utcnow() - timedelta(days=days)).isoformat()

    try:
        actions = supabase.table('actions_log').select('action, xp_earned, coins_earned, created_at').eq(
            'employee_id', request.user_id
        ).gte('created_at', since).execute()
    except Exception:
        return success_response({'stats': {'messages_sent':0,'calls_made':0,'deals_closed':0,'meetings_booked':0,'total_xp':0,'total_coins':0,'daily_breakdown':{},'deals_closed_total':0}, 'period_days': days})

    # حسب الإحصائيات
    stats = {
        'messages_sent': 0,
        'calls_made': 0,
        'deals_closed': 0,
        'meetings_booked': 0,
        'total_xp': 0,
        'total_coins': 0,
        'daily_breakdown': {}
    }

    for action in (actions.data or []):
        stats['total_xp'] += action['xp_earned']
        stats['total_coins'] += action['coins_earned']

        if action['action'] == 'whatsapp_sent':
            stats['messages_sent'] += 1
        elif action['action'] == 'call_made':
            stats['calls_made'] += 1
        elif action['action'] == 'deal_closed':
            stats['deals_closed'] += 1
        elif action['action'] == 'meeting_booked':
            stats['meetings_booked'] += 1

        # التقسيم اليومي
        day = action['created_at'][:10]
        if day not in stats['daily_breakdown']:
            stats['daily_breakdown'][day] = {'xp': 0, 'actions': 0}
        stats['daily_breakdown'][day]['xp'] += action['xp_earned']
        stats['daily_breakdown'][day]['actions'] += 1

    return jsonify({'stats': stats, 'period_days': days})


# ═══════════════════════════════════════════════════════════════
# DOCUMENT ARCHIVE
# ═══════════════════════════════════════════════════════════════

@app.route('/api/documents', methods=['GET'])
@require_auth
def list_documents():
    """قائمة المستندات (مع تصفية حسب entity_type + entity_id + category)."""
    try:
        q = supabase.table('documents').select('*').is_('deleted_at', 'null').order('uploaded_at', desc=True)
        et = request.args.get('entity_type')
        eid = request.args.get('entity_id')
        cat = request.args.get('category')
        if et:
            q = q.eq('entity_type', et)
        if eid:
            q = q.eq('entity_id', eid)
        if cat:
            q = q.eq('category', cat)
        limit = min(int(request.args.get('limit', 100)), 500)
        result = q.limit(limit).execute()
        return jsonify({'documents': result.data or []})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/documents', methods=['POST'])
@require_auth
def upload_document_meta():
    """تسجيل مستند جديد (الـ file URL بييجي من Supabase Storage بعد الـ upload الـ client-side)."""
    try:
        body = request.get_json() or {}
        required = ['title', 'file_url', 'file_name', 'entity_type']
        for k in required:
            if not body.get(k):
                return error_response(f'الحقل {k} مطلوب', 400)
        if body['entity_type'] not in ('employee', 'lead', 'transaction', 'property', 'contract', 'general'):
            return error_response('entity_type غير صحيح', 400)
        record = {
            'title': body['title'][:200],
            'description': (body.get('description') or '')[:1000] or None,
            'file_url': body['file_url'],
            'file_name': body['file_name'][:255],
            'file_size': body.get('file_size'),
            'mime_type': (body.get('mime_type') or '')[:100] or None,
            'entity_type': body['entity_type'],
            'entity_id': body.get('entity_id'),
            'category': (body.get('category') or 'other')[:50],
            'tags': body.get('tags') or [],
            'uploaded_by': request.user_id,
            'metadata': body.get('metadata') or {},
        }
        r = supabase.table('documents').insert(record).execute()
        return success_response({'document': (r.data or [None])[0]}, 'تم رفع المستند')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/documents/<doc_id>', methods=['DELETE'])
@check_role(['admin', 'manager'])
@audit_log('document_deleted', 'document')
def delete_document(doc_id):
    """حذف مستند (soft delete)."""
    try:
        supabase.table('documents').update({
            'deleted_at': datetime.utcnow().isoformat()
        }).eq('id', doc_id).execute()
        return success_response(None, 'تم الحذف')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/employees/<user_id>/profile', methods=['GET'])
@require_auth
@cached_response(ttl=20)
def admin_employee_profile(user_id):
    """بروفايل تفصيلي لموظف — بيانات أساسية + إحصائيات + شارات + آخر النشاطات."""
    try:
        # Basic employee info
        emp = supabase.table('employees').select(
            'id, email, full_name, avatar_color, role, level, title, total_xp, current_xp, '
            'syb_coins, total_deals, total_leads, total_messages, total_calls, '
            'is_active, is_locked, daily_streak, stamina, last_seen_at, created_at'
        ).eq('id', user_id).single().execute()
        emp_data = emp.data
        if not emp_data:
            return error_response('الموظف غير موجود', 404)

        # Last 30 days stats
        since = (datetime.utcnow() - timedelta(days=30)).isoformat()
        actions = supabase.table('actions_log').select('action, xp_earned, coins_earned, created_at').eq(
            'employee_id', user_id
        ).gte('created_at', since).execute()
        action_rows = actions.data or []
        stats_30d = {
            'messages_sent': sum(1 for a in action_rows if a.get('action') == 'whatsapp_sent'),
            'calls_made': sum(1 for a in action_rows if a.get('action') == 'call_made'),
            'deals_closed': sum(1 for a in action_rows if a.get('action') == 'deal_closed'),
            'meetings_booked': sum(1 for a in action_rows if a.get('action') == 'meeting_booked'),
            'total_xp_30d': sum(a.get('xp_earned', 0) for a in action_rows),
            'total_actions': len(action_rows),
        }

        # Daily breakdown for last 14 days
        daily = {}
        for a in action_rows:
            day = (a.get('created_at') or '')[:10]
            if not day:
                continue
            if day not in daily:
                daily[day] = {'xp': 0, 'actions': 0, 'deals': 0}
            daily[day]['xp'] += a.get('xp_earned', 0)
            daily[day]['actions'] += 1
            if a.get('action') == 'deal_closed':
                daily[day]['deals'] += 1

        # Badges earned (try; table may not exist)
        badges = []
        try:
            b = supabase.table('employee_badges').select(
                '*, badges:badge_id(name, description, icon, tier)'
            ).eq('employee_id', user_id).execute()
            badges = b.data or []
        except Exception:
            pass

        # Active leads count
        active_leads = 0
        try:
            r = supabase.table('leads').select('id', count='exact').eq(
                'assigned_to', user_id
            ).in_('status', ['new', 'contacted', 'interested', 'meeting_set', 'negotiation']).execute()
            active_leads = r.count or 0
        except Exception:
            pass

        # Recent leads (5 most recently updated)
        recent_leads = []
        try:
            r = supabase.table('leads').select('id, name, phone, status, updated_at').eq(
                'assigned_to', user_id
            ).order('updated_at', desc=True).limit(5).execute()
            recent_leads = r.data or []
        except Exception:
            pass

        # Current rank
        rank_info = {'rank': None, 'total_employees': 0}
        try:
            ranked = supabase.rpc('get_leaderboard_rank', {'p_employee_id': user_id}).execute()
            if ranked.data:
                rank_info = ranked.data if isinstance(ranked.data, dict) else {'rank': ranked.data}
        except Exception:
            # Fallback: count manually
            try:
                all_emps = supabase.table('employees').select('id, total_xp').eq('is_active', True).execute().data or []
                sorted_emps = sorted(all_emps, key=lambda e: e.get('total_xp', 0), reverse=True)
                for idx, e in enumerate(sorted_emps):
                    if e['id'] == user_id:
                        rank_info = {'rank': idx + 1, 'total_employees': len(sorted_emps)}
                        break
            except Exception:
                pass

        return jsonify({
            'employee': emp_data,
            'stats_30d': stats_30d,
            'daily_breakdown': daily,
            'badges': badges,
            'active_leads_count': active_leads,
            'recent_leads': recent_leads,
            'rank': rank_info,
        })
    except Exception as e:
        return error_response(f'فشل التحميل: {str(e)}', 500)


# ═══════════════════════════════════════════════════════════════
# AI BOT ADMIN ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.route('/api/admin/bot/conversations', methods=['GET'])
@check_role(['admin', 'manager'])
@cached_response(ttl=10)
def admin_bot_conversations():
    """قائمة محادثات البوت (آخر ٥٠ محادثة)."""
    try:
        state = request.args.get('state')
        q = supabase.table('bot_conversations').select('*').order('last_message_at', desc=True).limit(50)
        if state:
            q = q.eq('state', state)
        r = q.execute()
        return jsonify({'conversations': r.data or []})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/bot/conversations/<conv_id>/messages', methods=['GET'])
@check_role(['admin', 'manager'])
def admin_bot_messages(conv_id):
    """رسائل محادثة محددة."""
    try:
        r = supabase.table('bot_messages').select('*').eq(
            'conversation_id', conv_id
        ).order('created_at', desc=False).execute()
        return jsonify({'messages': r.data or []})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/job-applications', methods=['GET'])
@check_role(['admin', 'manager'])
@cached_response(ttl=10)
def admin_job_applications():
    """قائمة طلبات التوظيف من بوت سارة."""
    try:
        status = request.args.get('status')
        q = supabase.table('job_applications').select('*').order('created_at', desc=True).limit(100)
        if status:
            q = q.eq('status', status)
        r = q.execute()
        return jsonify({'applications': r.data or []})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/job-applications/<app_id>/status', methods=['PATCH'])
@check_role(['admin', 'manager'])
@audit_log('job_app_status_change', 'job_application')
def admin_change_job_status(app_id):
    """تعديل حالة متقدم وظيفة (مثلاً: قبول/رفض/تعيين)."""
    try:
        body = request.get_json() or {}
        new_status = (body.get('status') or '').strip()
        allowed = ['pending', 'screening', 'qualified', 'rejected', 'hired', 'archived']
        if new_status not in allowed:
            return error_response(f'حالة غير صحيحة. المسموح: {allowed}', 400)
        supabase.table('job_applications').update({
            'status': new_status,
            'updated_at': datetime.utcnow().isoformat(),
        }).eq('id', app_id).execute()
        return success_response(None, 'تم التحديث')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/online-users', methods=['GET'])
@require_auth
@cached_response(ttl=10)
def admin_online_users():
    """الموظفين الفاتحين السيستم في آخر 5 دقايق."""
    try:
        cutoff = (datetime.utcnow() - timedelta(minutes=5)).isoformat()
        r = supabase.table('employees').select(
            'id, full_name, email, avatar_color, role, title, level, last_seen_at, is_locked'
        ).gte('last_seen_at', cutoff).order('last_seen_at', desc=True).execute()
        users = r.data or []
        # Filter out locked + add online status (active in last 2 min)
        now = datetime.utcnow()
        for u in users:
            try:
                ls = datetime.fromisoformat(u['last_seen_at'].replace('Z', '+00:00')).replace(tzinfo=None)
                u['minutes_ago'] = int((now - ls).total_seconds() / 60)
                u['is_active'] = u['minutes_ago'] < 2
            except Exception:
                u['minutes_ago'] = 99
                u['is_active'] = False
        return jsonify({'users': users, 'count': len(users), 'active_count': sum(1 for u in users if u.get('is_active'))})
    except Exception as e:
        return jsonify({'users': [], 'count': 0, 'active_count': 0, 'error': str(e)})


@app.route('/api/ai/coach', methods=['GET'])
@require_auth
def ai_coach_tip():
    """نصيحة سريعة من AI بناء على بيانات الموظف الحالية (للعبة 3D)."""
    try:
        emp = supabase.table('employees').select(
            'full_name, level, total_xp, syb_coins, total_deals, total_leads, stamina, title'
        ).eq('id', request.user_id).single().execute().data or {}

        # Get recent activity (last 24h)
        since = (datetime.utcnow() - timedelta(hours=24)).isoformat()
        try:
            recent = supabase.table('actions_log').select('action, xp_earned').eq(
                'employee_id', request.user_id
            ).gte('created_at', since).execute().data or []
        except Exception:
            recent = []

        deals_today = sum(1 for a in recent if a.get('action') == 'deal_closed')
        msgs_today = sum(1 for a in recent if a.get('action') == 'whatsapp_sent')
        calls_today = sum(1 for a in recent if a.get('action') == 'call_made')

        # Build context for AI
        context = (
            f"اللاعب اسمه {emp.get('full_name','')} لقبه {emp.get('title','')}، "
            f"المستوى {emp.get('level',1)}، عدد الصفقات اليوم {deals_today}، "
            f"رسائل اليوم {msgs_today}، مكالمات اليوم {calls_today}، "
            f"رصيد العملات {emp.get('syb_coins',0)}، إجمالي الصفقات {emp.get('total_deals',0)}."
        )
        prompt = (
            f"{context}\n"
            "اقترح نصيحة قصيرة جداً (سطر واحد، أقل من 70 حرف، عربي) لتحفيز اللاعب أو "
            "تذكيره بحاجة مهمة بناء على أرقامه. لا تكتب مقدمات."
        )

        # Try Gemini → fallback to canned tip
        tip = ''
        try:
            from ai_client import generate_text
            tip = (generate_text(prompt, max_tokens=80) or '').strip().strip('"').strip("'")
            if len(tip) > 140:
                tip = tip[:140] + '...'
        except Exception as e:
            logger.warning('AI coach failed: %s', e)

        # Fallback tips based on player state
        if not tip:
            tips = []
            if deals_today == 0:
                tips.append('لسه ماقفلتش صفقة اليوم — يلا 💪')
            if msgs_today < 5:
                tips.append('ابعت رسالة WhatsApp جديدة لليدز بتاعك')
            if calls_today < 3:
                tips.append('مكالمة سريعة بترفع XP — جرب الآن')
            if emp.get('stamina', 100) < 30:
                tips.append('طاقتك قليلة — روح غرفة الاستراحة ☕')
            if not tips:
                tips.append(f"أنت في {emp.get('title','نجم')}! استمر 🚀")
            import random as _r
            tip = _r.choice(tips)

        return jsonify({'tip': tip})
    except Exception as e:
        return jsonify({'tip': 'استمر في العمل! 💪', 'error': str(e)}), 200


@app.route('/api/ai/npc-chat', methods=['POST'])
@require_auth
def ai_npc_chat():
    """محادثة مع زميل افتراضي (NPC) باستخدام Gemini AI."""
    try:
        body = request.get_json() or {}
        user_msg = (body.get('message') or '').strip()[:500]
        npc_name = (body.get('npc_name') or 'زميل')[:30]
        if not user_msg:
            return jsonify({'reply': 'إيه عاوز؟ كلمني'}), 200

        prompt = (
            f"أنت {npc_name}، مندوب مبيعات شغّال في شركة AlSaeb CRM. ردك يكون قصير "
            f"(جملة أو اثنين)، عربي عامية مصرية، وشخصيتك ودودة ومتحمسة للشغل. "
            f"الزميل بيقولك: \"{user_msg}\""
        )
        reply = ''
        try:
            from ai_client import generate_text
            reply = (generate_text(prompt, max_tokens=120) or '').strip().strip('"').strip("'")
            if len(reply) > 240:
                reply = reply[:240] + '...'
        except Exception as e:
            logger.warning('NPC chat Gemini failed: %s', e)

        if not reply:
            replies = [
                'هلا، أنا شغال على ليد جديد 💪',
                'صفقاتك كام النهارده؟',
                'لازم نقفل أكتر صفقات الأسبوع ده',
                'العميل بتاع آخر مكالمة طلع كويس!',
                'تعال نشرب قهوة بعد ما تخلص الليد ده ☕',
            ]
            import random as _r
            reply = _r.choice(replies)

        return jsonify({'reply': reply}), 200
    except Exception as e:
        return jsonify({'reply': 'مش فاهم، قول تاني', 'error': str(e)}), 200


# ==================== ADMIN: EMPLOYEE MANAGEMENT ====================

@app.route('/api/admin/employees', methods=['GET'])
@check_role(['admin'])
def list_employees():
    """كل الموظفين"""
    try:
        result = supabase.table('employees').select(
            'id, email, full_name, role, level, title, total_xp, syb_coins, total_deals, total_leads, status, avatar_color, professional_mode, created_at'
        ).order('created_at', desc=True).execute()
        return success_response({'employees': result.data})
    except Exception:
        result = supabase.table('employees').select(
            'id, email, full_name, role, level, title, total_xp, syb_coins, total_deals, total_leads, created_at'
        ).order('created_at', desc=True).execute()
        return success_response({'employees': result.data})


@app.route('/api/admin/employees', methods=['POST'])
@check_role(['admin'])
@audit_log('employee_create', target_type='employee')
def create_employee():
    """إنشاء موظف جديد (auth user + employee record)"""
    data = request.json or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    full_name = (data.get('full_name') or '').strip()
    role = data.get('role', 'sales_agent')
    title = data.get('title', '')

    if not email or not password or not full_name:
        return error_response('الإيميل والباسورد والاسم مطلوبين', 400)
    if len(password) < 8:
        return error_response('الباسورد لازم 8 حروف على الأقل', 400)
    if role not in ('sales_agent', 'manager', 'admin'):
        return error_response('الدور لازم يكون sales_agent أو manager أو admin', 400)

    # ── Seats limit check ──
    with contextlib.suppress(Exception):
        org = supabase.table('organizations').select('seats_limit').limit(1).execute()
        if org.data:
            seats_limit = org.data[0].get('seats_limit') or 0
            if seats_limit > 0:
                current_count = supabase.table('employees').select('id', count='exact') \
                    .eq('is_active', True).execute().count or 0
                if current_count >= seats_limit:
                    return error_response(
                        f'وصلت للحد الأقصى من الموظفين ({seats_limit}). رقّي خطتك لإضافة المزيد.', 403
                    )

    try:
        # 1. Create auth user via Supabase Admin API
        auth_res = supabase.auth.admin.create_user({
            'email': email,
            'password': password,
            'email_confirm': True,
        })
        user_id = auth_res.user.id

        # 2. Create employee record
        emp_data = {
            'id': user_id,
            'email': email,
            'full_name': full_name,
            'role': role,
            'title': title or ('مدير' if role == 'admin' else 'موظف مبيعات'),
            'level': 1,
            'total_xp': 0,
            'current_xp': 0,
            'syb_coins': 0,
            'total_deals': 0,
            'total_leads': 0,
            'avatar_color': '#00e5ff',
            'stamina': 100,
            'max_stamina': 100,
            'status': 'working',
        }
        result = supabase.table('employees').insert(emp_data).execute()
        return success_response(result.data[0] if result.data else emp_data, 'تم إنشاء الموظف بنجاح', 201)
    except Exception as e:
        err_msg = str(e)
        if 'already been registered' in err_msg or 'already exists' in err_msg:
            return error_response('الإيميل ده مسجل قبل كده', 409)
        return error_response(f'خطأ في إنشاء الموظف: {err_msg}', 500)


@app.route('/api/admin/employees/<user_id>', methods=['PATCH'])
@check_role(['admin'])
@audit_log('employee_update', target_type='employee')
def update_employee(user_id):
    """تعديل بيانات موظف (دور، وضع احترافي، الخ)"""
    try:
        body = request.get_json() or {}
        allowed = ['role', 'professional_mode', 'title', 'full_name', 'team', 'team_id', 'email']
        updates = {k: body[k] for k in allowed if k in body}
        # تحديث كلمة السر في Supabase Auth
        if body.get('password') and len(body['password']) >= 8:
            with contextlib.suppress(Exception):
                supabase.auth.admin.update_user_by_id(user_id, {'password': body['password']})
        # تحديث الإيميل في Supabase Auth
        if body.get('email'):
            with contextlib.suppress(Exception):
                supabase.auth.admin.update_user_by_id(user_id, {'email': body['email']})
        if not updates:
            if body.get('password'):
                return success_response({}, 'تم تعديل كلمة السر ✅')
            return error_response('لا توجد بيانات للتحديث', 400)
        result = supabase.table('employees').update(updates).eq('id', user_id).execute()
        return success_response(result.data[0] if result.data else {}, 'تم تعديل الموظف ✅')
    except Exception as e:
        return error_response(f"Employee Update Error: {str(e)}", 500)


@app.route('/api/admin/employees/<user_id>/lock', methods=['POST'])
@check_role(['admin'])
@audit_log('employee_locked', 'employee')
def lock_employee(user_id):
    """قفل الموظف — يمنعه من تسجيل الدخول."""
    try:
        body = request.get_json(silent=True) or {}
        reason = (body.get('reason') or '').strip()[:200] or 'تم القفل بواسطة الأدمن'
        supabase.table('employees').update({
            'is_locked': True,
            'locked_at': datetime.utcnow().isoformat(),
            'locked_reason': reason,
        }).eq('id', user_id).execute()
        _invalidate_lock_cache(user_id)
        return success_response(None, 'تم قفل الموظف')
    except Exception as e:
        return error_response(f'فشل القفل: {str(e)}', 500)


@app.route('/api/admin/employees/<user_id>/unlock', methods=['POST'])
@check_role(['admin'])
@audit_log('employee_unlocked', 'employee')
def unlock_employee(user_id):
    """فتح القفل — يسمح للموظف بتسجيل الدخول مرة أخرى."""
    try:
        supabase.table('employees').update({
            'is_locked': False,
            'locked_at': None,
            'locked_reason': None,
        }).eq('id', user_id).execute()
        _invalidate_lock_cache(user_id)
        return success_response(None, 'تم فتح القفل')
    except Exception as e:
        return error_response(f'فشل الفتح: {str(e)}', 500)


@app.route('/api/admin/employees/<user_id>', methods=['DELETE'])
@check_role(['admin'])
@audit_log('employee_delete', target_type='employee')
def delete_employee(user_id):
    """حذف موظف (auth + employee record)"""
    if user_id == request.user_id:
        return error_response('مش ممكن تحذف نفسك', 400)
    try:
        # حذف كل السجلات المرتبطة بالموظف — شامل كل الجداول
        emp_id_tables = [
            'employee_purchases', 'user_badges', 'actions_log',
            'attendance', 'positions', 'high_fives',
            'competition_participants', 'daily_highlights',
            'rate_limit_log', 'quests', 'activity_feed',
            'admin_audit_log', 'story_progress', 'break_room_sessions',
            'lead_status_history', 'whatsapp_messages',
            'squad_members', 'cosmetics_inventory',
        ]
        for tbl in emp_id_tables:
            with contextlib.suppress(Exception):
                supabase.table(tbl).delete().eq('employee_id', user_id).execute()
        # جداول بمفتاح مختلف عن employee_id
        alt_key_deletes = [
            ('high_fives', 'target_id'),
            ('high_fives', 'from_user'),
            ('high_fives', 'to_user'),
            ('admin_audit_log', 'admin_id'),
            ('hot_lead_events', 'spawned_by'),
            ('hot_lead_events', 'claimed_by'),
            ('power_hours', 'started_by'),
            ('daily_highlights', 'generated_by'),
            ('competitions', 'winner_emp_id'),
            ('squads', 'manager_id'),
            ('ceo_visits', 'created_by'),
            ('workflow_rules', 'created_by'),
            ('player_settings', 'user_id'),
            ('employee_permissions', 'user_id'),
        ]
        for tbl, col in alt_key_deletes:
            with contextlib.suppress(Exception):
                supabase.table(tbl).delete().eq(col, user_id).execute()
        # فك ربط الليدز + locks
        with contextlib.suppress(Exception):
            supabase.table('leads').update({'assigned_to': None}).eq('assigned_to', user_id).execute()
        with contextlib.suppress(Exception):
            supabase.table('leads').update({'locked_by': None, 'locked_at': None}).eq('locked_by', user_id).execute()
        # listings/transactions agent references (SET NULL)
        for tbl in ['listings', 'transactions', 'owner_contacts', 'viewing_requests', 'change_requests']:
            with contextlib.suppress(Exception):
                supabase.table(tbl).update({'agent_id': None}).eq('agent_id', user_id).execute()

        supabase.table('employees').delete().eq('id', user_id).execute()
        with contextlib.suppress(Exception):
            supabase.auth.admin.delete_user(user_id)
        return success_response(None, 'تم حذف الموظف')
    except Exception as e:
        return error_response(f'خطأ في الحذف: {str(e)}', 500)


# ==================== PUSH NOTIFICATIONS ====================

def send_push_to_employee(employee_id: str, title: str, body: str, url: str = '/agent'):
    """ابعت Push Notification لموظف محدد (كل أجهزته)."""
    vapid_private = os.getenv('VAPID_PRIVATE_KEY', '')
    vapid_email = os.getenv('VAPID_EMAIL', 'mailto:admin@al-saeb-crm.online')
    if not vapid_private:
        return
    def _send():
        try:
            from pywebpush import WebPushException, webpush
            subs = supabase.table('push_subscriptions').select('subscription').eq(
                'employee_id', employee_id).execute()
            for row in (subs.data or []):
                sub_info = row['subscription']
                if isinstance(sub_info, str):
                    sub_info = json.loads(sub_info)
                try:
                    webpush(
                        subscription_info=sub_info,
                        data=json.dumps({'title': title, 'body': body, 'url': url}),
                        vapid_private_key=vapid_private,
                        vapid_claims={'sub': vapid_email},
                        timeout=10,
                    )
                except WebPushException:
                    # Subscription expired — clean up
                    with contextlib.suppress(Exception):
                        supabase.table('push_subscriptions').delete().eq(
                            'subscription', json.dumps(sub_info)).execute()
        except Exception:
            pass
    threading.Thread(target=_send, daemon=True).start()


@app.route('/api/push/subscribe', methods=['POST'])
@require_auth
def push_subscribe():
    """سجّل اشتراك Push Notification لجهاز الموظف."""
    try:
        body = request.get_json() or {}
        subscription = body.get('subscription')
        if not subscription:
            return error_response('subscription مطلوب', 400)
        sub_str = json.dumps(subscription) if isinstance(subscription, dict) else subscription
        # Upsert — لو الاشتراك موجود قبل كده، متضيفوش تاني
        existing = supabase.table('push_subscriptions').select('id').eq(
            'employee_id', request.user_id).eq('subscription', sub_str).execute()
        if not existing.data:
            supabase.table('push_subscriptions').insert({
                'employee_id': request.user_id,
                'subscription': sub_str,
            }).execute()
        return success_response(None, 'تم تسجيل الإشعارات')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/push/unsubscribe', methods=['POST'])
@require_auth
def push_unsubscribe():
    """إلغاء اشتراك Push."""
    try:
        supabase.table('push_subscriptions').delete().eq(
            'employee_id', request.user_id).execute()
        return success_response(None, 'تم إلغاء الإشعارات')
    except Exception as e:
        return error_response(str(e), 500)


# ==================== TEMPLATES ====================

@app.route('/api/templates', methods=['GET'])
@require_auth
def get_templates():
    """جيب قوالب الرسايل"""
    category = request.args.get('category')
    language = request.args.get('language')

    query = supabase.table('message_templates').select('*').eq('is_active', True)

    if category:
        query = query.eq('category', category)
    if language:
        query = query.eq('language', language)

    result = query.order('times_used', desc=True).execute()
    return jsonify({'templates': result.data})


# ==================== AI HINTS ====================

@app.route('/api/ai/hint', methods=['POST'])
@rate_limit('ai_hint', max_per_min=20)
def get_ai_hint():
    """
    طلب تحليل من الـ AI.
    بيروح لـ Anthropic Claude لو في ANTHROPIC_API_KEY,
    أو OpenAI لو في OPENAI_API_KEY,
    ولو مفيش، heuristic fallback.
    """
    data = request.json or {}
    lead_id = data.get('lead_id')
    if not lead_id:
        return error_response('lead_id required')

    try:
        lead = supabase.table('leads').select(
            '*, projects(name)'
        ).eq('id', lead_id).single().execute()
        if not lead.data:
            return error_response('Lead not found', 404)

        msgs = supabase.table('whatsapp_messages').select('*').eq(
            'lead_id', lead_id
        ).order('created_at', desc=True).limit(15).execute()

        hint = ai_client.analyze_lead(lead.data, msgs.data or [])

        # حدث score على الليد لو الـ AI اقترح واحد
        if isinstance(hint.get('lead_score'), int | float):
            supabase.table('leads').update({
                'lead_score':       int(hint['lead_score']),
                'ai_summary':       hint.get('suggestion'),
                'score_updated_at': datetime.utcnow().isoformat(),
            }).eq('id', lead_id).execute()

        # 🎮 XP لاستخدام الـ AI
        xp_config = XP_CONFIG['ai_hint_used']
        supabase.rpc('award_xp_and_coins', {
            'p_employee_id': request.user_id,
            'p_xp':          xp_config['xp'],
            'p_coins':       xp_config['coins'],
            'p_reason':      'استخدام تلميح AI',
            'p_ref_type':    'action',
        }).execute()

        return success_response({'hint': hint})
    except Exception as e:
        return error_response(f"AI Hint Error: {str(e)}", 500)


@app.route('/api/admin/leads/score-batch', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('batch_lead_scoring')
def batch_lead_scoring():
    """
    Score unscored leads using heuristic rules (fast, no AI API needed).
    Factors: status progression, contact_count, recency, campaign quality.
    """
    try:
        leads = supabase.table('leads').select(
            'id, status, contact_count, created_at, updated_at, quality'
        ).is_('lead_score', 'null').eq('is_active', True).limit(200).execute()

        scored = 0
        for lead in (leads.data or []):
            score = _calculate_lead_score(lead)
            supabase.table('leads').update({
                'lead_score': score,
                'score_updated_at': datetime.utcnow().isoformat(),
            }).eq('id', lead['id']).execute()
            scored += 1

        return success_response({'scored': scored}, f'تم تقييم {scored} ليد')
    except Exception as e:
        return error_response(f"Scoring Error: {str(e)}", 500)


def _calculate_lead_score(lead: dict) -> int:
    """Heuristic lead score 0-100 based on engagement signals."""
    score = 30  # base

    # Status progression bonus
    status_scores = {
        'new': 0, 'contacted': 10, 'interested': 25,
        'meeting_set': 40, 'negotiation': 55,
        'closed_won': 95, 'closed_lost': 5,
    }
    score += status_scores.get(lead.get('status', 'new'), 0)

    # Contact frequency
    contacts = lead.get('contact_count') or 0
    if contacts >= 5:
        score += 15
    elif contacts >= 2:
        score += 8

    # Recency (how recently was the lead updated)
    try:
        updated = datetime.fromisoformat(lead['updated_at'].replace('Z', '+00:00'))
        days_ago = (datetime.now(updated.tzinfo) - updated).days
        if days_ago <= 1:
            score += 10
        elif days_ago <= 7:
            score += 5
        elif days_ago > 30:
            score -= 15
    except Exception:
        pass

    # Quality bonus
    quality = lead.get('quality', '')
    if quality == 'hot':
        score += 10
    elif quality == 'warm':
        score += 5
    elif quality == 'cold':
        score -= 10

    return max(0, min(100, score))


@app.route('/api/analytics/lead-scores', methods=['GET'])
@check_role(['admin', 'manager'])
def get_lead_score_distribution():
    """Distribution of lead scores for analytics."""
    try:
        leads = supabase.table('leads').select(
            'lead_score, status'
        ).eq('is_active', True).not_.is_('lead_score', 'null').execute()

        buckets = {'0-20': 0, '21-40': 0, '41-60': 0, '61-80': 0, '81-100': 0}
        for lead in (leads.data or []):
            s = lead.get('lead_score', 0)
            if s <= 20:
                buckets['0-20'] += 1
            elif s <= 40:
                buckets['21-40'] += 1
            elif s <= 60:
                buckets['41-60'] += 1
            elif s <= 80:
                buckets['61-80'] += 1
            else:
                buckets['81-100'] += 1

        return success_response({
            'distribution': buckets,
            'total_scored': len(leads.data or []),
        })
    except Exception as e:
        return error_response(str(e), 500)


# ==================== ADMIN: IMPORT ====================

@app.route('/api/admin/import', methods=['POST'])
@check_role(['admin'])
def import_leads():
    """استيراد ليدز من ملف (Admin فقط)"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    # TODO: حفظ الملف مؤقتاً وتشغيل import_leads.py
    # هنا ممكن نستخدم نفس الـ logic من import_leads.py

    return jsonify({'message': 'Import started', 'filename': file.filename})


# ==================== ADMIN: DASHBOARD (Advanced) ====================

@app.route('/api/admin/dashboard/stats', methods=['GET'])
@check_role(['admin'])
def admin_stats_advanced():
    """
    Complex Admin Analytics including conversion funnel and performance trends.
    This query is designed to be heavy but highly detailed for a CEO level overview.
    """
    try:
        # 1. Pipeline Overview
        counts = supabase.rpc('get_pipeline_stats').execute()

        # 2. Daily Performance (Last 7 days)
        today = datetime.utcnow().date()
        week_ago = (today - timedelta(days=7)).isoformat()

        deals_trend = supabase.table('actions_log') \
            .select('created_at, action') \
            .eq('action', 'deal_closed') \
            .gte('created_at', week_ago) \
            .execute()

        # 3. Top Performers (XP Based)
        top_hunters = supabase.table('employees') \
            .select('full_name, level, total_xp, total_deals') \
            .eq('role', 'sales_agent') \
            .order('total_xp', desc=True) \
            .limit(5) \
            .execute()

        # 4. Lead Efficiency (Conversion Rate)
        # Logic: (Deals Won / Total Contacted Leads) * 100
        # We perform these calculations on the backend to keep the Flutter UI light

        # 5. Conversion Funnel
        funnel_order = ['new', 'contacted', 'interested', 'meeting_set', 'negotiation', 'closed_won', 'closed_lost']
        funnel = {}
        for row in (counts.data or []):
            funnel[row['status']] = row['count']

        # 6. Campaign Performance
        try:
            campaigns = supabase.table('projects').select('id, name').eq('is_active', True).execute()
            campaign_stats = []
            for c in (campaigns.data or [])[:10]:
                cstat = supabase.table('leads').select('status', count='exact').eq('project_id', c['id']).execute()
                won = supabase.table('leads').select('id', count='exact').eq('project_id', c['id']).eq('status', 'closed_won').execute()
                total = cstat.count or 0
                won_c = won.count or 0
                campaign_stats.append({
                    'name': c['name'],
                    'total': total,
                    'won': won_c,
                    'rate': round(won_c / total * 100, 1) if total > 0 else 0,
                })
        except Exception:
            campaign_stats = []

        # 7. Agent Activity (last 7 days)
        try:
            actions_7d = supabase.table('actions_log').select(
                'employee_id, action, created_at'
            ).gte('created_at', week_ago).execute()
            action_counts = {}
            for a in (actions_7d.data or []):
                act = a.get('action', 'other')
                action_counts[act] = action_counts.get(act, 0) + 1
        except Exception:
            action_counts = {}

        return success_response({
            'pipeline': counts.data,
            'trends': deals_trend.data,
            'leaderboard_mini': top_hunters.data,
            'funnel': [{'status': s, 'count': funnel.get(s, 0)} for s in funnel_order],
            'campaigns': campaign_stats,
            'action_breakdown': action_counts,
            'summary': {
                'total_xp_pool': sum(h.get('total_xp', 0) for h in (top_hunters.data or [])),
                'total_leads': sum(funnel.values()),
                'total_won': funnel.get('closed_won', 0),
                'conversion_rate': round(funnel.get('closed_won', 0) / max(1, sum(funnel.values())) * 100, 1),
            }
        })
    except Exception as e:
        return error_response(f"Analytics Engine Error: {str(e)}", 500)


# ==================== SECURITY HELPERS ====================

WHATSAPP_APP_SECRET = os.getenv('WHATSAPP_APP_SECRET', '')

def verify_whatsapp_signature(raw_body: bytes, signature_header: str) -> bool:
    """
    تحقق من صحة توقيع Meta على الـ webhook.
    Meta بتبعت X-Hub-Signature-256: sha256=<hex>
    """
    if not WHATSAPP_APP_SECRET or not signature_header:
        return False
    try:
        prefix, hex_sig = signature_header.split('=', 1)
    except ValueError:
        return False
    if prefix != 'sha256':
        return False
    expected = hmac.new(
        WHATSAPP_APP_SECRET.encode('utf-8'),
        raw_body,
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected, hex_sig)


# (audit_log + rate_limit moved up near check_role for import-time availability)


# ==================== LEAD LOCKING ====================

@app.route('/api/leads/<lead_id>/lock', methods=['POST'])
@rate_limit('lead_lock', max_per_min=60)
def lock_lead(lead_id):
    """احجز ليد عشان محدش تاني يفتحه في نفس الوقت."""
    data = request.json or {}
    minutes = int(data.get('duration_minutes', 15))
    try:
        result = supabase.rpc('acquire_lead_lock', {
            'p_lead_id':          lead_id,
            'p_employee_id':      request.user_id,
            'p_duration_minutes': minutes,
        }).execute()
        payload = result.data or {}
        if payload.get('error'):
            status = 409 if payload['error'] == 'already_locked' else 400
            return error_response(payload, status)

        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id':     lead_id,
            'action':      'lead_locked',
            'details':     payload,
        }).execute()
        return success_response(payload, 'Lead locked')
    except Exception as e:
        return error_response(f"Lock Error: {str(e)}", 500)


@app.route('/api/leads/<lead_id>/lock', methods=['DELETE'])
@require_auth
def unlock_lead(lead_id):
    """حرر الـ lock بتاعك."""
    try:
        result = supabase.rpc('release_lead_lock', {
            'p_lead_id':     lead_id,
            'p_employee_id': request.user_id,
        }).execute()
        payload = result.data or {}
        if payload.get('error'):
            return error_response(payload, 403)

        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id':     lead_id,
            'action':      'lead_unlocked',
        }).execute()
        return success_response(payload, 'Lead unlocked')
    except Exception as e:
        return error_response(f"Unlock Error: {str(e)}", 500)


# ==================== LEAD NOTES ====================

@app.route('/api/leads/<lead_id>/notes', methods=['GET'])
@require_auth
def list_lead_notes(lead_id):
    """كل الملاحظات المرتبطة بالليد."""
    try:
        result = supabase.table('lead_notes').select(
            '*, employees(full_name, avatar_url)'
        ).eq('lead_id', lead_id).order('is_pinned', desc=True).order(
            'created_at', desc=True
        ).execute()
        return success_response({'notes': result.data})
    except Exception as e:
        return error_response(f"Notes Retrieval Error: {str(e)}", 500)


@app.route('/api/leads/<lead_id>/notes', methods=['POST'])
@rate_limit('note_add', max_per_min=30)
def add_lead_note(lead_id):
    """ضيف ملاحظة على الليد (+ XP)."""
    data = request.json or {}
    note_text = (data.get('note_text') or '').strip()
    is_pinned = bool(data.get('is_pinned', False))

    if not note_text:
        return error_response('note_text is required')

    try:
        owner = supabase.table('leads').select('assigned_to').eq(
            'id', lead_id
        ).single().execute()
        if not owner.data:
            return error_response('Lead not found', 404)
        if owner.data['assigned_to'] != request.user_id:
            return error_response('Access denied', 403)

        inserted = supabase.table('lead_notes').insert({
            'lead_id':     lead_id,
            'employee_id': request.user_id,
            'note_text':   note_text,
            'is_pinned':   is_pinned,
        }).execute()

        # حدّث عداد الملاحظات على الليد
        lead_row = supabase.table('leads').select('notes_count').eq(
            'id', lead_id
        ).single().execute()
        supabase.table('leads').update({
            'notes_count': (lead_row.data.get('notes_count') or 0) + 1,
            'updated_at':  datetime.utcnow().isoformat(),
        }).eq('id', lead_id).execute()

        # 🎮 XP صغير (with Power Hour multiplier)
        xp_config = XP_CONFIG['note_added']
        ph_mult = get_power_hour_multiplier()
        note_xp = int(xp_config['xp'] * ph_mult)
        supabase.rpc('award_xp_and_coins', {
            'p_employee_id': request.user_id,
            'p_xp':          note_xp,
            'p_coins':       int(xp_config['coins'] * ph_mult),
            'p_reason':      'إضافة ملاحظة على ليد',
            'p_ref_type':    'note',
        }).execute()

        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id':     lead_id,
            'action':      'note_added',
            'xp_earned':   note_xp,
        }).execute()

        return success_response(inserted.data[0], 'Note added')
    except Exception as e:
        return error_response(f"Add Note Error: {str(e)}", 500)


@app.route('/api/notes/<note_id>', methods=['PATCH'])
@require_auth
def update_lead_note(note_id):
    """تعديل ملاحظة (المالك فقط)."""
    data = request.json or {}
    updates = {}
    if 'note_text' in data:
        updates['note_text'] = data['note_text']
    if 'is_pinned' in data:
        updates['is_pinned'] = bool(data['is_pinned'])
    if not updates:
        return error_response('No fields to update')
    updates['updated_at'] = datetime.utcnow().isoformat()

    try:
        result = supabase.table('lead_notes').update(updates).eq(
            'id', note_id
        ).eq('employee_id', request.user_id).execute()
        if not result.data:
            return error_response('Note not found or access denied', 404)
        return success_response(result.data[0], 'Note updated')
    except Exception as e:
        return error_response(f"Update Note Error: {str(e)}", 500)


@app.route('/api/notes/<note_id>', methods=['DELETE'])
@require_auth
def delete_lead_note(note_id):
    """مسح ملاحظة (المالك فقط)."""
    try:
        result = supabase.table('lead_notes').delete().eq(
            'id', note_id
        ).eq('employee_id', request.user_id).execute()
        if not result.data:
            return error_response('Note not found or access denied', 404)
        return success_response({'deleted': note_id}, 'Note deleted')
    except Exception as e:
        return error_response(f"Delete Note Error: {str(e)}", 500)


# ==================== LEAD HISTORY ====================

@app.route('/api/leads/<lead_id>/history', methods=['GET'])
@require_auth
def get_lead_history(lead_id):
    """
    تاريخ كامل لليد: تحويلات الحالة + آخر الأكشنز + الملاحظات.
    بنعتمد على RLS في Supabase للتحقق من الصلاحية.
    """
    try:
        statuses = supabase.table('lead_status_history').select(
            '*, employees(full_name)'
        ).eq('lead_id', lead_id).order('changed_at', desc=True).execute()

        actions = supabase.table('actions_log').select(
            '*, employees(full_name)'
        ).eq('lead_id', lead_id).order('created_at', desc=True).limit(100).execute()

        return success_response({
            'status_history': statuses.data,
            'actions':        actions.data,
        })
    except Exception as e:
        return error_response(f"History Error: {str(e)}", 500)


# ==================== MEETING ACTION ====================

@app.route('/api/actions/meeting', methods=['POST'])
@require_auth
def book_meeting():
    """حجز ميتنج مع ليد (+ XP كبير)."""
    data = request.json or {}
    lead_id = data.get('lead_id')
    meeting_at = data.get('meeting_at')
    notes = data.get('notes')

    if not lead_id:
        return error_response('lead_id required')

    try:
        # حرك الحالة لـ meeting_set لو ينفع (بدون كسر الـ state machine)
        supabase.rpc('change_lead_status', {
            'p_lead_id':     lead_id,
            'p_employee_id': request.user_id,
            'p_new_status':  'meeting_set',
            'p_reason':      'Meeting booked',
        }).execute()

        xp_config = XP_CONFIG['meeting_booked']
        xp_result = supabase.rpc('award_xp_and_coins', {
            'p_employee_id': request.user_id,
            'p_xp':          xp_config['xp'],
            'p_coins':       xp_config['coins'],
            'p_reason':      'حجز ميتنج',
            'p_ref_type':    'action',
        }).execute()

        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'lead_id':     lead_id,
            'action':      'meeting_booked',
            'xp_earned':   xp_config['xp'],
            'coins_earned': xp_config['coins'],
            'details': {'meeting_at': meeting_at, 'notes': notes},
        }).execute()

        return success_response({
            'xp_earned':    xp_config['xp'],
            'coins_earned': xp_config['coins'],
            'level_up':     (xp_result.data or {}).get('leveled_up', False),
        }, 'Meeting booked')
    except Exception as e:
        return error_response(f"Meeting Error: {str(e)}", 500)


# ==================== PROJECTS (CAMPAIGNS) ====================

@app.route('/api/projects', methods=['GET'])
@require_auth
def list_projects():
    """قايمة الحملات كلها."""
    try:
        result = supabase.table('projects').select('*').order(
            'created_at', desc=True
        ).execute()
        return success_response({'projects': result.data})
    except Exception as e:
        return error_response(f"Projects Error: {str(e)}", 500)


@app.route('/api/projects', methods=['POST'])
@check_role(['admin'])
@audit_log('project_create', target_type='project')
def create_project():
    """إنشاء حملة جديدة (Admin)."""
    data = request.json or {}
    name = (data.get('name') or '').strip()
    slug = (data.get('slug') or '').strip().lower()

    if not name or not slug:
        return error_response('name and slug are required')
    if not re.fullmatch(r'[a-z0-9_-]{2,40}', slug):
        return error_response('Invalid slug format')

    try:
        result = supabase.table('projects').insert({
            'name': name,
            'slug': slug,
        }).execute()
        return success_response(result.data[0], 'Project created', 201)
    except Exception as e:
        return error_response(f"Create Project Error: {str(e)}", 500)


@app.route('/api/projects/<project_id>', methods=['PATCH'])
@check_role(['admin'])
@audit_log('project_update', target_type='project')
def update_project(project_id):
    """تعديل حملة."""
    data = request.json or {}
    updates = {k: v for k, v in data.items() if k in {'name', 'slug'}}
    if not updates:
        return error_response('No fields to update')
    try:
        result = supabase.table('projects').update(updates).eq(
            'id', project_id
        ).execute()
        if not result.data:
            return error_response('Project not found', 404)
        return success_response(result.data[0], 'Project updated')
    except Exception as e:
        return error_response(f"Update Project Error: {str(e)}", 500)


@app.route('/api/projects/<project_id>/stats', methods=['GET'])
@require_auth
def project_stats(project_id):
    """إحصائيات حملة واحدة."""
    try:
        all_leads = supabase.table('leads').select(
            'status', count='exact'
        ).eq('project_id', project_id).execute()

        buckets = {}
        for row in (all_leads.data or []):
            buckets[row['status']] = buckets.get(row['status'], 0) + 1

        return success_response({
            'project_id':  project_id,
            'total_leads': all_leads.count,
            'by_status':   buckets,
        })
    except Exception as e:
        return error_response(f"Project Stats Error: {str(e)}", 500)


# ==================== TEAMS ====================

@app.route('/api/teams', methods=['GET'])
@require_auth
def list_teams():
    """قايمة الفرق."""
    try:
        result = supabase.table('teams').select(
            '*, manager:employees!teams_manager_id_fkey(full_name, avatar_url)'
        ).eq('is_active', True).order('total_xp', desc=True).execute()
        return success_response({'teams': result.data})
    except Exception as e:
        return error_response(f"Teams Error: {str(e)}", 500)


@app.route('/api/teams', methods=['POST'])
@check_role(['admin'])
@audit_log('team_create', target_type='team')
def create_team():
    """إنشاء فريق جديد."""
    data = request.json or {}
    name = (data.get('name') or '').strip()
    slug = (data.get('slug') or '').strip().lower()
    manager_id = data.get('manager_id')
    color = data.get('color', '#6366F1')

    if not name or not slug:
        return error_response('name and slug are required')

    try:
        result = supabase.table('teams').insert({
            'name':       name,
            'slug':       slug,
            'manager_id': manager_id,
            'color':      color,
        }).execute()
        return success_response(result.data[0], 'Team created', 201)
    except Exception as e:
        return error_response(f"Create Team Error: {str(e)}", 500)


@app.route('/api/teams/<team_id>/members', methods=['POST'])
@check_role(['admin', 'manager'])
def add_team_member(team_id):
    """ضيف موظف لفريق."""
    data = request.json or {}
    employee_id = data.get('employee_id')
    if not employee_id:
        return error_response('employee_id required')

    try:
        result = supabase.table('employees').update({
            'team_id': team_id,
        }).eq('id', employee_id).execute()
        if not result.data:
            return error_response('Employee not found', 404)
        return success_response(result.data[0], 'Member added')
    except Exception as e:
        return error_response(f"Add Member Error: {str(e)}", 500)


# ==================== MANAGER DASHBOARD ====================

@app.route('/api/manager/dashboard', methods=['GET'])
@check_role(['admin', 'manager'])
def manager_dashboard():
    """
    لوحة تحكم المدير: أداء الفريق التفصيلي.
    - المدير بيشوف فريقه فقط
    - الأدمن بيشوف الكل
    """
    try:
        # هات team بتاع المدير الحالي
        me = supabase.table('employees').select('role, team').eq(
            'id', request.user_id
        ).single().execute()

        emp_query = supabase.table('employees').select(
            'id, full_name, avatar_url, level, title, total_xp, current_xp,'
            ' syb_coins, total_deals, total_messages, total_calls,'
            ' daily_streak, last_action_at, team'
        ).eq('is_active', True)

        if me.data.get('role') == 'manager' and me.data.get('team'):
            emp_query = emp_query.eq('team', me.data['team'])

        members = emp_query.order('total_xp', desc=True).execute()

        today = datetime.utcnow().date()
        week_ago = (today - timedelta(days=7)).isoformat()
        member_ids = [m['id'] for m in (members.data or [])]

        # أكشنز آخر 7 أيام
        actions = []
        if member_ids:
            actions_res = supabase.table('actions_log').select(
                'employee_id, action, xp_earned, created_at'
            ).gte('created_at', week_ago).in_(
                'employee_id', member_ids
            ).execute()
            actions = actions_res.data or []

        # لخّص الأكشنز لكل موظف
        per_emp = {}
        for a in actions:
            bucket = per_emp.setdefault(a['employee_id'], {
                'messages': 0, 'calls': 0, 'deals': 0,
                'meetings': 0, 'xp_7d': 0,
            })
            bucket['xp_7d'] += a['xp_earned'] or 0
            if a['action'] == 'whatsapp_sent':
                bucket['messages'] += 1
            elif a['action'] == 'call_made':
                bucket['calls'] += 1
            elif a['action'] == 'deal_closed':
                bucket['deals'] += 1
            elif a['action'] == 'meeting_booked':
                bucket['meetings'] += 1

        enriched = []
        for m in (members.data or []):
            stats = per_emp.get(m['id'], {
                'messages': 0, 'calls': 0, 'deals': 0, 'meetings': 0, 'xp_7d': 0,
            })
            m['week_stats'] = stats
            enriched.append(m)

        return success_response({
            'members':   enriched,
            'summary': {
                'team_size':   len(enriched),
                'total_xp':    sum(m['total_xp'] for m in enriched),
                'total_deals': sum(m['total_deals'] for m in enriched),
                'xp_7d':       sum(m['week_stats']['xp_7d'] for m in enriched),
            },
        })
    except Exception as e:
        return error_response(f"Manager Dashboard Error: {str(e)}", 500)


# ==================== ACTIVITY FEED (Real-time via polling) ====================

@app.route('/api/feed', methods=['GET'])
@require_auth
def get_activity_feed():
    """
    آخر الأحداث في الشركة (closed deals, level ups, purchases...).
    يدعم pagination عن طريق `since` timestamp عشان polling خفيف.
    """
    try:
        since = request.args.get('since')
        limit = min(int(request.args.get('limit', 50)), 200)

        query = supabase.table('activity_feed').select(
            '*, employees(full_name, avatar_url, level, title)'
        ).eq('is_public', True).order('created_at', desc=True).limit(limit)

        if since:
            query = query.gt('created_at', since)

        result = query.execute()
        return success_response({
            'events': result.data,
            'count':  len(result.data or []),
        })
    except Exception as e:
        return error_response(f"Feed Error: {str(e)}", 500)


# ==================== COMPETITIONS ====================

@app.route('/api/competitions', methods=['GET'])
@require_auth
def list_competitions():
    """كل المسابقات مع ترتيب المشاركين."""
    try:
        status = request.args.get('status', 'active')
        comps = supabase.table('competitions').select('*').eq(
            'status', status
        ).order('starts_at', desc=True).execute()
        return success_response({'competitions': comps.data})
    except Exception as e:
        return error_response(f"Competitions Error: {str(e)}", 500)


@app.route('/api/competitions', methods=['POST'])
@check_role(['admin'])
@audit_log('competition_create', target_type='competition')
def create_competition():
    """إنشاء مسابقة جديدة."""
    data = request.json or {}
    required = {'title', 'starts_at', 'ends_at'}
    if not required.issubset(data.keys()):
        return error_response(f'Missing fields: {required - set(data.keys())}')

    try:
        result = supabase.table('competitions').insert({
            'title':            data['title'],
            'description':      data.get('description'),
            'competition_type': data.get('competition_type', 'team_vs_team'),
            'metric':           data.get('metric', 'total_xp'),
            'starts_at':        data['starts_at'],
            'ends_at':          data['ends_at'],
            'prize_xp':         data.get('prize_xp', 0),
            'prize_coins':      data.get('prize_coins', 0),
        }).execute()
        return success_response(result.data[0], 'Competition created', 201)
    except Exception as e:
        return error_response(f"Create Competition Error: {str(e)}", 500)


@app.route('/api/competitions/<comp_id>/join', methods=['POST'])
@require_auth
def join_competition(comp_id):
    """اشترك في مسابقة كفرد أو كفريق."""
    data = request.json or {}
    team_id = data.get('team_id')

    try:
        exists = supabase.table('competition_participants').select('id').eq(
            'competition_id', comp_id
        ).or_(
            f'employee_id.eq.{request.user_id}' +
            (f',team_id.eq.{team_id}' if team_id else '')
        ).execute()
        if exists.data:
            return error_response('Already joined', 409)

        result = supabase.table('competition_participants').insert({
            'competition_id': comp_id,
            'employee_id':    None if team_id else request.user_id,
            'team_id':        team_id,
        }).execute()
        return success_response(result.data[0], 'Joined')
    except Exception as e:
        return error_response(f"Join Error: {str(e)}", 500)


@app.route('/api/competitions/<comp_id>/leaderboard', methods=['GET'])
@require_auth
def competition_leaderboard(comp_id):
    """ترتيب مسابقة — بيحدّث السكور قبل الرد."""
    try:
        supabase.rpc('refresh_competition_scores', {
            'p_comp_id': comp_id,
        }).execute()

        result = supabase.table('competition_participants').select(
            '*, employees(full_name, avatar_url, level), teams(name, color)'
        ).eq('competition_id', comp_id).order('rank').execute()

        return success_response({'standings': result.data})
    except Exception as e:
        return error_response(f"Competition Leaderboard Error: {str(e)}", 500)


@app.route('/api/competitions/<comp_id>/finalize', methods=['POST'])
@check_role(['admin'])
@audit_log('competition_finalize', target_type='competition')
def finalize_comp(comp_id):
    """اقفل المسابقة ووزع الجوايز."""
    try:
        result = supabase.rpc('finalize_competition', {
            'p_comp_id': comp_id,
        }).execute()
        payload = result.data or {}
        if payload.get('error'):
            return error_response(payload, 400)
        return success_response(payload, 'Competition finalized')
    except Exception as e:
        return error_response(f"Finalize Error: {str(e)}", 500)


# ==================== ADMIN: QUESTS GENERATION ====================

@app.route('/api/admin/quests/generate-daily', methods=['POST'])
@check_role(['admin'])
@audit_log('quests_generate_daily')
def trigger_daily_quests():
    """شغّل generate_daily_quests يدوياً (أو من cron)."""
    try:
        result = supabase.rpc('generate_daily_quests').execute()
        return success_response({'created': result.data}, 'Daily quests generated')
    except Exception as e:
        return error_response(f"Quest Generation Error: {str(e)}", 500)


@app.route('/api/admin/audit-log', methods=['GET'])
@check_role(['admin'])
def get_audit_log():
    """آخر أكشنز الأدمن (audit trail)."""
    try:
        limit = min(int(request.args.get('limit', 100)), 500)
        action = request.args.get('action')
        query = supabase.table('admin_audit_log').select(
            '*, admin:employees!admin_audit_log_admin_id_fkey(full_name, email)'
        ).order('created_at', desc=True).limit(limit)
        if action:
            query = query.eq('action', action)
        result = query.execute()
        return success_response({'entries': result.data})
    except Exception as e:
        return error_response(f"Audit Log Error: {str(e)}", 500)


@app.route('/api/leads/<lead_id>/assign', methods=['POST'])
@check_role(['admin'])
@audit_log('lead_reassign', target_type='lead')
def assign_single_lead(lead_id):
    """تعيين ليد لموظف معين (يدوي)."""
    data = request.json or {}
    employee_id = data.get('employee_id')
    if not employee_id:
        return error_response('employee_id required')
    try:
        result = supabase.rpc('assign_lead_to', {
            'p_lead_id':     lead_id,
            'p_employee_id': employee_id,
            'p_admin_id':    request.user_id,
        }).execute()
        payload = result.data or {}
        if payload.get('error'):
            return error_response(payload, 404)
        return success_response(payload, 'Lead reassigned')
    except Exception as e:
        return error_response(f"Assign Error: {str(e)}", 500)


@app.route('/api/admin/leads/export', methods=['GET'])
@check_role(['admin'])
@audit_log('leads_export', target_type='lead')
def export_leads_csv():
    """تصدير الليدز إلى CSV مع الفلاتر الحالية."""
    try:
        status = request.args.get('status')
        project_id = request.args.get('project_id')

        query = supabase.table('leads').select(
            'id, name, phone, status, quality, contact_count,'
            ' last_contact_at, created_at, assigned_to,'
            ' projects(name)'
        ).eq('is_active', True).limit(10000)
        if status:
            query = query.eq('status', status)
        if project_id:
            query = query.eq('project_id', project_id)

        rows = query.execute().data or []

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            'id', 'name', 'phone', 'status', 'quality',
            'project', 'contact_count', 'last_contact_at',
            'created_at', 'assigned_to',
        ])
        for r in rows:
            writer.writerow([
                r.get('id', ''),
                r.get('name') or '',
                r.get('phone') or '',
                r.get('status') or '',
                r.get('quality') or '',
                (r.get('projects') or {}).get('name', ''),
                r.get('contact_count') or 0,
                r.get('last_contact_at') or '',
                r.get('created_at') or '',
                r.get('assigned_to') or '',
            ])

        csv_data = buf.getvalue()
        buf.close()
        filename = f"leads_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            '\ufeff' + csv_data,  # BOM for Excel Arabic support
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        return error_response(f"Export Error: {str(e)}", 500)


@app.route('/api/admin/leaderboard/refresh', methods=['POST'])
@check_role(['admin'])
@audit_log('leaderboard_refresh')
def refresh_leaderboard_period():
    """حدّث الـ leaderboard لفترة معينة (daily/weekly/monthly/all_time)."""
    period = (request.json or {}).get('period', 'weekly')
    if period not in {'daily', 'weekly', 'monthly', 'all_time'}:
        return error_response('Invalid period')
    try:
        result = supabase.rpc('refresh_leaderboard', {
            'p_period': period,
        }).execute()
        return success_response({
            'period': period,
            'rows':   result.data,
        }, 'Leaderboard refreshed')
    except Exception as e:
        return error_response(f"Leaderboard Refresh Error: {str(e)}", 500)


# ==================== WHATSAPP INBOUND WEBHOOK ====================

def _is_bot_conversation_active(phone: str) -> bool:
    """Check if there's an active bot conversation (not yet handed off) for this phone."""
    try:
        r = supabase.table('bot_conversations').select('state').eq(
            'whatsapp_number', phone
        ).order('created_at', desc=True).limit(1).execute()
        if not r.data:
            return False
        return (r.data[0].get('state') or '') == 'active'
    except Exception:
        return False


def _auto_create_lead_from_whatsapp(phone_clean: str, msg: dict) -> dict | None:
    """
    إنشاء ليد جديد تلقائي لما رقم مش موجود يبعت رسالة واتساب.
    بيكتشف لو جاي من إعلان (Click-to-WhatsApp) ويربطه بالحملة.
    """
    try:
        referral = msg.get('referral') or {}
        profile_name = msg.get('profile_name') or None

        # تحديد المصدر والحملة
        source = 'whatsapp'
        ad_campaign_id = None
        utm_source = None
        utm_medium = None
        utm_campaign = None
        quality = 'warm'  # جاي من واتساب = مهتم

        # لو جاي من إعلان (Click-to-WhatsApp ad)
        if referral.get('source_type') == 'ad' or referral.get('source_id'):
            source = 'instagram_click_to_whatsapp'
            utm_source = 'instagram'
            utm_medium = 'paid'
            utm_campaign = referral.get('headline') or referral.get('body') or None
            quality = 'hot'  # جاي من إعلان وبادر بالتواصل = ساخن

            # حاول تلاقي الحملة من platform_campaign_id أو من الاسم
            source_id = referral.get('source_id', '')
            if source_id:
                camp = supabase.table('ad_campaigns').select('id').eq(
                    'platform_campaign_id', source_id
                ).limit(1).execute()
                if camp.data:
                    ad_campaign_id = camp.data[0]['id']

            # لو ما لقيناش بالـ ID، جرب بالـ headline
            if not ad_campaign_id and referral.get('headline'):
                camp = supabase.table('ad_campaigns').select('id').ilike(
                    'name', f"%{referral['headline'][:50]}%"
                ).limit(1).execute()
                if camp.data:
                    ad_campaign_id = camp.data[0]['id']

            # لو لسه ما لقيناش، جرب أي حملة Instagram نشطة
            if not ad_campaign_id:
                camp = supabase.table('ad_campaigns').select('id').eq(
                    'platform', 'instagram'
                ).eq('status', 'active').order('created_at', desc=True).limit(1).execute()
                if camp.data:
                    ad_campaign_id = camp.data[0]['id']

        # إنشاء الليد
        row = {
            'phone': phone_clean,
            'phone_clean': phone_clean,
            'name': profile_name,
            'source': source,
            'status': 'new',
            'quality': quality,
            'ad_campaign_id': ad_campaign_id,
            'utm_source': utm_source,
            'utm_medium': utm_medium,
            'utm_campaign': utm_campaign,
            'imported_from': 'whatsapp_auto',
            'last_contact_at': datetime.utcnow().isoformat(),
            'contact_count': 0,
        }

        result = supabase.table('leads').insert(row).execute()
        if not result.data:
            return None

        new_lead = result.data[0]

        # تحديث عداد الحملة لو مربوط
        if ad_campaign_id:
            supabase.table('ad_campaigns').update({
                'webhook_leads_count': supabase.table('ad_campaigns').select(
                    'webhook_leads_count'
                ).eq('id', ad_campaign_id).single().execute().data.get('webhook_leads_count', 0) + 1,
                'updated_at': datetime.utcnow().isoformat(),
            }).eq('id', ad_campaign_id).execute()

        # التوزيع التلقائي — round-robin على الإيجنت النشطين
        _auto_assign_lead(new_lead['id'])

        # Outbound webhook
        fire_webhooks('lead.created', new_lead)

        return new_lead
    except Exception:
        return None


@app.route('/api/webhooks/whatsapp', methods=['POST'])
def whatsapp_inbound():
    """
    يستقبل الرسايل الجاية من WhatsApp Business API.
    بيربطها بالليد عن طريق phone_clean.
    لو الرقم مش موجود — بينشئ ليد جديد تلقائي.
    لو جاي من إعلان (Click-to-WhatsApp) — بيربطه بالحملة.
    محمي بـ HMAC signature verification (Meta standard).
    """
    # 🔐 تحقق من التوقيع قبل أي حاجة
    if WHATSAPP_APP_SECRET:
        sig = request.headers.get('X-Hub-Signature-256', '')
        if not verify_whatsapp_signature(request.get_data(), sig):
            return error_response('Invalid signature', 401)

    try:
        payload = request.json or {}
        # يدعم صيغة Meta الأصلية (entry[].changes[]) والصيغة المبسطة
        if 'entry' in payload:
            msgs = whatsapp_client.parse_webhook(payload)
        else:
            msgs = [{
                'from':         (payload.get('from') or '').lstrip('+'),
                'text':         payload.get('text'),
                'message_id':   payload.get('message_id'),
                'type':         payload.get('type', 'text'),
                'profile_name': payload.get('profile_name'),
                'referral':     payload.get('referral'),
            }]

        processed = 0
        new_leads = 0
        for m in msgs:
            from_phone = (m.get('from') or '').strip()
            text = m.get('text') or ''
            if not from_phone:
                continue

            is_new = False
            lead = supabase.table('leads').select(
                'id, assigned_to, contact_count'
            ).eq('phone_clean', from_phone).eq('is_active', True).limit(1).execute()

            # ===== AUTO-CREATE LEAD لو الرقم مش موجود =====
            if not lead.data:
                new_lead = _auto_create_lead_from_whatsapp(from_phone, m)
                if not new_lead:
                    continue
                lead_data = [new_lead]
                new_leads += 1
                is_new = True
            else:
                lead_data = lead.data

            target = lead_data[0]

            supabase.table('whatsapp_messages').insert({
                'lead_id':       target['id'],
                'employee_id':   target.get('assigned_to'),
                'direction':     'inbound',
                'message_text':  text,
                'message_type':  m.get('type', 'text'),
                'wa_message_id': m.get('message_id'),
                'wa_status':     'received',
            }).execute()

            supabase.table('leads').update({
                'last_contact_at': datetime.utcnow().isoformat(),
                'contact_count':   (target.get('contact_count') or 0) + 1,
            }).eq('id', target['id']).execute()

            # mark as read in Meta
            whatsapp_client.mark_as_read(m.get('message_id'))

            # 🤖 BOT HANDLING — لو الليد جديد أو مفيش موظف مكلف بيه أو لسه البوت بيتعامل معاه
            try:
                from bot_agents import handle_incoming_message
                bot_active = _is_bot_conversation_active(from_phone)
                if (not target.get('assigned_to')) or is_new or bot_active:
                    bot_result = handle_incoming_message(supabase, from_phone, text)
                    if bot_result.get('reply'):
                        whatsapp_client.send_message(from_phone, bot_result['reply'])
                    # If bot handed off, assign the lead to the chosen human
                    if bot_result.get('handoff_to') and target.get('id'):
                        supabase.table('leads').update({
                            'assigned_to': bot_result['handoff_to'],
                        }).eq('id', target['id']).execute()
                        target['assigned_to'] = bot_result['handoff_to']
                    # Send notification to the human about the handoff
                    notify = bot_result.get('notify')
                    if notify and notify.get('employee_id'):
                        send_push_to_employee(
                            notify['employee_id'],
                            notify.get('title', 'إشعار جديد'),
                            notify.get('body', ''),
                            f"/agent#lead-{target['id']}" if target.get('id') else '/agent',
                        )
            except Exception as bot_err:
                logger.error('Bot handler error: %s', bot_err, exc_info=True)

            # 🎮 XP للموظف لأن العميل رد عليه
            if target.get('assigned_to'):
                xp_config = XP_CONFIG['whatsapp_received']
                supabase.rpc('award_xp_and_coins', {
                    'p_employee_id': target['assigned_to'],
                    'p_xp':          xp_config['xp'],
                    'p_coins':       xp_config['coins'],
                    'p_reason':      'العميل رد على رسالة',
                    'p_ref_type':    'action',
                }).execute()
                supabase.table('actions_log').insert({
                    'employee_id': target['assigned_to'],
                    'lead_id':     target['id'],
                    'action':      'whatsapp_received',
                    'xp_earned':   xp_config['xp'],
                }).execute()
            processed += 1

            # 📱 Push Notification للموظف
            if target.get('assigned_to'):
                send_push_to_employee(
                    target['assigned_to'],
                    f"📩 رسالة من {target.get('name') or from_phone}",
                    text[:100] if text else 'رسالة جديدة',
                    f"/agent#lead-{target['id']}",
                )

            # Outbound webhook event
            fire_webhooks('whatsapp.inbound', {
                'lead_id': target['id'], 'from': from_phone,
                'text': text, 'is_new_lead': is_new,
            })

        return success_response({
            'processed': processed,
            'new_leads_created': new_leads,
        }, 'Webhook handled')
    except Exception as e:
        return error_response(f"Webhook Error: {str(e)}", 500)


# Meta GET verification (initial webhook setup)
@app.route('/api/webhooks/whatsapp', methods=['GET'])
def whatsapp_verify():
    """Meta بيبعت GET request أول مرة للتحقق من الـ webhook."""
    verify_token = os.getenv('WHATSAPP_VERIFY_TOKEN', '') or FB_VERIFY_TOKEN
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    if mode == 'subscribe' and token == verify_token and verify_token:
        return challenge or '', 200
    return 'forbidden', 403


# =====================================================================
# PHASE 9 — BUSINESS FEATURES
#   * Password reset + 2FA
#   * Granular permissions
#   * Data retention (archive + GDPR delete)
#   * Stripe billing webhook (scaffold)
# =====================================================================

# ---------- 9.0  permission helpers ----------------------------------

def has_permission(user_id: str, perm_code: str) -> bool:
    """Check role→permission via the role_permissions table."""
    try:
        emp = supabase.table('employees').select('role').eq('id', user_id).single().execute()
        role = (emp.data or {}).get('role')
        if not role:
            return False
        row = supabase.rpc('has_permission', {'p_role': role, 'p_permission': perm_code}).execute()
        return bool(row.data)
    except Exception:
        return False


def require_permission(perm_code):
    """Decorator: 403 unless the caller's role holds the named permission."""
    def decorator(f):
        @wraps(f)
        @require_auth
        def wrapped(*args, **kwargs):
            if not has_permission(request.user_id, perm_code):
                return error_response(f'Missing permission: {perm_code}', 403)
            return f(*args, **kwargs)
        return wrapped
    return decorator


@app.route('/api/me/permissions', methods=['GET'])
@require_auth
def my_permissions():
    try:
        emp = supabase.table('employees').select('role').eq('id', request.user_id).single().execute()
        role = (emp.data or {}).get('role', 'agent')
        perms = supabase.table('role_permissions').select('permission').eq('role', role).execute()
        return success_response({
            'role': role,
            'permissions': [p['permission'] for p in (perms.data or [])],
        })
    except Exception as e:
        return error_response(str(e), 500)


# ---------- 9.1  password reset -------------------------------------

@app.route('/api/auth/password/reset-request', methods=['POST'])
def password_reset_request():
    """Kicks off Supabase's email-reset flow. Always returns success to avoid email enumeration."""
    data = request.json or {}
    email = (data.get('email') or '').strip().lower()
    if not email:
        return error_response('email required', 400)
    try:
        redirect_to = os.getenv('PASSWORD_RESET_REDIRECT', '')
        supabase.auth.reset_password_for_email(
            email,
            {'redirect_to': redirect_to} if redirect_to else None,
        )
    except Exception:
        pass  # swallow — do not leak whether the email exists
    return success_response({'sent': True}, 'If that email exists, a reset link was sent')


@app.route('/api/auth/password/update', methods=['POST'])
@require_auth
def password_update():
    data = request.json or {}
    new_password = data.get('new_password') or ''
    if len(new_password) < 8:
        return error_response('password must be at least 8 characters', 400)
    try:
        supabase.auth.update_user({'password': new_password})
        return success_response({'updated': True})
    except Exception as e:
        return error_response(str(e), 400)


# ---------- 9.2  MFA / 2FA ------------------------------------------

def _log_mfa(event, factor_type=None):
    with contextlib.suppress(Exception):
        supabase.table('mfa_audit').insert({
            'user_id': request.user_id,
            'event': event,
            'factor_type': factor_type,
            'ip': request.remote_addr,
            'user_agent': request.headers.get('User-Agent', '')[:255],
        }).execute()


@app.route('/api/auth/mfa/enroll', methods=['POST'])
@require_auth
def mfa_enroll():
    """Returns TOTP secret + QR code URL. Caller then verifies with /mfa/verify."""
    try:
        body = request.json or {}
        friendly = body.get('friendly_name', 'AlSaeb CRM')
        resp = supabase.auth.mfa.enroll({'factor_type': 'totp', 'friendly_name': friendly})
        _log_mfa('enroll', 'totp')
        # resp.data: { id, type, totp: { qr_code, secret, uri } }
        return success_response(getattr(resp, 'data', resp))
    except Exception as e:
        return error_response(f'MFA enroll failed: {e}', 400)


@app.route('/api/auth/mfa/challenge', methods=['POST'])
@require_auth
def mfa_challenge():
    body = request.json or {}
    factor_id = body.get('factor_id')
    if not factor_id:
        return error_response('factor_id required', 400)
    try:
        resp = supabase.auth.mfa.challenge({'factor_id': factor_id})
        return success_response(getattr(resp, 'data', resp))
    except Exception as e:
        _log_mfa('challenge_failed', 'totp')
        return error_response(str(e), 400)


@app.route('/api/auth/mfa/verify', methods=['POST'])
@require_auth
def mfa_verify():
    body = request.json or {}
    factor_id    = body.get('factor_id')
    challenge_id = body.get('challenge_id')
    code         = body.get('code')
    if not all([factor_id, challenge_id, code]):
        return error_response('factor_id, challenge_id, code all required', 400)
    try:
        resp = supabase.auth.mfa.verify({
            'factor_id': factor_id, 'challenge_id': challenge_id, 'code': code,
        })
        _log_mfa('verify', 'totp')
        return success_response(getattr(resp, 'data', resp))
    except Exception as e:
        _log_mfa('challenge_failed', 'totp')
        return error_response(str(e), 400)


@app.route('/api/auth/mfa/unenroll', methods=['POST'])
@require_auth
def mfa_unenroll():
    body = request.json or {}
    factor_id = body.get('factor_id')
    if not factor_id:
        return error_response('factor_id required', 400)
    try:
        supabase.auth.mfa.unenroll({'factor_id': factor_id})
        _log_mfa('unenroll', 'totp')
        return success_response({'unenrolled': True})
    except Exception as e:
        return error_response(str(e), 400)


# ---------- 9.3  data retention: archive + GDPR ---------------------

@app.route('/api/admin/leads/<lead_id>/archive', methods=['POST'])
@require_permission('leads.archive')
@audit_log('lead_archive', 'lead')
def archive_lead(lead_id):
    try:
        supabase.table('leads').update({
            'archived_at': datetime.utcnow().isoformat(),
            'is_active': False,
        }).eq('id', lead_id).execute()
        return success_response({'archived': True, 'lead_id': lead_id})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/retention/run', methods=['POST'])
@require_permission('leads.archive')
@audit_log('retention_run')
def run_retention():
    """Auto-archive closed_won/lost leads older than N days."""
    days = int((request.json or {}).get('days', 365))
    try:
        resp = supabase.rpc('auto_archive_stale_leads', {'p_days': days}).execute()
        return success_response({'archived_count': resp.data, 'days': days})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/admin/users/<user_id>/gdpr-delete', methods=['POST'])
@require_permission('gdpr.delete')
@audit_log('gdpr_delete', 'user')
def gdpr_delete_user(user_id):
    """Right-to-delete: anonymise + revoke. Cannot be undone."""
    confirm = (request.json or {}).get('confirm')
    if confirm != 'DELETE':
        return error_response('confirm field must equal "DELETE"', 400)
    try:
        resp = supabase.rpc('gdpr_delete_employee', {'p_user_id': user_id}).execute()
        payload = resp.data or {}
        if not payload.get('success'):
            return error_response(payload.get('error', 'gdpr_failed'), 400)
        return success_response(payload)
    except Exception as e:
        return error_response(str(e), 500)


# ---------- 9.4  billing — Stripe webhook scaffold ------------------

def _verify_stripe_signature(raw_body: bytes, sig_header: str) -> bool:
    """Stripe's t=...,v1=... HMAC-SHA256 scheme. Skips if no secret configured."""
    secret = os.getenv('STRIPE_WEBHOOK_SECRET', '')
    if not secret:
        return True  # dev mode: accept unsigned
    if not sig_header:
        return False
    try:
        parts = dict(p.split('=', 1) for p in sig_header.split(','))
        timestamp, v1 = parts.get('t'), parts.get('v1')
        if not timestamp or not v1:
            return False
        signed = f'{timestamp}.{raw_body.decode("utf-8", "ignore")}'.encode()
        expected = hmac.new(secret.encode(), signed, hashlib.sha256).hexdigest()
        return hmac.compare_digest(expected, v1)
    except Exception:
        return False


@app.route('/api/billing/webhooks/stripe', methods=['POST'])
def stripe_webhook():
    raw = request.get_data()
    sig = request.headers.get('Stripe-Signature', '')
    if not _verify_stripe_signature(raw, sig):
        return error_response('bad signature', 400)
    try:
        event = request.json or {}
        event_id = event.get('id')
        etype    = event.get('type', '')
        if not event_id:
            return error_response('missing event id', 400)

        # idempotency: skip if we've already processed this event
        existing = supabase.table('stripe_events').select('id').eq('id', event_id).execute()
        if existing.data:
            return success_response({'status': 'duplicate', 'id': event_id})

        supabase.table('stripe_events').insert({
            'id': event_id, 'type': etype, 'payload': event,
        }).execute()

        obj = (event.get('data') or {}).get('object', {}) or {}

        # minimal event handlers — expand as billing needs grow
        if etype in ('customer.subscription.created', 'customer.subscription.updated'):
            org_id = (obj.get('metadata') or {}).get('organization_id')
            if org_id:
                supabase.table('subscriptions').upsert({
                    'organization_id':        org_id,
                    'stripe_customer_id':     obj.get('customer'),
                    'stripe_subscription_id': obj.get('id'),
                    'plan':   ((obj.get('items', {}).get('data') or [{}])[0]
                               .get('price', {}).get('lookup_key') or 'free'),
                    'status': obj.get('status', 'active'),
                    'cancel_at_period_end': obj.get('cancel_at_period_end', False),
                    'current_period_end':   (
                        datetime.utcfromtimestamp(obj['current_period_end']).isoformat()
                        if obj.get('current_period_end') else None
                    ),
                }, on_conflict='stripe_subscription_id').execute()

        elif etype == 'customer.subscription.deleted':
            sub_id = obj.get('id')
            if sub_id:
                supabase.table('subscriptions').update({
                    'status': 'canceled',
                }).eq('stripe_subscription_id', sub_id).execute()

        supabase.table('stripe_events').update({
            'processed_at': datetime.utcnow().isoformat(),
        }).eq('id', event_id).execute()

        return success_response({'received': True, 'type': etype})
    except Exception as e:
        return error_response(f'stripe webhook error: {e}', 500)


@app.route('/api/billing/subscription', methods=['GET'])
@require_permission('billing.view')
def get_subscription():
    try:
        emp = supabase.table('employees').select('organization_id').eq(
            'id', request.user_id).single().execute()
        org_id = (emp.data or {}).get('organization_id')
        if not org_id:
            return success_response({'subscription': None})
        sub = supabase.table('subscriptions').select('*').eq(
            'organization_id', org_id).order('created_at', desc=True).limit(1).execute()
        return success_response({'subscription': (sub.data or [None])[0]})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/billing/checkout', methods=['POST'])
@check_role(['admin'])
def billing_checkout():
    """Create Stripe Checkout session for plan upgrade."""
    stripe_key = os.getenv('STRIPE_SECRET_KEY')
    if not stripe_key:
        return success_response({'url': None, 'message': 'Stripe not configured'})
    try:
        import stripe
        stripe.api_key = stripe_key
        data = request.json or {}
        plan = data.get('plan', 'pro')
        price_lookup = {'pro': os.getenv('STRIPE_PRICE_PRO', 'price_pro'),
                        'enterprise': os.getenv('STRIPE_PRICE_ENTERPRISE', 'price_enterprise')}
        emp = supabase.table('employees').select('organization_id, email').eq(
            'id', request.user_id).single().execute()
        org_id = (emp.data or {}).get('organization_id', request.user_id)
        session = stripe.checkout.Session.create(
            mode='subscription',
            line_items=[{'price': price_lookup.get(plan, plan), 'quantity': 1}],
            success_url=request.host_url + 'admin?billing=success',
            cancel_url=request.host_url + 'admin?billing=cancel',
            metadata={'organization_id': org_id},
            customer_email=(emp.data or {}).get('email'),
        )
        return success_response({'url': session.url})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/billing/portal', methods=['POST'])
@check_role(['admin'])
def billing_portal():
    """Create Stripe Customer Portal session for self-serve management."""
    stripe_key = os.getenv('STRIPE_SECRET_KEY')
    if not stripe_key:
        return success_response({'url': None, 'message': 'Stripe not configured'})
    try:
        import stripe
        stripe.api_key = stripe_key
        emp = supabase.table('employees').select('organization_id').eq(
            'id', request.user_id).single().execute()
        org_id = (emp.data or {}).get('organization_id')
        if not org_id:
            return error_response('No organization found', 400)
        sub = supabase.table('subscriptions').select('stripe_customer_id').eq(
            'organization_id', org_id).order('created_at', desc=True).limit(1).execute()
        cust_id = (sub.data[0] if sub.data else {}).get('stripe_customer_id')
        if not cust_id:
            return error_response('No Stripe customer found', 400)
        session = stripe.billing_portal.Session.create(
            customer=cust_id,
            return_url=request.host_url + 'admin?view=billing',
        )
        return success_response({'url': session.url})
    except Exception as e:
        return error_response(str(e), 500)


# ==================== LISTINGS ====================

@app.route('/api/listings', methods=['GET'])
@require_auth
def get_listings():
    try:
        per_page = min(int(request.args.get('per_page', 50)), 200)
        listing_type = request.args.get('listing_type')
        status = request.args.get('status')
        search = request.args.get('search', '').strip()

        q = supabase.table('listings').select('*, owners(name, phone), agents:employees!listings_agent_id_fkey(full_name)').order('created_at', desc=True).limit(per_page)
        if listing_type:
            q = q.eq('listing_type', listing_type)
        if status:
            q = q.eq('status', status)
        if search:
            q = q.or_(f"title.ilike.%{search}%,location.ilike.%{search}%,building_name.ilike.%{search}%,reference_number.ilike.%{search}%")
        data = q.execute()
        return success_response({'listings': data.data})
    except Exception:
        return success_response({'listings': []})


@app.route('/api/listings', methods=['POST'])
@check_role(['admin', 'manager', 'agent'])
def create_listing():
    try:
        body = request.get_json()
        if not body or not body.get('title'):
            return error_response('Title is required')
        row = {
            'title':         body['title'],
            'description':   body.get('description'),
            'property_type': body.get('property_type', 'apartment'),
            'listing_type':  body.get('listing_type', 'sale'),
            'price':         body.get('price'),
            'area_sqft':     body.get('area_sqft'),
            'bedrooms':      body.get('bedrooms'),
            'bathrooms':     body.get('bathrooms'),
            'location':      body.get('location'),
            'building_name': body.get('building_name'),
            'owner_id':      body.get('owner_id'),
            'agent_id':      request.user_id,
            'permit_number': body.get('permit_number'),
            'lat':           body.get('lat'),
            'lng':           body.get('lng'),
        }
        data = supabase.table('listings').insert(row).execute()
        return success_response({'listing': data.data[0] if data.data else {}}, 'Listing created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/listings/<listing_id>', methods=['PATCH'])
@check_role(['admin', 'manager', 'agent'])
def update_listing(listing_id):
    try:
        body = request.get_json()
        allowed = ['title', 'description', 'property_type', 'listing_type', 'price',
                   'area_sqft', 'bedrooms', 'bathrooms', 'location', 'building_name',
                   'owner_id', 'status', 'permit_number', 'lat', 'lng', 'features', 'images']
        updates = {k: v for k, v in body.items() if k in allowed}
        data = supabase.table('listings').update(updates).eq('id', listing_id).execute()
        return success_response({'listing': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


# ==================== OWNERS ====================

@app.route('/api/owners', methods=['GET'])
@require_auth
def get_owners():
    try:
        search = request.args.get('search', '').strip()
        q = supabase.table('owners').select('*, listings(count)').order('created_at', desc=True).limit(100)
        if search:
            q = q.or_(f"name.ilike.%{search}%,phone.ilike.%{search}%,email.ilike.%{search}%")
        data = q.execute()
        owners = data.data or []
        for o in owners:
            o['listings_count'] = o.get('listings', [{}])[0].get('count', 0) if o.get('listings') else 0
        return success_response({'owners': owners})
    except Exception:
        return success_response({'owners': []})


@app.route('/api/owners', methods=['POST'])
@check_role(['admin', 'manager', 'agent'])
def create_owner():
    try:
        body = request.get_json()
        if not body or not body.get('name'):
            return error_response('Name is required')
        row = {
            'name':         body['name'],
            'phone':        body.get('phone'),
            'email':        body.get('email'),
            'type':         body.get('type', 'individual'),
            'company_name': body.get('company_name'),
            'notes':        body.get('notes'),
            'agent_id':     request.user_id,
        }
        data = supabase.table('owners').insert(row).execute()
        return success_response({'owner': data.data[0] if data.data else {}}, 'Owner created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/owners/<owner_id>', methods=['PATCH'])
@check_role(['admin', 'manager'])
def update_owner(owner_id):
    try:
        body = request.get_json()
        allowed = ['name', 'phone', 'email', 'type', 'company_name', 'nationality', 'notes']
        updates = {k: v for k, v in body.items() if k in allowed}
        data = supabase.table('owners').update(updates).eq('id', owner_id).execute()
        return success_response({'owner': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/owners/import', methods=['POST'])
@check_role(['admin', 'manager'])
def import_owners():
    """استيراد ملاك من ملف Excel (.xlsx)
    الأعمدة المتوقعة: الاسم | الموبايل | الإيميل | النوع (فرد/شركة) | الجنسية | اسم الشركة | نوع العقار | ملاحظات
    """
    import openpyxl

    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        return error_response('يرجى رفع ملف Excel (.xlsx)', 400)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
        ws = wb.active

        rows_list = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows_list) < 2:
            return error_response('الملف فاضي — لازم يكون فيه صف عناوين + بيانات', 400)

        # Map Arabic/English header names to DB fields
        header_map = {
            'الاسم': 'name', 'name': 'name', 'اسم المالك': 'name', 'owner name': 'name',
            'الموبايل': 'phone', 'phone': 'phone', 'رقم الموبايل': 'phone', 'الهاتف': 'phone', 'mobile': 'phone',
            'الإيميل': 'email', 'email': 'email', 'البريد': 'email', 'الايميل': 'email',
            'النوع': 'type', 'type': 'type', 'نوع المالك': 'type',
            'الجنسية': 'nationality', 'nationality': 'nationality',
            'اسم الشركة': 'company_name', 'company': 'company_name', 'company name': 'company_name', 'الشركة': 'company_name',
            'نوع العقار': 'property_type', 'property type': 'property_type', 'نوع الملكية': 'property_type',
            'ملاحظات': 'notes', 'notes': 'notes',
        }

        # Parse header row
        raw_headers = [str(h).strip().lower() if h else '' for h in rows_list[0]]
        col_map = {}
        for i, h in enumerate(raw_headers):
            if h in header_map:
                col_map[i] = header_map[h]

        if 'name' not in col_map.values():
            return error_response('الملف لازم يحتوي على عمود "الاسم" أو "name"', 400)

        # Type value normalization
        type_map = {
            'فرد': 'individual', 'individual': 'individual', 'شخص': 'individual',
            'شركة': 'company', 'company': 'company',
        }

        created = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(rows_list[1:], start=2):
            parsed = {}
            for col_idx, field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    parsed[field] = str(val).strip()

            name = parsed.get('name', '').strip()
            if not name:
                skipped += 1
                continue

            # Normalize type
            raw_type = parsed.get('type', '').strip().lower()
            owner_type = type_map.get(raw_type, 'individual')

            # Build notes — include property_type in notes if provided
            notes_parts = []
            if parsed.get('property_type'):
                notes_parts.append(f"نوع العقار: {parsed['property_type']}")
            if parsed.get('notes'):
                notes_parts.append(parsed['notes'])

            owner_row = {
                'name':         name,
                'phone':        parsed.get('phone', ''),
                'email':        parsed.get('email', ''),
                'type':         owner_type,
                'nationality':  parsed.get('nationality', ''),
                'company_name': parsed.get('company_name', ''),
                'notes':        ' | '.join(notes_parts) if notes_parts else '',
                'agent_id':     request.user_id,
            }

            try:
                supabase.table('owners').insert(owner_row).execute()
                created += 1
            except Exception as e:
                errors.append(f'صف {row_idx}: {str(e)[:80]}')

        wb.close()
        msg = f'تم استيراد {created} مالك'
        if skipped:
            msg += f'، تم تخطي {skipped} صف فاضي'
        if errors:
            msg += f'، {len(errors)} خطأ'

        return success_response({
            'created': created,
            'skipped': skipped,
            'errors': errors[:10],
        }, msg)

    except Exception as e:
        return error_response(f'خطأ في قراءة الملف: {str(e)}', 400)


@app.route('/api/owners/template')
@require_auth
def owners_template():
    """تحميل نموذج Excel لاستيراد الملاك"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الملاك'
    ws.sheet_view.rightToLeft = True

    headers = ['الاسم', 'الموبايل', 'الإيميل', 'النوع', 'الجنسية', 'اسم الشركة', 'نوع العقار', 'ملاحظات']
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border

    # Sample rows
    samples = [
        ['أحمد محمد', '0501234567', 'ahmed@email.com', 'فرد', 'إماراتي', '', 'فيلا', 'عميل VIP'],
        ['شركة النخيل', '0441234567', 'info@nakheel.ae', 'شركة', '', 'شركة النخيل العقارية', 'برج سكني', ''],
        ['سارة أحمد', '0551234567', '', 'فرد', 'مصري', '', 'استوديو', ''],
    ]
    for r_idx, sample in enumerate(samples, 2):
        for c_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border

    # Column widths
    widths = [20, 16, 25, 10, 12, 22, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Notes row
    ws.cell(row=6, column=1, value='ملاحظات:').font = Font(bold=True, color='EF4444')
    ws.cell(row=7, column=1, value='• عمود "الاسم" مطلوب — الباقي اختياري')
    ws.cell(row=8, column=1, value='• النوع: فرد أو شركة')
    ws.cell(row=9, column=1, value='• نوع العقار: فيلا، شقة، استوديو، مكتب، أرض، بنتهاوس، تاون هاوس، برج سكني')
    ws.cell(row=10, column=1, value='• امسح الصفوف التجريبية قبل الاستيراد')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='owners_template.xlsx',
    )


# ==================== OFF-PLAN PROJECTS ====================

@app.route('/api/offplan', methods=['GET'])
@require_auth
def get_offplan():
    try:
        q = supabase.table('offplan_projects').select('*').order('created_at', desc=True)
        status = request.args.get('status')
        if status:
            q = q.eq('status', status)
        data = q.execute()
        return success_response({'projects': data.data})
    except Exception:
        return success_response({'projects': []})


@app.route('/api/offplan', methods=['POST'])
@check_role(['admin', 'manager'])
def create_offplan():
    try:
        body = request.get_json()
        if not body or not body.get('name'):
            return error_response('Name is required')
        row = {
            'name':          body['name'],
            'developer':     body.get('developer'),
            'location':      body.get('location'),
            'total_units':   body.get('total_units'),
            'handover_date': body.get('handover_date'),
            'price_from':    body.get('price_from'),
            'price_to':      body.get('price_to'),
            'status':        body.get('status', 'upcoming'),
            'description':   body.get('description'),
        }
        data = supabase.table('offplan_projects').insert(row).execute()
        return success_response({'project': data.data[0] if data.data else {}}, 'Off-plan project created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/offplan/<project_id>', methods=['PATCH'])
@check_role(['admin', 'manager'])
def update_offplan(project_id):
    try:
        body = request.get_json()
        allowed = ['name', 'developer', 'location', 'total_units', 'sold_units',
                   'handover_date', 'price_from', 'price_to', 'status', 'description']
        updates = {k: v for k, v in body.items() if k in allowed}
        data = supabase.table('offplan_projects').update(updates).eq('id', project_id).execute()
        return success_response({'project': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


# ==================== TRANSACTIONS & COMMISSIONS ====================

@app.route('/api/transactions', methods=['GET'])
@require_auth
def get_transactions():
    try:
        status = request.args.get('status')
        q = supabase.table('transactions').select(
            '*, listings(title), leads(name, phone), agents:employees!transactions_agent_id_fkey(full_name)'
        ).order('created_at', desc=True).limit(100)
        if status:
            q = q.eq('status', status)
        data = q.execute()
        txns = data.data or []

        # Summary
        total_amount = sum(float(t.get('amount', 0) or 0) for t in txns)
        total_commission = sum(float(t.get('commission_amount', 0) or 0) for t in txns)
        pending_count = sum(1 for t in txns if t.get('status') == 'pending')

        return success_response({
            'transactions': txns,
            'summary': {
                'total_count': len(txns),
                'total_amount': total_amount,
                'total_commission': total_commission,
                'pending_count': pending_count,
            },
        })
    except Exception:
        return success_response({
            'transactions': [],
            'summary': {'total_count': 0, 'total_amount': 0, 'total_commission': 0, 'pending_count': 0},
        })


@app.route('/api/transactions', methods=['POST'])
@check_role(['admin', 'manager', 'agent'])
def create_transaction():
    try:
        body = request.get_json()
        amount = body.get('amount', 0)
        commission_rate = body.get('commission_rate', 2)
        row = {
            'listing_id':      body.get('listing_id'),
            'lead_id':         body.get('lead_id'),
            'agent_id':        request.user_id,
            'type':            body.get('type', 'sale'),
            'amount':          amount,
            'commission_rate':  commission_rate,
            'commission_amount': body.get('commission_amount', amount * commission_rate / 100),
            'notes':           body.get('notes'),
            # ─── Property details ───
            'property_type':       body.get('property_type'),
            'property_title':      body.get('property_title'),
            'property_address':    body.get('property_address'),
            'property_size':       body.get('property_size'),
            'property_bedrooms':   body.get('property_bedrooms'),
            'property_bathrooms':  body.get('property_bathrooms'),
            'property_furnished':  body.get('property_furnished'),
            'property_emirate':    body.get('property_emirate'),
            # ─── Party details (owner/buyer/tenant) ───
            'party_role':        body.get('party_role'),
            'party_name':        body.get('party_name'),
            'party_phone':       body.get('party_phone'),
            'party_email':       body.get('party_email'),
            'party_id_number':   body.get('party_id_number'),
            'party_nationality': body.get('party_nationality'),
            # ─── Broker company details ───
            'broker_company':           body.get('broker_company'),
            'broker_agent_name':        body.get('broker_agent_name'),
            'broker_phone':             body.get('broker_phone'),
            'broker_email':             body.get('broker_email'),
            'broker_license':           body.get('broker_license'),
            'broker_commission_split':  body.get('broker_commission_split'),
        }
        # Strip None values so we don't overwrite defaults
        row = {k: v for k, v in row.items() if v is not None}
        # Auto-set owner from listing
        if body.get('listing_id'):
            try:
                listing = supabase.table('listings').select('owner_id').eq('id', body['listing_id']).single().execute()
                if listing.data:
                    row['owner_id'] = listing.data.get('owner_id')
            except Exception:
                pass
        data = supabase.table('transactions').insert(row).execute()
        return success_response({'transaction': data.data[0] if data.data else {}}, 'Transaction created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/transactions/<txn_id>', methods=['PATCH'])
@check_role(['admin', 'manager', 'agent'])
@audit_log('transaction_update', 'transaction')
def update_transaction(txn_id):
    """Update transaction details — property, party, broker, etc."""
    try:
        body = request.get_json() or {}
        allowed_keys = {
            'type', 'amount', 'commission_rate', 'commission_amount', 'status',
            'contract_date', 'completion_date', 'notes',
            'property_type', 'property_title', 'property_address', 'property_size',
            'property_bedrooms', 'property_bathrooms', 'property_furnished', 'property_emirate',
            'party_role', 'party_name', 'party_phone', 'party_email',
            'party_id_number', 'party_nationality',
            'broker_company', 'broker_agent_name', 'broker_phone', 'broker_email',
            'broker_license', 'broker_commission_split',
        }
        row = {k: v for k, v in body.items() if k in allowed_keys}
        row['updated_at'] = datetime.utcnow().isoformat()
        if not row:
            return error_response('لا توجد حقول للتحديث', 400)
        data = supabase.table('transactions').update(row).eq('id', txn_id).execute()
        return success_response({'transaction': data.data[0] if data.data else {}}, 'تم التحديث')
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/transactions/<txn_id>/approve', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('approve_transaction', 'transaction')
def approve_transaction(txn_id):
    try:
        data = supabase.table('transactions').update({
            'status': 'approved',
            'approved_by': request.user_id,
            'approved_at': datetime.utcnow().isoformat(),
        }).eq('id', txn_id).execute()
        return success_response({'transaction': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


# ==================== LEAD ROTATION ====================

@app.route('/api/rotation-rules', methods=['GET'])
@check_role(['admin', 'manager'])
def get_rotation_rules():
    try:
        data = supabase.table('lead_rotation_rules').select('*').order('created_at', desc=True).execute()
        return success_response({'rules': data.data})
    except Exception:
        return success_response({'rules': []})


@app.route('/api/rotation-rules', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('create_rotation_rule', 'rotation')
def create_rotation_rule():
    try:
        body = request.get_json()
        if not body or not body.get('name'):
            return error_response('Name is required')
        row = {
            'name':    body['name'],
            'method':  body.get('method', 'round_robin'),
            'agents':  body.get('agents', []),
            'active':  True,
        }
        data = supabase.table('lead_rotation_rules').insert(row).execute()
        return success_response({'rule': data.data[0] if data.data else {}}, 'Rule created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/rotation-rules/<rule_id>', methods=['PATCH'])
@check_role(['admin', 'manager'])
def update_rotation_rule(rule_id):
    try:
        body = request.get_json()
        allowed = ['name', 'method', 'agents', 'active', 'source_filter']
        updates = {k: v for k, v in body.items() if k in allowed}
        data = supabase.table('lead_rotation_rules').update(updates).eq('id', rule_id).execute()
        return success_response({'rule': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/rotation-rules/<rule_id>', methods=['DELETE'])
@check_role(['admin', 'manager'])
@audit_log('delete_rotation_rule', 'rotation')
def delete_rotation_rule(rule_id):
    try:
        supabase.table('lead_rotation_rules').delete().eq('id', rule_id).execute()
        return success_response({}, 'Rule deleted')
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/rotation/assign', methods=['POST'])
@check_role(['admin', 'manager'])
def rotation_assign_lead():
    """Auto-assign a lead using active rotation rules."""
    try:
        body = request.get_json()
        lead_id = body.get('lead_id')
        if not lead_id:
            return error_response('lead_id required')

        rules = supabase.table('lead_rotation_rules').select('*').eq('active', True).execute()
        if not rules.data:
            return error_response('No active rotation rules')

        rule = rules.data[0]
        agents = rule.get('agents', [])
        if not agents:
            return error_response('No agents in rotation rule')

        method = rule.get('method', 'round_robin')
        idx = rule.get('current_index', 0) % len(agents)

        if method == 'round_robin':
            agent_id = agents[idx]
            supabase.table('lead_rotation_rules').update({'current_index': idx + 1}).eq('id', rule['id']).execute()
        elif method == 'random':
            import random
            agent_id = random.choice(agents)
        else:
            agent_id = agents[idx]

        supabase.table('leads').update({'assigned_to': agent_id}).eq('id', lead_id).execute()
        return success_response({'assigned_to': agent_id, 'rule': rule['name']})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


# ==================== WORKFLOWS & APPROVALS ====================

@app.route('/api/workflows', methods=['GET'])
@require_auth
def get_workflows():
    try:
        status = request.args.get('status')
        q = supabase.table('workflows').select(
            '*, requester:employees!workflows_requested_by_fkey(full_name), approver:employees!workflows_approved_by_fkey(full_name)'
        ).order('created_at', desc=True).limit(50)
        if status:
            q = q.eq('status', status)
        data = q.execute()
        return success_response({'workflows': data.data})
    except Exception:
        return success_response({'workflows': []})


@app.route('/api/workflows', methods=['POST'])
@require_auth
def create_workflow():
    try:
        body = request.get_json()
        row = {
            'type':          body.get('type', 'general'),
            'entity_id':     body.get('entity_id'),
            'entity_type':   body.get('entity_type'),
            'requested_by':  request.user_id,
            'notes':         body.get('notes'),
        }
        data = supabase.table('workflows').insert(row).execute()
        return success_response({'workflow': data.data[0] if data.data else {}}, 'Workflow created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/workflows/<wf_id>/approve', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('approve_workflow', 'workflow')
def approve_workflow(wf_id):
    try:
        data = supabase.table('workflows').update({
            'status': 'approved',
            'approved_by': request.user_id,
            'responded_at': datetime.utcnow().isoformat(),
        }).eq('id', wf_id).execute()
        return success_response({'workflow': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/workflows/<wf_id>/reject', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('reject_workflow', 'workflow')
def reject_workflow(wf_id):
    try:
        body = request.get_json(silent=True) or {}
        data = supabase.table('workflows').update({
            'status': 'rejected',
            'approved_by': request.user_id,
            'responded_at': datetime.utcnow().isoformat(),
            'response_notes': body.get('notes', ''),
        }).eq('id', wf_id).execute()
        return success_response({'workflow': data.data[0] if data.data else {}})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


# ==================== KPI INSIGHTS ====================

@app.route('/api/kpi', methods=['GET'])
@check_role(['admin', 'manager'])
def get_kpi():
    try:
        now = datetime.utcnow()
        month_start = now.replace(day=1, hour=0, minute=0, second=0).isoformat()

        # Monthly deals
        try:
            deals = supabase.table('leads').select('id', count='exact').eq('status', 'closed_won').gte('updated_at', month_start).execute()
            monthly_deals = deals.count or 0
        except Exception:
            monthly_deals = 0

        # Monthly revenue from transactions
        try:
            txns = supabase.table('transactions').select('amount').gte('created_at', month_start).execute()
            monthly_revenue = sum(float(t.get('amount', 0) or 0) for t in (txns.data or []))
        except Exception:
            monthly_revenue = 0

        # Conversion rate
        try:
            total_leads = supabase.table('leads').select('id', count='exact').gte('created_at', month_start).execute()
            total_count = total_leads.count or 1
            conversion_rate = round(monthly_deals / total_count * 100, 1) if total_count > 0 else 0
        except Exception:
            conversion_rate = 0

        # Team performance
        try:
            team = supabase.table('employees').select('full_name, total_deals, total_xp').eq('role', 'agent').order('total_deals', desc=True).limit(10).execute()
            team_performance = [{'name': t['full_name'], 'deals': t.get('total_deals', 0)} for t in (team.data or [])]
        except Exception:
            team_performance = []

        # KPI Targets
        try:
            targets = supabase.table('kpi_targets').select('*').eq('period', 'monthly').execute()
            targets_data = targets.data or []
        except Exception:
            targets_data = []

        return success_response({
            'kpis': {
                'monthly_deals': monthly_deals,
                'monthly_revenue': monthly_revenue,
                'conversion_rate': conversion_rate,
                'avg_close_days': 0,
            },
            'team_performance': team_performance,
            'targets': targets_data,
        })
    except Exception:
        return success_response({
            'kpis': {'monthly_deals': 0, 'monthly_revenue': 0, 'conversion_rate': 0, 'avg_close_days': 0},
            'team_performance': [],
            'targets': [],
        })


# ==================== CUSTOM FIELDS ====================

@app.route('/api/custom-fields', methods=['GET'])
@require_auth
def get_custom_fields():
    try:
        entity_type = request.args.get('entity_type', 'lead')
        data = supabase.table('custom_fields').select('*').eq('entity_type', entity_type).eq('active', True).order('sort_order').execute()
        return success_response({'fields': data.data})
    except Exception:
        return success_response({'fields': []})


@app.route('/api/custom-fields', methods=['POST'])
@check_role(['admin', 'manager'])
def create_custom_field():
    try:
        body = request.get_json()
        if not body or not body.get('field_name') or not body.get('field_label'):
            return error_response('field_name and field_label required')
        row = {
            'entity_type': body.get('entity_type', 'lead'),
            'field_name':  body['field_name'],
            'field_label': body['field_label'],
            'field_type':  body.get('field_type', 'text'),
            'options':     body.get('options', []),
            'required':    body.get('required', False),
        }
        data = supabase.table('custom_fields').insert(row).execute()
        return success_response({'field': data.data[0] if data.data else {}}, 'Field created', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/custom-fields/<field_id>', methods=['DELETE'])
@check_role(['admin'])
def delete_custom_field(field_id):
    try:
        supabase.table('custom_fields').update({'active': False}).eq('id', field_id).execute()
        return success_response({}, 'Field deactivated')
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/custom-fields/values/<entity_id>', methods=['GET'])
@require_auth
def get_custom_field_values(entity_id):
    try:
        data = supabase.table('custom_field_values').select('*, custom_fields(*)').eq('entity_id', entity_id).execute()
        return success_response({'values': data.data})
    except Exception:
        return success_response({'values': []})


@app.route('/api/custom-fields/values/<entity_id>', methods=['POST'])
@require_auth
def set_custom_field_values(entity_id):
    try:
        body = request.get_json()
        values = body.get('values', {})
        for field_id, value in values.items():
            supabase.table('custom_field_values').upsert({
                'field_id': field_id,
                'entity_id': entity_id,
                'value': str(value),
            }, on_conflict='field_id,entity_id').execute()
        return success_response({}, 'Values saved')
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)



# ==================== DATABASE MANAGEMENT ====================

@app.route('/api/admin/database/stats', methods=['GET'])
@check_role(['admin'])
def database_stats():
    try:
        result = supabase.rpc('get_table_counts').execute()
        tables = [{'name': r['table_name'], 'count': r['row_count']} for r in (result.data or [])]
    except Exception:
        # Fallback: count main tables manually
        tables = []
        for tbl in ['leads', 'employees', 'listings', 'owners', 'transactions',
                     'offplan_projects', 'workflows', 'custom_fields', 'activity_feed',
                     'leaderboard', 'quests', 'admin_audit_log']:
            try:
                res = supabase.table(tbl).select('id', count='exact').limit(0).execute()
                tables.append({'name': tbl, 'count': res.count or 0})
            except Exception:
                pass
    return success_response({'tables': tables})


@app.route('/api/admin/database/backup', methods=['POST'])
@check_role(['admin'])
@audit_log('database_backup', 'system')
def database_backup():
    """Export all main tables as CSV (triggers download)."""
    return success_response({'message': 'Use CSV export per-table for now', 'download_url': None})


@app.route('/api/admin/database/cleanup', methods=['POST'])
@check_role(['admin'])
@audit_log('database_cleanup', 'system')
def database_cleanup():
    """Clean old activity feed entries and expired data."""
    cutoff = (datetime.utcnow() - timedelta(days=90)).isoformat()
    with contextlib.suppress(Exception):
        supabase.table('activity_feed').delete().lt('created_at', cutoff).execute()
    return success_response({'message': 'Cleanup complete — removed activity older than 90 days'})


# ==================== HOT LEADS (Random Events) ====================

@app.route('/api/hot-leads/spawn', methods=['POST'])
@check_role(['admin', 'manager'])
def spawn_hot_lead():
    """Manager spawns a golden Hot Lead on a random desk. First to claim wins +100 XP."""
    data = request.json or {}
    desk_index = data.get('desk_index')
    xp_reward = data.get('xp_reward', 100)
    duration_secs = data.get('duration_secs', 60)

    if desk_index is None:
        import random
        desk_index = random.randint(0, 5)

    expires_at = (datetime.utcnow() + timedelta(seconds=duration_secs)).isoformat()

    try:
        row = supabase.table('hot_lead_events').insert({
            'desk_index': desk_index,
            'xp_reward': xp_reward,
            'spawned_by': request.user_id,
            'expires_at': expires_at,
        }).execute()
        return success_response({'event': row.data, 'desk_index': desk_index, 'expires_at': expires_at})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/hot-leads/claim', methods=['POST'])
@require_auth
def claim_hot_lead():
    """Player claims the active Hot Lead. Race-condition safe via RPC."""
    data = request.json or {}
    event_id = data.get('event_id')
    if not event_id:
        return error_response('event_id required', 400)

    try:
        result = supabase.rpc('claim_hot_lead', {
            'p_user_id': request.user_id,
            'p_event_id': event_id,
        }).execute()

        claim = result.data or {}
        if claim.get('ok'):
            xp = claim.get('xp_reward', 100)
            with contextlib.suppress(Exception):
                supabase.rpc('award_xp_and_coins', {
                    'p_employee_id': request.user_id,
                    'p_xp': xp,
                    'p_coins': xp // 2,
                    'p_reason': 'hot_lead_claimed',
                }).execute()
            with contextlib.suppress(Exception):
                supabase.table('actions_log').insert({
                    'employee_id': request.user_id,
                    'action': 'deal_closed',
                    'xp_earned': xp,
                    'coins_earned': xp // 2,
                    'details': {'type': 'hot_lead_claimed', 'event_id': event_id},
                }).execute()
        return success_response(claim)
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/hot-leads/active', methods=['GET'])
@require_auth
def get_active_hot_leads():
    """Get currently active (unclaimed, not expired) Hot Lead events."""
    try:
        now = datetime.utcnow().isoformat()
        res = supabase.table('hot_lead_events').select('*') \
            .is_('claimed_by', 'null') \
            .gte('expires_at', now) \
            .order('created_at', desc=True) \
            .limit(5).execute()
        return success_response({'events': res.data or []})
    except Exception:
        return success_response({'events': []})


# ==================== HIGH FIVES (Social) ====================

@app.route('/api/high-five/send', methods=['POST'])
@require_auth
def send_high_five():
    """Send a High Five to another player. Both get +5 XP. 5-min cooldown per pair."""
    data = request.json or {}
    to_user = data.get('to_user_id')
    if not to_user:
        return error_response('to_user_id required', 400)

    try:
        result = supabase.rpc('send_high_five', {
            'p_from': request.user_id,
            'p_to': to_user,
        }).execute()

        hf = result.data or {}
        if hf.get('ok'):
            # Award +5 XP to both players
            for uid in [request.user_id, to_user]:
                with contextlib.suppress(Exception):
                    supabase.rpc('award_xp_and_coins', {
                        'p_employee_id': uid,
                        'p_xp': 5,
                        'p_coins': 2,
                        'p_reason': 'high_five',
                    }).execute()
        return success_response(hf)
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/high-fives/today', methods=['GET'])
@require_auth
def get_high_fives_today():
    """Get today's High Five history for the current user."""
    try:
        today = datetime.utcnow().strftime('%Y-%m-%d')
        given = supabase.table('high_fives').select('*, to_emp:employees!high_fives_to_user_fkey(full_name)') \
            .eq('from_user', request.user_id).gte('created_at', today).execute()
        received = supabase.table('high_fives').select('*, from_emp:employees!high_fives_from_user_fkey(full_name)') \
            .eq('to_user', request.user_id).gte('created_at', today).execute()
        return success_response({
            'given': given.data or [],
            'received': received.data or [],
            'total_given': len(given.data or []),
            'total_received': len(received.data or []),
        })
    except Exception:
        return success_response({'given': [], 'received': [], 'total_given': 0, 'total_received': 0})


# ==================== POWER HOURS (Manager Events) ====================

@app.route('/api/power-hours/activate', methods=['POST'])
@check_role(['admin', 'manager'])
def activate_power_hour():
    """Manager activates a Power Hour event (happy_hour, lightning_round, sniper_mode)."""
    data = request.json or {}
    ph_type = data.get('type', 'happy_hour')
    duration = data.get('duration', 60)  # minutes

    if ph_type not in ('happy_hour', 'lightning_round', 'sniper_mode'):
        return error_response('Invalid type. Use: happy_hour, lightning_round, sniper_mode', 400)

    multiplier = {'happy_hour': 2, 'lightning_round': 1.5, 'sniper_mode': 1}.get(ph_type, 2)
    ends_at = (datetime.utcnow() + timedelta(minutes=duration)).isoformat()

    try:
        # Deactivate any existing active power hour
        with contextlib.suppress(Exception):
            supabase.table('power_hours').update({'is_active': False}).eq('is_active', True).execute()

        row = supabase.table('power_hours').insert({
            'type': ph_type,
            'multiplier': multiplier,
            'duration': duration,
            'started_by': request.user_id,
            'ends_at': ends_at,
            'is_active': True,
        }).execute()
        return success_response({'power_hour': row.data, 'ends_at': ends_at, 'type': ph_type})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/power-hours/active', methods=['GET'])
@require_auth
def get_active_power_hour():
    """Get currently active Power Hour (if any)."""
    try:
        now = datetime.utcnow().isoformat()
        res = supabase.table('power_hours').select('*') \
            .eq('is_active', True) \
            .gte('ends_at', now) \
            .order('started_at', desc=True) \
            .limit(1).execute()
        active = (res.data or [None])[0]
        return success_response({'power_hour': active})
    except Exception:
        return success_response({'power_hour': None})


# ==================== DAILY HIGHLIGHTS ====================

@app.route('/api/daily-highlights/generate', methods=['POST'])
@check_role(['admin', 'manager'])
def generate_daily_highlights():
    """Generate end-of-day highlights cinematic data."""
    try:
        result = supabase.rpc('generate_daily_highlights', {
            'p_admin_id': request.user_id,
        }).execute()
        return success_response(result.data or {})
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/daily-highlights/latest', methods=['GET'])
@require_auth
def get_latest_highlights():
    """Get the latest daily highlights."""
    try:
        res = supabase.table('daily_highlights').select('*') \
            .order('highlight_date', desc=True) \
            .limit(1).execute()
        highlights = (res.data or [None])[0]
        return success_response({'highlights': highlights})
    except Exception:
        return success_response({'highlights': None})


# ==================== BREAK ROOM ====================

@app.route('/api/breakroom/coffee', methods=['POST'])
@require_auth
def breakroom_coffee():
    """Coffee machine: instant +30 stamina (player must be on break, 10-min cooldown client-side)."""
    try:
        result = supabase.rpc('regen_stamina', {
            'p_employee_id': request.user_id,
            'p_amount': 30,
        }).execute()
        return success_response(result.data, 'قهوة! طاقة +30 ☕')
    except Exception:
        return success_response({'stamina': 100, 'regen': 30}, 'قهوة!')


@app.route('/api/breakroom/minigame', methods=['POST'])
@require_auth
def breakroom_minigame():
    """Award small XP for completing break room mini-game (reaction time)."""
    data = request.json or {}
    score = data.get('score', 0)
    xp = min(10, max(2, score // 100))  # 2-10 XP based on score
    try:
        multiplier = get_power_hour_multiplier()
        xp = int(xp * multiplier)
        supabase.rpc('award_xp_and_coins', {
            'p_employee_id': request.user_id,
            'p_xp': xp,
            'p_coins': 1,
            'p_reason': 'minigame',
            'p_ref_type': 'action',
        }).execute()
        supabase.table('actions_log').insert({
            'employee_id': request.user_id,
            'action': 'minigame',
            'xp_earned': xp,
            'coins_earned': 1,
            'details': {'score': score},
        }).execute()
    except Exception:
        pass
    return success_response({'xp': xp, 'coins': 1}, f'ميني جيم! +{xp} XP')


# ==================== STORY MODE ====================

STORY_MILESTONES = [
    {'id': 0, 'xp': 0, 'ch': 'Chapter 0', 'title': 'The Rookie'},
    {'id': 1, 'xp': 100, 'ch': 'Chapter 1', 'title': 'First Contact'},
    {'id': 2, 'xp': 300, 'ch': 'Chapter 2', 'title': 'The Caller'},
    {'id': 3, 'xp': 600, 'ch': 'Chapter 3', 'title': 'The Hunter'},
    {'id': 4, 'xp': 1000, 'ch': 'Chapter 4', 'title': 'Face to Face'},
    {'id': 5, 'xp': 1500, 'ch': 'Chapter 5', 'title': 'First Blood'},
    {'id': 6, 'xp': 2200, 'ch': 'Chapter 6', 'title': 'The Machine'},
    {'id': 7, 'xp': 3000, 'ch': 'Chapter 7', 'title': 'The Persuader'},
    {'id': 8, 'xp': 4000, 'ch': 'Chapter 8', 'title': 'Rising Star'},
    {'id': 9, 'xp': 5000, 'ch': 'Chapter 9', 'title': 'The Negotiator'},
    {'id': 10, 'xp': 6500, 'ch': 'Chapter 10', 'title': 'The Obelisk'},
    {'id': 11, 'xp': 8000, 'ch': 'Chapter 11', 'title': 'The Sniper'},
    {'id': 12, 'xp': 10000, 'ch': 'Chapter 12', 'title': 'Elite Access'},
    {'id': 13, 'xp': 12500, 'ch': 'Chapter 13', 'title': "Legend's Gate"},
    {'id': 14, 'xp': 15000, 'ch': 'Chapter 14', 'title': 'The Millionaire'},
    {'id': 15, 'xp': 18000, 'ch': 'Chapter 15', 'title': 'The Mentor'},
    {'id': 16, 'xp': 22000, 'ch': 'Chapter 16', 'title': 'Commander'},
    {'id': 17, 'xp': 27000, 'ch': 'Chapter 17', 'title': 'Champion'},
    {'id': 18, 'xp': 35000, 'ch': 'Chapter 18', 'title': 'The Samurai'},
    {'id': 19, 'xp': 50000, 'ch': 'Chapter 19', 'title': 'The Legend'},
]


@app.route('/api/story/progress', methods=['GET'])
@require_auth
def story_progress():
    """Returns player's story mode progress — current chapter + unlocked milestones."""
    try:
        emp = supabase.table('employees').select('total_xp, level').eq(
            'id', request.user_id).single().execute()
        xp = (emp.data or {}).get('total_xp', 0)
        current = 0
        for m in reversed(STORY_MILESTONES):
            if xp >= m['xp']:
                current = m['id']
                break
        return success_response({
            'current_chapter': current,
            'total_chapters': len(STORY_MILESTONES),
            'milestones': [
                {**m, 'unlocked': xp >= m['xp']} for m in STORY_MILESTONES
            ],
        })
    except Exception:
        return success_response({'current_chapter': 0, 'total_chapters': 20, 'milestones': []})


# ==================== CEO VISITS ====================

@app.route('/api/ceo-visit/start', methods=['POST'])
@check_role(['admin'])
def start_ceo_visit():
    """Start a CEO visit event — 3x multiplier stored in power_hours table."""
    data = request.json or {}
    duration = data.get('duration', 30)
    ends_at = (datetime.utcnow() + timedelta(minutes=duration)).isoformat()
    try:
        supabase.table('power_hours').insert({
            'type': 'ceo_visit',
            'multiplier': 3.0,
            'started_by': request.user_id,
            'ends_at': ends_at,
            'is_active': True,
        }).execute()
        return success_response({
            'type': 'ceo_visit', 'duration': duration, 'multiplier': 3.0,
        }, 'CEO في المكتب! 🕴️')
    except Exception:
        return success_response({
            'type': 'ceo_visit', 'duration': duration,
        }, 'CEO visit started')


@app.route('/api/ceo-visit/active', methods=['GET'])
@require_auth
def get_active_ceo_visit():
    """Returns currently active CEO visit (if any)."""
    try:
        result = supabase.table('power_hours').select('*') \
            .eq('type', 'ceo_visit').eq('is_active', True).execute()
        visit = result.data[0] if result.data else None
        return success_response({'visit': visit})
    except Exception:
        return success_response({'visit': None})


# ==================== SQUADS ====================

@app.route('/api/squads/create', methods=['POST'])
@require_auth
def create_squad():
    """Create a squad (max 4 members, named by creator)."""
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return error_response('اسم السكواد مطلوب', 400)
    try:
        squad = supabase.table('teams').insert({
            'name': name,
            'manager_id': request.user_id,
        }).execute()
        # Add creator as member
        supabase.table('employees').update({
            'team_id': squad.data[0]['id'],
        }).eq('id', request.user_id).execute()
        return success_response(squad.data[0], 'سكواد جديد! 🤝')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/squads/join', methods=['POST'])
@require_auth
def join_squad():
    """Join an existing squad (max 4 members)."""
    data = request.json or {}
    squad_id = data.get('squad_id')
    if not squad_id:
        return error_response('squad_id مطلوب', 400)
    try:
        # Check squad size (max 4)
        members = supabase.table('employees').select('id', count='exact') \
            .eq('team_id', squad_id).execute()
        if members.count and members.count >= 4:
            return error_response('السكواد ممتلئ (حد أقصى 4)', 400)
        supabase.table('employees').update({
            'team_id': squad_id,
        }).eq('id', request.user_id).execute()
        return success_response({'squad_id': squad_id}, 'انضممت للسكواد! 🎯')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/squads', methods=['GET'])
@require_auth
def list_squads():
    """List all squads with member details."""
    try:
        squads = supabase.table('teams').select(
            '*, employees(id, name, level, total_xp)'
        ).execute()
        return success_response({'squads': squads.data})
    except Exception:
        return success_response({'squads': []})


@app.route('/api/squads/my', methods=['GET'])
@require_auth
def my_squad():
    """Get my squad details + members + combined stats."""
    try:
        emp = supabase.table('employees').select('team_id') \
            .eq('id', request.user_id).single().execute()
        if not emp.data or not emp.data.get('team_id'):
            return success_response({'squad': None})
        tid = emp.data['team_id']
        squad = supabase.table('teams').select(
            '*, employees(id, name, level, total_xp, avatar_color, status)'
        ).eq('id', tid).single().execute()
        return success_response({'squad': squad.data})
    except Exception:
        return success_response({'squad': None})


@app.route('/api/squads/leave', methods=['POST'])
@require_auth
def leave_squad():
    """Leave current squad."""
    try:
        supabase.table('employees').update({
            'team_id': None,
        }).eq('id', request.user_id).execute()
        return success_response({}, 'خرجت من السكواد')
    except Exception:
        return success_response({})


# ==================== MAP REWARDS (AL SAEB Tower) ====================
#
# Server-authoritative milestone + skill-branch rewards for /map.
# The MILESTONES array in templates/map.html is the visual mirror; this dict
# is the trust boundary — client-supplied reward values are never honored.

MAP_MILESTONES = [
    {'idx': 0, 'key': 'deira',           'xp_req': 0,     'reward_xp': 0,    'reward_coins': 0},
    {'idx': 1, 'key': 'bur_dubai',       'xp_req': 200,   'reward_xp': 50,   'reward_coins': 25},
    {'idx': 2, 'key': 'karama',          'xp_req': 500,   'reward_xp': 100,  'reward_coins': 50},
    {'idx': 3, 'key': 'al_quoz',         'xp_req': 1000,  'reward_xp': 150,  'reward_coins': 75},
    {'idx': 4, 'key': 'business_bay',    'xp_req': 2000,  'reward_xp': 250,  'reward_coins': 150},
    {'idx': 5, 'key': 'dubai_marina',    'xp_req': 3500,  'reward_xp': 400,  'reward_coins': 250},
    {'idx': 6, 'key': 'downtown_dubai',  'xp_req': 5500,  'reward_xp': 600,  'reward_coins': 400},
    {'idx': 7, 'key': 'jbr',             'xp_req': 8000,  'reward_xp': 800,  'reward_coins': 600},
    {'idx': 8, 'key': 'palm_jumeirah',   'xp_req': 11000, 'reward_xp': 1000, 'reward_coins': 1000},
    {'idx': 9, 'key': 'burj_khalifa',    'xp_req': 15000, 'reward_xp': 2000, 'reward_coins': 2500},
]

MAP_BRANCH_DEFAULTS = {
    'calls':     {'target': 10, 'xp': 150,  'coins': 50},
    'whatsapp':  {'target': 50, 'xp': 250,  'coins': 100},
    'meetings':  {'target': 20, 'xp': 400,  'coins': 200},
    'big_deals': {'target': 1,  'xp': 1000, 'coins': 500},
}
MAP_BRANCH_KEYS = tuple(MAP_BRANCH_DEFAULTS.keys())


def _branch_view(key, row):
    """Project a branch_progress row (or absent row) to the API shape."""
    d = MAP_BRANCH_DEFAULTS[key]
    if not row:
        return {
            'branch_key': key, 'current': 0, 'target': d['target'],
            'xp_reward': d['xp'], 'coin_reward': d['coins'],
            'claimed_at': None, 'claimable': False,
        }
    current = row.get('current_count', 0) or 0
    target = row.get('target_count', d['target']) or d['target']
    claimed = row.get('claimed_at') is not None
    return {
        'branch_key':  key,
        'current':     current,
        'target':      target,
        'xp_reward':   row.get('xp_reward', d['xp']),
        'coin_reward': row.get('coin_reward', d['coins']),
        'claimed_at':  row.get('claimed_at'),
        'claimable':   (not claimed) and current >= target,
    }


@app.route('/api/map/milestones', methods=['GET'])
@require_auth
def map_milestones():
    """Milestone definitions + claim status + branch progress for current user."""
    try:
        emp_res = supabase.table('employees') \
            .select('total_xp, level, syb_coins, total_deals, full_name, title') \
            .eq('id', request.user_id).single().execute()
        emp_data = emp_res.data or {}
        total_xp = emp_data.get('total_xp', 0) or 0

        claims_res = supabase.table('milestone_claims') \
            .select('milestone_idx, xp_awarded, coins_awarded, claimed_at') \
            .eq('employee_id', request.user_id).execute()
        claimed_by_idx = {c['milestone_idx']: c for c in (claims_res.data or [])}

        milestones_out = []
        for m in MAP_MILESTONES:
            unlocked = total_xp >= m['xp_req']
            claimed = m['idx'] in claimed_by_idx
            milestones_out.append({
                **m,
                'unlocked':   unlocked,
                'claimed':    claimed,
                'claimable':  unlocked and not claimed and (m['reward_xp'] > 0 or m['reward_coins'] > 0),
                'claimed_at': claimed_by_idx.get(m['idx'], {}).get('claimed_at'),
            })

        branches_res = supabase.table('branch_progress') \
            .select('branch_key, current_count, target_count, xp_reward, coin_reward, claimed_at') \
            .eq('employee_id', request.user_id).execute()
        branches_by_key = {b['branch_key']: b for b in (branches_res.data or [])}
        branches_out = [_branch_view(k, branches_by_key.get(k)) for k in MAP_BRANCH_KEYS]

        return success_response({
            'employee':   emp_data,
            'milestones': milestones_out,
            'branches':   branches_out,
        })
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/map/claim/<int:idx>', methods=['POST'])
@rate_limit('map_claim', max_per_min=10)
def claim_map_milestone(idx):
    """Claim the reward for milestone <idx>. Server validates eligibility + uniqueness."""
    if idx < 0 or idx >= len(MAP_MILESTONES):
        return error_response('invalid milestone index', 400)
    m = MAP_MILESTONES[idx]
    if m['reward_xp'] <= 0 and m['reward_coins'] <= 0:
        return error_response('no reward for this milestone', 400)
    try:
        result = supabase.rpc('claim_milestone_reward', {
            'p_employee_id':   request.user_id,
            'p_milestone_idx': m['idx'],
            'p_milestone_key': m['key'],
            'p_required_xp':   m['xp_req'],
            'p_reward_xp':     m['reward_xp'],
            'p_reward_coins':  m['reward_coins'],
        }).execute()
        data = result.data or {}
        if isinstance(data, dict) and data.get('error'):
            return error_response(data.get('error'), 400)
        invalidate_user_cache(request.user_id)
        return success_response(data, 'تم استلام كافأة المحطة 🏆')
    except Exception as e:
        return error_response(str(e), 500)


@app.route('/api/map/branches/<branch_key>/claim', methods=['POST'])
@rate_limit('branch_claim', max_per_min=10)
def claim_map_branch(branch_key):
    """Claim the reward for a completed skill branch."""
    if branch_key not in MAP_BRANCH_KEYS:
        return error_response('invalid branch key', 400)
    try:
        result = supabase.rpc('claim_branch_reward', {
            'p_employee_id': request.user_id,
            'p_branch_key':  branch_key,
        }).execute()
        data = result.data or {}
        if isinstance(data, dict) and data.get('error'):
            return error_response(data.get('error'), 400)
        invalidate_user_cache(request.user_id)
        return success_response(data, 'تم استلام كافأة الفرع 🎯')
    except Exception as e:
        return error_response(str(e), 500)


# ==================== ATTENDANCE (Check-in / Check-out) ====================

@app.route('/api/attendance/checkin', methods=['POST'])
@rate_limit('attendance', max_per_min=5)
def attendance_checkin():
    """تسجيل حضور الموظف — مرة واحدة في اليوم (مع GPS اختياري)"""
    try:
        user_id = request.user_id
        today = datetime.utcnow().strftime('%Y-%m-%d')
        body = request.json or {}

        # ── GPS verification ──
        lat = body.get('latitude')
        lng = body.get('longitude')
        gps_verified = False
        gps_distance_m = None

        if lat is not None and lng is not None and OFFICE_LAT and OFFICE_LNG:
            gps_distance_m = round(haversine_m(float(lat), float(lng), OFFICE_LAT, OFFICE_LNG), 1)
            gps_verified = gps_distance_m <= OFFICE_RADIUS_M
            if REQUIRE_GPS_CHECKIN and not gps_verified:
                return error_response(
                    f'أنت خارج نطاق المكتب ({int(gps_distance_m)} متر — المسموح {OFFICE_RADIUS_M} متر)', 403
                )

        # تحقق إن ما سجلش حضور النهاردة
        existing = supabase.table('attendance').select('id') \
            .eq('employee_id', user_id).eq('date', today).execute()
        if existing.data:
            return error_response('أنت سجلت حضور النهاردة بالفعل', 409)

        # سجل الحضور
        row = {
            'employee_id': user_id,
            'date': today,
            'check_in': datetime.utcnow().isoformat(),
            'status': 'present',
        }
        # GPS fields — may not exist if schema not fully applied
        gps_fields = {}
        if gps_verified:
            gps_fields['gps_verified'] = gps_verified
        if lat is not None:
            gps_fields['latitude'] = float(lat)
            gps_fields['longitude'] = float(lng)
        if gps_distance_m is not None:
            gps_fields['gps_distance_m'] = gps_distance_m

        try:
            record = supabase.table('attendance').insert({**row, **gps_fields}).execute()
        except Exception:
            # GPS columns may not exist yet — retry without them
            record = supabase.table('attendance').insert(row).execute()

        # 🎮 أدي XP للحضور
        ph_mult = get_power_hour_multiplier()
        xp = int(20 * ph_mult)
        coins = int(5 * ph_mult)
        with contextlib.suppress(Exception):
            supabase.rpc('award_xp_and_coins', {
                'p_employee_id': user_id,
                'p_xp': xp,
                'p_coins': coins,
                'p_reason': 'تسجيل حضور',
                'p_ref_type': 'action',
            }).execute()

        # سجل الأكشن
        with contextlib.suppress(Exception):
            supabase.table('actions_log').insert({
                'employee_id': user_id,
                'action': 'check_in',
                'xp_earned': xp,
                'coins_earned': coins,
                'details': {'date': today}
            }).execute()

        with contextlib.suppress(Exception):
            fire_webhooks('employee_checkin', {
                'employee_id': user_id, 'date': today,
                'gps_verified': gps_verified, 'gps_distance_m': gps_distance_m,
            })

        return success_response({
            'attendance': record.data[0] if record.data else None,
            'xp_earned': xp,
            'coins_earned': coins,
            'gps_verified': gps_verified,
            'gps_distance_m': gps_distance_m,
        }, 'تم تسجيل الحضور بنجاح ✅')
    except Exception as e:
        if 'attendance' in str(e) and ('not find' in str(e) or 'does not exist' in str(e)):
            return error_response('جدول الحضور لسه ما اتعمل — شغّل schema_complete.sql في Supabase', 503)
        return error_response(f'خطأ في تسجيل الحضور: {str(e)}', 500)


@app.route('/api/attendance/checkout', methods=['POST'])
@rate_limit('attendance', max_per_min=5)
def attendance_checkout():
    """تسجيل انصراف الموظف"""
    try:
        user_id = request.user_id
        today = datetime.utcnow().strftime('%Y-%m-%d')

        # لازم يكون سجل حضور الأول
        existing = supabase.table('attendance').select('*') \
            .eq('employee_id', user_id).eq('date', today).execute()
        if not existing.data:
            return error_response('لازم تسجل حضور الأول', 400)

        record = existing.data[0]
        if record.get('check_out'):
            return error_response('أنت سجلت انصراف النهاردة بالفعل', 409)

        # حساب ساعات العمل
        check_in_time = datetime.fromisoformat(record['check_in'].replace('Z', '+00:00').replace('+00:00', ''))
        now = datetime.utcnow()
        hours_worked = round((now - check_in_time).total_seconds() / 3600, 2)

        # سجل الانصراف
        updated = supabase.table('attendance').update({
            'check_out': now.isoformat(),
            'hours_worked': hours_worked
        }).eq('id', record['id']).execute()

        # 🎮 XP إضافي لو اشتغل أكتر من 6 ساعات
        bonus_xp = 0
        bonus_coins = 0
        if hours_worked >= 6:
            ph_mult = get_power_hour_multiplier()
            bonus_xp = int(30 * ph_mult)
            bonus_coins = int(10 * ph_mult)
            with contextlib.suppress(Exception):
                supabase.rpc('award_xp_and_coins', {
                    'p_employee_id': user_id,
                    'p_xp': bonus_xp,
                    'p_coins': bonus_coins,
                    'p_reason': f'يوم عمل كامل ({hours_worked:.1f} ساعة)',
                    'p_ref_type': 'action',
                }).execute()

        with contextlib.suppress(Exception):
            supabase.table('actions_log').insert({
                'employee_id': user_id,
                'action': 'check_out',
                'xp_earned': bonus_xp,
                'coins_earned': bonus_coins,
                'details': {'date': today, 'hours_worked': hours_worked}
            }).execute()

        return success_response({
            'attendance': updated.data[0] if updated.data else None,
            'hours_worked': hours_worked,
            'bonus_xp': bonus_xp,
            'bonus_coins': bonus_coins
        }, f'تم تسجيل الانصراف — اشتغلت {hours_worked:.1f} ساعة 👋')
    except Exception as e:
        if 'attendance' in str(e) and ('not find' in str(e) or 'does not exist' in str(e)):
            return error_response('جدول الحضور لسه ما اتعمل — شغّل schema_complete.sql في Supabase', 503)
        return error_response(f'خطأ في تسجيل الانصراف: {str(e)}', 500)


@app.route('/api/attendance/status', methods=['GET'])
@rate_limit('attendance_status', max_per_min=30)
def attendance_status():
    """حالة حضور الموظف النهاردة"""
    try:
        user_id = request.user_id
        today = datetime.utcnow().strftime('%Y-%m-%d')

        record = supabase.table('attendance').select('*') \
            .eq('employee_id', user_id).eq('date', today).execute()

        if not record.data:
            return success_response({'status': 'not_checked_in', 'record': None})

        rec = record.data[0]
        if rec.get('check_out'):
            return success_response({'status': 'checked_out', 'record': rec})
        return success_response({'status': 'checked_in', 'record': rec})
    except Exception:
        return success_response({'status': 'not_checked_in', 'record': None})


@app.route('/api/attendance/history', methods=['GET'])
@rate_limit('attendance_history', max_per_min=15)
def attendance_history():
    """سجل حضور الموظف آخر 30 يوم"""
    try:
        user_id = request.user_id
        limit = min(int(request.args.get('limit', 30)), 90)

        records = supabase.table('attendance').select('*') \
            .eq('employee_id', user_id) \
            .order('date', desc=True) \
            .limit(limit).execute()

        return success_response({'history': records.data})
    except Exception:
        return success_response({'history': []})


@app.route('/api/admin/attendance', methods=['GET'])
@check_role(['admin', 'manager'])
def admin_attendance():
    """سجل حضور كل الموظفين (للأدمن)"""
    try:
        date = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))

        records = supabase.table('attendance').select(
            '*, employees(full_name, email, role, avatar_color)'
        ).eq('date', date).order('check_in', desc=True).execute()

        return success_response({'date': date, 'records': records.data})
    except Exception:
        return success_response({'date': request.args.get('date', ''), 'records': []})


@app.route('/api/admin/attendance/office', methods=['GET', 'POST'])
@check_role(['admin'])
def attendance_office():
    """إعدادات الـ GPS للمكتب — GET يجيب الإحداثيات، POST يحفظها"""
    global OFFICE_LAT, OFFICE_LNG, OFFICE_RADIUS_M, REQUIRE_GPS_CHECKIN
    if request.method == 'GET':
        return success_response({
            'lat': OFFICE_LAT,
            'lng': OFFICE_LNG,
            'radius_m': OFFICE_RADIUS_M,
            'require_gps': REQUIRE_GPS_CHECKIN,
            'configured': bool(OFFICE_LAT and OFFICE_LNG),
        })
    # POST — إعادة إعداد (بيحدّث .env الـ runtime فقط — للحفظ الدائم عدّل .env يدوياً)
    data = request.json or {}
    if 'lat' in data:
        OFFICE_LAT = float(data['lat'])
        os.environ['OFFICE_LAT'] = str(OFFICE_LAT)
    if 'lng' in data:
        OFFICE_LNG = float(data['lng'])
        os.environ['OFFICE_LNG'] = str(OFFICE_LNG)
    if 'radius_m' in data:
        OFFICE_RADIUS_M = int(data['radius_m'])
        os.environ['OFFICE_RADIUS_METERS'] = str(OFFICE_RADIUS_M)
    if 'require_gps' in data:
        REQUIRE_GPS_CHECKIN = bool(data['require_gps'])
        os.environ['REQUIRE_GPS_CHECKIN'] = 'true' if REQUIRE_GPS_CHECKIN else 'false'
    return success_response({
        'lat': OFFICE_LAT, 'lng': OFFICE_LNG,
        'radius_m': OFFICE_RADIUS_M, 'require_gps': REQUIRE_GPS_CHECKIN,
    }, 'تم حفظ إعدادات GPS ✅')


@app.route('/api/admin/attendance/monthly', methods=['GET'])
@check_role(['admin', 'manager'])
def attendance_monthly():
    """تقرير الحضور الشهري لكل موظف"""
    try:
        year = int(request.args.get('year', datetime.utcnow().year))
        month = int(request.args.get('month', datetime.utcnow().month))
        from_date = f'{year}-{month:02d}-01'
        # آخر يوم في الشهر
        to_date = f'{year + 1}-01-01' if month == 12 else f'{year}-{month + 1:02d}-01'

        records = supabase.table('attendance').select(
            '*, employees(full_name, email, role, avatar_color)'
        ).gte('date', from_date).lt('date', to_date).execute()

        # تجميع per employee
        emp_map: dict = {}
        for r in (records.data or []):
            eid = r['employee_id']
            if eid not in emp_map:
                emp_map[eid] = {
                    'employee_id': eid,
                    'full_name': (r.get('employees') or {}).get('full_name', ''),
                    'email': (r.get('employees') or {}).get('email', ''),
                    'role': (r.get('employees') or {}).get('role', ''),
                    'avatar_color': (r.get('employees') or {}).get('avatar_color', '#00e5ff'),
                    'days_present': 0, 'days_late': 0, 'days_absent': 0,
                    'total_hours': 0.0, 'avg_checkin': [],
                    'gps_verified_count': 0,
                }
            e = emp_map[eid]
            status = r.get('status', 'present')
            if status in ('present', 'remote'):
                e['days_present'] += 1
            elif status == 'late':
                e['days_late'] += 1
            elif status == 'absent':
                e['days_absent'] += 1
            e['total_hours'] = round(e['total_hours'] + float(r.get('hours_worked') or 0), 2)
            if r.get('gps_verified'):
                e['gps_verified_count'] += 1
            if r.get('check_in'):
                with contextlib.suppress(Exception):
                    t = datetime.fromisoformat(r['check_in'].replace('Z', ''))
                    e['avg_checkin'].append(t.hour * 60 + t.minute)

        # حساب متوسط وقت الحضور
        summary = []
        for e in emp_map.values():
            if e['avg_checkin']:
                avg_min = int(sum(e['avg_checkin']) / len(e['avg_checkin']))
                e['avg_checkin_time'] = f'{avg_min // 60:02d}:{avg_min % 60:02d}'
            else:
                e['avg_checkin_time'] = '—'
            del e['avg_checkin']
            summary.append(e)

        summary.sort(key=lambda x: x['days_present'], reverse=True)
        return success_response({
            'year': year, 'month': month,
            'from_date': from_date, 'to_date': to_date,
            'summary': summary,
        })
    except Exception as e:
        return success_response({
            'year': int(request.args.get('year', datetime.utcnow().year)),
            'month': int(request.args.get('month', datetime.utcnow().month)),
            'from_date': '', 'to_date': '',
            'summary': [],
        })


# ==================== OUTBOUND WEBHOOKS ====================

@app.route('/api/webhooks', methods=['GET'])
@check_role(['admin'])
def list_webhooks():
    """قائمة الـ webhooks المسجلة"""
    try:
        rows = supabase.table('webhooks').select('id,url,events,is_active,created_at').order('created_at', desc=True).execute()
        return success_response({'webhooks': rows.data})
    except Exception:
        return success_response({'webhooks': []})


@app.route('/api/webhooks', methods=['POST'])
@check_role(['admin'])
@audit_log('webhook_create', target_type='webhook')
def create_webhook():
    """إضافة webhook جديد"""
    try:
        data = request.json or {}
        url = (data.get('url') or '').strip()
        events = data.get('events', ['*'])
        secret = data.get('secret', '')
        if not url or not url.startswith('http'):
            return error_response('URL غير صالح', 400)
        row = supabase.table('webhooks').insert({
            'url': url,
            'events': events,
            'secret': secret,
            'is_active': True,
        }).execute()
        return success_response(row.data[0] if row.data else {}, 'تم إضافة الـ webhook ✅', 201)
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/webhooks/<wh_id>', methods=['PATCH'])
@check_role(['admin'])
def update_webhook(wh_id):
    """تعديل webhook"""
    try:
        data = request.json or {}
        allowed = {k: v for k, v in data.items() if k in ('url', 'events', 'secret', 'is_active')}
        row = supabase.table('webhooks').update(allowed).eq('id', wh_id).execute()
        return success_response(row.data[0] if row.data else {})
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/webhooks/<wh_id>', methods=['DELETE'])
@check_role(['admin'])
@audit_log('webhook_delete', target_type='webhook')
def delete_webhook(wh_id):
    """حذف webhook"""
    try:
        supabase.table('webhooks').delete().eq('id', wh_id).execute()
        return success_response(None, 'تم الحذف')
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


@app.route('/api/webhooks/test/<wh_id>', methods=['POST'])
@check_role(['admin'])
def test_webhook(wh_id):
    """إرسال ping تجريبي للـ webhook"""
    try:
        row = supabase.table('webhooks').select('*').eq('id', wh_id).single().execute()
        if not row.data:
            return error_response('Webhook غير موجود', 404)
        fire_webhooks('ping', {'message': 'AlSaeb webhook test', 'webhook_id': wh_id})
        return success_response(None, 'تم إرسال الـ ping ✅')
    except Exception as e:
        return error_response(f'خطأ: {str(e)}', 500)


# ==================== AI WEEKLY REPORT ====================

@app.route('/api/admin/reports/weekly', methods=['GET'])
@check_role(['admin', 'manager'])
def weekly_ai_report():
    """تقرير AI أسبوعي لأداء الفريق"""
    try:
        days = int(request.args.get('days', 7))
        since = (datetime.utcnow() - timedelta(days=days)).isoformat()

        # جمع إحصائيات الأسبوع
        actions = supabase.table('actions_log').select(
            'action,xp_earned,coins_earned,employee_id'
        ).gte('created_at', since).execute()

        employees = supabase.table('employees').select(
            'id,full_name,total_xp,total_deals,level'
        ).eq('is_active', True).execute()

        # تجميع الإحصائيات
        totals: dict = {}
        action_counts: dict = {}
        for a in (actions.data or []):
            eid = a['employee_id']
            totals[eid] = totals.get(eid, 0) + (a.get('xp_earned') or 0)
            act = a.get('action', 'other')
            action_counts[act] = action_counts.get(act, 0) + 1

        top_performers = sorted(
            [{'id': k, 'xp': v} for k, v in totals.items()],
            key=lambda x: x['xp'], reverse=True
        )[:5]

        emp_map = {e['id']: e['full_name'] for e in (employees.data or [])}
        for p in top_performers:
            p['name'] = emp_map.get(p['id'], 'unknown')

        stats_summary = {
            'period_days': days,
            'total_actions': len(actions.data or []),
            'action_breakdown': action_counts,
            'top_performers': top_performers,
            'total_xp_awarded': sum(totals.values()),
        }

        # استدعاء AI لتلخيص الأداء
        prompt = (
            f"أنت مدير مبيعات خبير. هنا إحصائيات فريق المبيعات خلال {days} أيام الأخيرة:\n"
            f"{json.dumps(stats_summary, ensure_ascii=False, indent=2)}\n\n"
            "اكتب تقرير موجز (150-200 كلمة) بالعربي يشمل:\n"
            "1. أبرز الإنجازات\n"
            "2. نقاط القوة في الفريق\n"
            "3. توصيات للأسبوع القادم\n"
            "اكتب بأسلوب إيجابي ومحفز."
        )

        ai_report = ai_client.generate_text(prompt) if hasattr(ai_client, 'generate_text') else (
            f"أداء الفريق خلال {days} أيام: {stats_summary['total_actions']} إجراء، "
            f"أعلى XP: {top_performers[0]['name'] if top_performers else '—'} "
            f"بـ {top_performers[0]['xp'] if top_performers else 0} نقطة."
        )

        return success_response({
            'stats': stats_summary,
            'ai_report': ai_report,
            'generated_at': datetime.utcnow().isoformat(),
        })
    except Exception:
        return success_response({
            'stats': {'period_days': 7, 'total_actions': 0, 'action_breakdown': {}, 'top_performers': [], 'total_xp_awarded': 0},
            'ai_report': '',
            'generated_at': datetime.utcnow().isoformat(),
        })


# ==================== AD CAMPAIGNS (الإعلانات) ====================

@app.route('/api/admin/ads', methods=['GET'])
@check_role(['admin', 'manager'])
def list_ad_campaigns():
    """كل الإعلانات مع إحصائيات الليدز"""
    try:
        platform = request.args.get('platform')
        status = request.args.get('status')
        project_id = request.args.get('project_id')
        search = request.args.get('search', '').strip()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        q = supabase.table('ad_campaigns').select(
            '*, projects(name, slug)'
        ).order('created_at', desc=True)

        if platform:
            q = q.eq('platform', platform)
        if status:
            q = q.eq('status', status)
        if project_id:
            q = q.eq('project_id', project_id)
        if search:
            q = q.or_(f"name.ilike.%{search}%,platform_campaign_id.ilike.%{search}%,notes.ilike.%{search}%")

        q = q.range((page - 1) * per_page, page * per_page - 1)
        result = q.execute()
        campaigns = result.data or []

        # إحصائيات الليدز لكل إعلان
        for camp in campaigns:
            leads_q = supabase.table('leads').select(
                'id', count='exact'
            ).eq('ad_campaign_id', camp['id']).eq('is_active', True).execute()
            camp['leads_count'] = leads_q.count or 0

            won_q = supabase.table('leads').select(
                'id', count='exact'
            ).eq('ad_campaign_id', camp['id']).eq('status', 'closed_won').execute()
            camp['won_count'] = won_q.count or 0

            spent = float(camp.get('spent') or 0)
            camp['cpl'] = round(spent / camp['leads_count'], 2) if camp['leads_count'] > 0 else 0

        return success_response({
            'campaigns': campaigns,
            'page': page,
            'per_page': per_page,
        })
    except Exception as e:
        if 'ad_campaigns' in str(e) and ('not find' in str(e) or 'does not exist' in str(e)):
            return success_response({'campaigns': [], 'page': 1, 'per_page': 20},
                                    'جدول الإعلانات لسه ما اتعمل — شغّل schema_complete.sql في Supabase')
        return error_response(f"Ads List Error: {str(e)}", 500)


@app.route('/api/admin/ads', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('ad_campaign_create')
def create_ad_campaign():
    """إنشاء إعلان جديد"""
    try:
        body = request.get_json()
        if not body.get('name') or not body.get('platform'):
            return error_response('الاسم والمنصة مطلوبين', 400)

        def safe_int(v, mx=2_000_000_000):
            try:
                return min(int(v), mx)
            except (TypeError, ValueError):
                return 0

        row = {
            'name': body['name'],
            'platform': body['platform'],
            'project_id': body.get('project_id') or None,
            'platform_campaign_id': body.get('platform_campaign_id') or None,
            'status': body.get('status', 'active'),
            'budget': float(body.get('budget') or 0),
            'spent': float(body.get('spent') or 0),
            'currency': body.get('currency', 'AED'),
            'start_date': body.get('start_date') or None,
            'end_date': body.get('end_date') or None,
            'target_audience': body.get('target_audience') or None,
            'ad_content': body.get('ad_content', {}),
            'impressions': safe_int(body.get('impressions', 0)),
            'clicks': safe_int(body.get('clicks', 0)),
            'utm_source': body.get('utm_source') or None,
            'utm_medium': body.get('utm_medium') or None,
            'utm_campaign': body.get('utm_campaign') or None,
            'notes': body.get('notes') or None,
            'created_by': request.user_id,
        }
        result = supabase.table('ad_campaigns').insert(row).execute()
        return success_response(result.data[0] if result.data else {}, 'تم إنشاء الإعلان ✅', 201)
    except Exception as e:
        return error_response(f"Ad Create Error: {str(e)}", 500)


@app.route('/api/admin/ads/<ad_id>', methods=['GET'])
@check_role(['admin', 'manager'])
def get_ad_campaign(ad_id):
    """تفاصيل إعلان واحد مع الليدز المرتبطة"""
    try:
        camp = supabase.table('ad_campaigns').select(
            '*, projects(name, slug)'
        ).eq('id', ad_id).single().execute()

        leads = supabase.table('leads').select(
            'id, name, phone, email, status, lead_score, created_at'
        ).eq('ad_campaign_id', ad_id).order('created_at', desc=True).limit(50).execute()

        return success_response({
            'campaign': camp.data,
            'leads': leads.data or [],
        })
    except Exception as e:
        return error_response(f"Ad Detail Error: {str(e)}", 500)


@app.route('/api/admin/ads/<ad_id>', methods=['PATCH'])
@check_role(['admin', 'manager'])
@audit_log('ad_campaign_update')
def update_ad_campaign(ad_id):
    """تعديل بيانات إعلان"""
    try:
        body = request.get_json()
        allowed = [
            'name', 'platform', 'project_id', 'platform_campaign_id',
            'status', 'budget', 'spent', 'currency', 'start_date', 'end_date',
            'target_audience', 'ad_content', 'impressions', 'clicks',
            'utm_source', 'utm_medium', 'utm_campaign', 'notes',
        ]
        updates = {k: body[k] for k in allowed if k in body}
        updates['updated_at'] = datetime.utcnow().isoformat()
        result = supabase.table('ad_campaigns').update(updates).eq('id', ad_id).execute()
        return success_response(result.data[0] if result.data else {}, 'تم التعديل ✅')
    except Exception as e:
        return error_response(f"Ad Update Error: {str(e)}", 500)


@app.route('/api/admin/ads/<ad_id>', methods=['DELETE'])
@check_role(['admin', 'manager'])
@audit_log('ad_campaign_delete')
def delete_ad_campaign(ad_id):
    """حذف إعلان (soft: فك ارتباط الليدز أولاً)"""
    try:
        # فك ارتباط الليدز قبل الحذف
        supabase.table('leads').update(
            {'ad_campaign_id': None}
        ).eq('ad_campaign_id', ad_id).execute()

        supabase.table('ad_campaigns').delete().eq('id', ad_id).execute()
        return success_response(None, 'تم حذف الإعلان')
    except Exception as e:
        return error_response(f"Ad Delete Error: {str(e)}", 500)


@app.route('/api/admin/ads/dashboard', methods=['GET'])
@check_role(['admin', 'manager'])
def ads_dashboard():
    """إحصائيات شاملة لكل الإعلانات"""
    try:
        camps = supabase.table('ad_campaigns').select('*').execute()
        all_camps = camps.data or []

        total_budget = sum(float(c.get('budget') or 0) for c in all_camps)
        total_spent = sum(float(c.get('spent') or 0) for c in all_camps)
        total_impressions = sum(int(c.get('impressions') or 0) for c in all_camps)
        total_clicks = sum(int(c.get('clicks') or 0) for c in all_camps)

        # ليدز من إعلانات
        ad_leads = supabase.table('leads').select(
            'id', count='exact'
        ).not_.is_('ad_campaign_id', 'null').eq('is_active', True).execute()
        total_ad_leads = ad_leads.count or 0

        won_leads = supabase.table('leads').select(
            'id', count='exact'
        ).not_.is_('ad_campaign_id', 'null').eq('status', 'closed_won').execute()
        total_won = won_leads.count or 0

        cpl = round(total_spent / total_ad_leads, 2) if total_ad_leads > 0 else 0
        ctr = round((total_clicks / total_impressions) * 100, 2) if total_impressions > 0 else 0
        conversion_rate = round((total_won / total_ad_leads) * 100, 2) if total_ad_leads > 0 else 0

        # حسب المنصة
        by_platform = {}
        for c in all_camps:
            p = c.get('platform', 'other')
            if p not in by_platform:
                by_platform[p] = {'count': 0, 'spent': 0, 'budget': 0, 'impressions': 0, 'clicks': 0}
            by_platform[p]['count'] += 1
            by_platform[p]['spent'] += float(c.get('spent') or 0)
            by_platform[p]['budget'] += float(c.get('budget') or 0)
            by_platform[p]['impressions'] += int(c.get('impressions') or 0)
            by_platform[p]['clicks'] += int(c.get('clicks') or 0)

        # حسب الحالة
        by_status = {}
        for c in all_camps:
            s = c.get('status', 'active')
            by_status[s] = by_status.get(s, 0) + 1

        return success_response({
            'total_campaigns': len(all_camps),
            'total_budget': total_budget,
            'total_spent': total_spent,
            'total_impressions': total_impressions,
            'total_clicks': total_clicks,
            'total_leads': total_ad_leads,
            'total_won': total_won,
            'cpl': cpl,
            'ctr': ctr,
            'conversion_rate': conversion_rate,
            'by_platform': by_platform,
            'by_status': by_status,
        })
    except Exception as e:
        return error_response(f"Ads Dashboard Error: {str(e)}", 500)


@app.route('/api/admin/ads/<ad_id>/update-stats', methods=['POST'])
@check_role(['admin', 'manager'])
def update_ad_stats(ad_id):
    """تحديث إحصائيات الإعلان يدوياً (impressions, clicks, spent)"""
    try:
        body = request.get_json()
        updates = {}
        if 'impressions' in body:
            updates['impressions'] = int(body['impressions'])
        if 'clicks' in body:
            updates['clicks'] = int(body['clicks'])
        if 'spent' in body:
            updates['spent'] = float(body['spent'])
        if not updates:
            return error_response('لا توجد بيانات للتحديث', 400)
        updates['updated_at'] = datetime.utcnow().isoformat()
        result = supabase.table('ad_campaigns').update(updates).eq('id', ad_id).execute()
        return success_response(result.data[0] if result.data else {}, 'تم تحديث الإحصائيات ✅')
    except Exception as e:
        return error_response(f"Stats Update Error: {str(e)}", 500)


# ==================== AD WEBHOOK — LEAD INGESTION ====================
# Facebook Lead Ads webhook + Generic webhook for any platform
# الليدز بتنزل تلقائي من الإعلانات وتتربط بالحملة

FB_VERIFY_TOKEN = os.getenv('FB_VERIFY_TOKEN', 'alsaeb_crm_verify')
FB_APP_SECRET = os.getenv('FB_APP_SECRET', '')


def _auto_assign_lead(lead_id):
    """Round-robin assign a webhook lead to the next available agent."""
    try:
        agents = supabase.table('employees').select('id').eq(
            'role', 'agent'
        ).eq('is_active', True).order('last_lead_assigned_at', desc=False).limit(1).execute()
        if agents.data:
            agent_id = agents.data[0]['id']
            supabase.table('leads').update({
                'assigned_to': agent_id,
                'status': 'new',
            }).eq('id', lead_id).execute()
            supabase.table('employees').update({
                'last_lead_assigned_at': datetime.utcnow().isoformat(),
            }).eq('id', agent_id).execute()
    except Exception:
        pass  # auto-assign failure shouldn't block lead creation


@app.route('/webhook/facebook', methods=['GET'])
def facebook_webhook_verify():
    """Facebook webhook verification (hub.mode=subscribe)."""
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    if mode == 'subscribe' and token == FB_VERIFY_TOKEN:
        return Response(challenge, status=200, mimetype='text/plain')
    return Response('Forbidden', status=403)


@app.route('/webhook/facebook', methods=['POST'])
def facebook_webhook_receive():
    """استقبال ليدز من Facebook Lead Ads (leadgen events)."""
    try:
        # التحقق من التوقيع (اختياري لو FB_APP_SECRET موجود)
        if FB_APP_SECRET:
            sig_header = request.headers.get('X-Hub-Signature-256', '')
            expected = 'sha256=' + hmac.new(
                FB_APP_SECRET.encode(), request.data, hashlib.sha256
            ).hexdigest()
            if not hmac.compare_digest(sig_header, expected):
                return jsonify({'error': 'Invalid signature'}), 403

        payload = request.get_json(force=True)
        leads_created = 0

        for entry in payload.get('entry', []):
            page_id = str(entry.get('id', ''))
            for change in entry.get('changes', []):
                if change.get('field') != 'leadgen':
                    continue
                value = change.get('value', {})
                leadgen_id = value.get('leadgen_id')
                form_id = str(value.get('form_id', ''))

                # ابحث عن الحملة المرتبطة بهذا الـ Page + Form
                q = supabase.table('ad_campaigns').select('*').eq(
                    'facebook_page_id', page_id
                )
                if form_id:
                    q = q.eq('facebook_form_id', form_id)
                camp_result = q.limit(1).execute()

                # لو مفيش form_id match، جرب page_id بس
                if not camp_result.data and form_id:
                    camp_result = supabase.table('ad_campaigns').select('*').eq(
                        'facebook_page_id', page_id
                    ).is_('facebook_form_id', 'null').limit(1).execute()

                campaign = camp_result.data[0] if camp_result.data else None

                # حاول تجيب بيانات الليد من Graph API
                lead_data = _fetch_facebook_lead(leadgen_id) if leadgen_id else {}

                phone = lead_data.get('phone', '')
                name = lead_data.get('name')
                email = lead_data.get('email')

                if not phone:
                    continue

                phone_clean = re.sub(r'\D', '', phone)

                row = {
                    'phone': phone,
                    'phone_clean': phone_clean,
                    'name': name,
                    'email': email,
                    'source': 'facebook_lead_ad',
                    'status': 'new',
                    'quality': 'warm',
                    'imported_from': f'fb_leadgen_{leadgen_id}',
                }

                if campaign:
                    row['ad_campaign_id'] = campaign['id']
                    row['utm_source'] = campaign.get('utm_source') or 'facebook'
                    row['utm_medium'] = campaign.get('utm_medium') or 'paid'
                    row['utm_campaign'] = campaign.get('utm_campaign') or campaign.get('name')
                    if campaign.get('project_id'):
                        row['project_id'] = campaign['project_id']

                result = supabase.table('leads').insert(row).execute()
                if result.data:
                    leads_created += 1
                    lead_id = result.data[0]['id']

                    # تحديث عداد الويب هوك
                    if campaign:
                        supabase.table('ad_campaigns').update({
                            'webhook_leads_count': (campaign.get('webhook_leads_count') or 0) + 1,
                            'updated_at': datetime.utcnow().isoformat(),
                        }).eq('id', campaign['id']).execute()

                    # توزيع تلقائي
                    if campaign and campaign.get('auto_assign'):
                        _auto_assign_lead(lead_id)

                    # webhook outbound
                    fire_webhooks('lead.created', result.data[0])

        return jsonify({'status': 'ok', 'leads_created': leads_created}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _fetch_facebook_lead(leadgen_id):
    """جلب بيانات الليد من Facebook Graph API باستخدام leadgen_id."""
    token = os.getenv('WHATSAPP_TOKEN') or os.getenv('FB_PAGE_ACCESS_TOKEN', '')
    if not token or not leadgen_id:
        return {}
    try:
        import requests as req_lib
        url = f"https://graph.facebook.com/v18.0/{leadgen_id}"
        resp = req_lib.get(url, params={'access_token': token}, timeout=10)
        if resp.status_code != 200:
            return {}
        data = resp.json()
        fields = {}
        for fd in data.get('field_data', []):
            name_key = fd.get('name', '').lower()
            val = fd.get('values', [''])[0] if fd.get('values') else ''
            if 'phone' in name_key or 'mobile' in name_key:
                fields['phone'] = val
            elif 'email' in name_key:
                fields['email'] = val
            elif 'name' in name_key:
                if 'full' in name_key or name_key == 'name':
                    fields['name'] = val
                elif 'first' in name_key:
                    fields['first_name'] = val
                elif 'last' in name_key:
                    fields['last_name'] = val
        # لو فيه first + last بس مفيش full name
        if not fields.get('name') and (fields.get('first_name') or fields.get('last_name')):
            fields['name'] = f"{fields.get('first_name', '')} {fields.get('last_name', '')}".strip()
        return fields
    except Exception:
        return {}


@app.route('/webhook/lead/<campaign_id>', methods=['POST'])
def generic_webhook_lead(campaign_id):
    """
    Generic webhook — أي منصة تقدر تبعت ليدز.
    Auth: ?key=WEBHOOK_KEY (من ad_campaigns.webhook_key)

    POST body JSON:
    {
      "phone": "05XXXXXXXX",           (مطلوب)
      "name": "Ahmed",                  (اختياري)
      "email": "a@b.com",               (اختياري)
      "source": "google_ads",           (اختياري — default: webhook)
      "quality": "warm",                (اختياري — default: unknown)
      "notes": "Interested in...",       (اختياري)
      "project_id": "uuid"              (اختياري)
    }
    """
    try:
        # Auth by webhook key
        key = request.args.get('key', '')
        if not key:
            return jsonify({'error': 'Missing key parameter'}), 401

        camp = supabase.table('ad_campaigns').select('*').eq(
            'id', campaign_id
        ).eq('webhook_key', key).limit(1).execute()

        if not camp.data:
            return jsonify({'error': 'Invalid campaign ID or key'}), 403

        campaign = camp.data[0]
        body = request.get_json(force=True) or {}

        # قبول ليد واحد أو مصفوفة ليدز
        leads_list = body if isinstance(body, list) else [body]
        created = []

        for item in leads_list:
            phone = (item.get('phone') or '').strip()
            if not phone:
                continue

            phone_clean = re.sub(r'\D', '', phone)

            row = {
                'phone': phone,
                'phone_clean': phone_clean,
                'name': item.get('name') or None,
                'email': item.get('email') or None,
                'source': item.get('source') or 'webhook',
                'quality': item.get('quality') or 'unknown',
                'status': 'new',
                'ad_campaign_id': campaign['id'],
                'utm_source': campaign.get('utm_source') or item.get('utm_source'),
                'utm_medium': campaign.get('utm_medium') or item.get('utm_medium'),
                'utm_campaign': campaign.get('utm_campaign') or item.get('utm_campaign'),
                'imported_from': f"webhook_{campaign['name']}",
            }
            if item.get('project_id'):
                row['project_id'] = item['project_id']
            elif campaign.get('project_id'):
                row['project_id'] = campaign['project_id']

            result = supabase.table('leads').insert(row).execute()
            if result.data:
                created.append(result.data[0])

                if campaign.get('auto_assign'):
                    _auto_assign_lead(result.data[0]['id'])

                fire_webhooks('lead.created', result.data[0])

        # تحديث عداد الويب هوك
        if created:
            supabase.table('ad_campaigns').update({
                'webhook_leads_count': (campaign.get('webhook_leads_count') or 0) + len(created),
                'updated_at': datetime.utcnow().isoformat(),
            }).eq('id', campaign['id']).execute()

        return jsonify({
            'status': 'ok',
            'leads_created': len(created),
            'leads': [{'id': ld['id'], 'phone': ld['phone']} for ld in created],
        }), 201 if created else 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/ads/<ad_id>/webhook', methods=['POST'])
@check_role(['admin', 'manager'])
@audit_log('webhook_key_generate')
def generate_webhook_key(ad_id):
    """توليد أو تجديد مفتاح الويب هوك لحملة إعلانية."""
    try:
        import secrets
        new_key = secrets.token_urlsafe(32)
        result = supabase.table('ad_campaigns').update({
            'webhook_key': new_key,
            'updated_at': datetime.utcnow().isoformat(),
        }).eq('id', ad_id).execute()
        if not result.data:
            return error_response('الحملة غير موجودة', 404)
        base_url = request.host_url.rstrip('/')
        return success_response({
            'webhook_key': new_key,
            'webhook_url': f"{base_url}/webhook/lead/{ad_id}?key={new_key}",
            'facebook_webhook_url': f"{base_url}/webhook/facebook",
        }, 'تم توليد مفتاح الويب هوك ✅')
    except Exception as e:
        return error_response(f"Webhook Key Error: {str(e)}", 500)


@app.route('/api/admin/ads/<ad_id>/webhook', methods=['GET'])
@check_role(['admin', 'manager'])
def get_webhook_info(ad_id):
    """عرض معلومات الويب هوك لحملة إعلانية."""
    try:
        camp = supabase.table('ad_campaigns').select(
            'id,name,webhook_key,facebook_page_id,facebook_form_id,auto_assign,webhook_leads_count'
        ).eq('id', ad_id).single().execute()
        if not camp.data:
            return error_response('الحملة غير موجودة', 404)
        base_url = request.host_url.rstrip('/')
        data = camp.data
        data['webhook_url'] = f"{base_url}/webhook/lead/{ad_id}?key={data.get('webhook_key', 'GENERATE_KEY_FIRST')}"
        data['facebook_webhook_url'] = f"{base_url}/webhook/facebook"
        return success_response(data)
    except Exception as e:
        return error_response(f"Webhook Info Error: {str(e)}", 500)


@app.route('/api/admin/ads/<ad_id>/webhook', methods=['PATCH'])
@check_role(['admin', 'manager'])
@audit_log('webhook_settings_update')
def update_webhook_settings(ad_id):
    """تحديث إعدادات الويب هوك (Facebook IDs, auto_assign)."""
    try:
        body = request.get_json() or {}
        updates = {}
        if 'facebook_page_id' in body:
            updates['facebook_page_id'] = body['facebook_page_id'] or None
        if 'facebook_form_id' in body:
            updates['facebook_form_id'] = body['facebook_form_id'] or None
        if 'auto_assign' in body:
            updates['auto_assign'] = bool(body['auto_assign'])
        if not updates:
            return error_response('لا توجد بيانات للتحديث', 400)
        updates['updated_at'] = datetime.utcnow().isoformat()
        result = supabase.table('ad_campaigns').update(updates).eq('id', ad_id).execute()
        if not result.data:
            return error_response('الحملة غير موجودة', 404)
        return success_response(result.data[0], 'تم تحديث إعدادات الويب هوك ✅')
    except Exception as e:
        return error_response(f"Webhook Settings Error: {str(e)}", 500)


# ==================== HEALTH CHECK ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'ok',
        'service': 'AlSaeb CRM',
        'version': '2.0.0',
        'timestamp': datetime.utcnow().isoformat()
    })


# ==================== RUN ====================

if __name__ == '__main__':
    app.run(debug=True, port=5000)
